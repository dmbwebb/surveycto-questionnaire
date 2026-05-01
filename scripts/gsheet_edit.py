"""Edit helpers for XLSForm Google Sheets — gsheet IS the source of truth.

These wrap the Sheets API around XLSForm-shaped intent: "find the row whose
'name' column equals X and set its 'label' to Y", "rename a variable and
update every ${var} reference that points at it", "insert a question inside
a begin/end group block at row N", "add a choice list to the choices tab and
a select_one row in the survey tab referencing it".

All public functions take ``doc_id`` (Drive file id) and a tab title (typically
'survey', 'choices', or 'settings'). Row indices are 1-based to match the
Sheets UI. Header row is row 1; data starts at row 2.
"""
from __future__ import annotations

import re
import time
from dataclasses import dataclass
from typing import Iterable, Mapping, Sequence

from googleapiclient.errors import HttpError
from openpyxl.utils import get_column_letter

from gsheet_io import sheets_service


def _execute_with_retry(request, *, max_attempts: int = 6, base_delay: float = 5.0):
    """Execute a Google API request with exponential backoff on HTTP 429.

    Sheets quotas a user to 60 reads/min and 60 writes/min. When a concurrent
    caller (another agent, the user's browser) is also burning calls, a single
    request can transiently fail with 429. This wrapper retries on 429 only;
    any other HttpError or non-HTTP exception re-raises immediately.

    Backoff: 5s, 10s, 20s, 40s, 80s, 160s. Max ~5min total before giving up.
    """
    last_err = None
    for attempt in range(max_attempts):
        try:
            return request.execute()
        except HttpError as e:
            status = getattr(getattr(e, "resp", None), "status", None)
            if status != 429:
                raise
            last_err = e
            if attempt < max_attempts - 1:
                time.sleep(base_delay * (2 ** attempt))
    raise last_err

# updates.updatedRange comes back like "survey!A11:D11" — capture the END row.
# Must use [A-Z]+ (not \w+) — \w+ would greedy-eat digits and backtrack wrong.
_APPENDED_RANGE_RE = re.compile(r"![A-Z]+\d+:[A-Z]+(\d+)$")


def _parse_last_appended_row(append_response: dict) -> int:
    updated_range = append_response.get("updates", {}).get("updatedRange", "")
    m = _APPENDED_RANGE_RE.search(updated_range)
    if not m:
        raise RuntimeError(
            f"Could not parse appended row from updatedRange={updated_range!r}; "
            f"full response: {append_response!r}"
        )
    return int(m.group(1))


class StaleDataError(RuntimeError):
    """Raised when a compare-and-swap write detects out-of-band modification.

    Carries enough context for the caller to retry intelligently:
      .row, .header, .expected, .actual
    """
    def __init__(self, row: int, header: str, expected, actual):
        super().__init__(
            f"stale-data: cell ({row}, {header!r}) expected={expected!r} "
            f"but found={actual!r}; another writer changed it since you read it"
        )
        self.row = row
        self.header = header
        self.expected = expected
        self.actual = actual


@dataclass(frozen=True)
class TabHandle:
    """Cached header layout for one tab, so repeated edits don't re-fetch headers."""
    doc_id: str
    tab_title: str
    sheet_id: int
    headers: tuple[str, ...]   # column titles in order; index 0 = column A

    def col_idx_0(self, header_name: str) -> int:
        try:
            return self.headers.index(header_name)
        except ValueError as e:
            raise ValueError(
                f"Tab '{self.tab_title}' has no column named {header_name!r}. "
                f"Headers: {list(self.headers)}"
            ) from e

    def col_letter(self, header_name: str) -> str:
        return _col_idx_to_letter(self.col_idx_0(header_name))


def _col_idx_to_letter(idx0: int) -> str:
    """0-based column index -> A1 letters (0->A, 25->Z, 26->AA).

    Thin wrapper over openpyxl's 1-based ``get_column_letter`` so call sites
    can stay 0-based (consistent with the Sheets API request schema).
    """
    if idx0 < 0:
        raise ValueError(f"negative column index: {idx0}")
    return get_column_letter(idx0 + 1)


def open_tab(doc_id: str, tab_title: str) -> TabHandle:
    """Fetch a tab's sheetId + header row in a single API round-trip."""
    res = sheets_service().spreadsheets().get(
        spreadsheetId=doc_id,
        ranges=[f"'{tab_title}'!1:1"],
        includeGridData=True,
        fields=("sheets.properties(sheetId,title),"
                "sheets.data.rowData.values.formattedValue"),
    ).execute()

    for s in res.get("sheets", []):
        props = s["properties"]
        if props["title"] != tab_title:
            continue
        sheet_id = props["sheetId"]
        rows = (s.get("data", [{}])[0] or {}).get("rowData", [])
        first_row_cells = rows[0].get("values", []) if rows else []
        headers = tuple(c.get("formattedValue", "") for c in first_row_cells)
        return TabHandle(doc_id=doc_id, tab_title=tab_title,
                         sheet_id=sheet_id, headers=headers)
    raise ValueError(f"No tab named {tab_title!r} in spreadsheet {doc_id}")


# ---------- Cell-level reads/writes ----------

def find_row_by_value(tab: TabHandle, header_name: str, value: str,
                      *, start_row: int = 2) -> int | None:
    """Return 1-based row of the first row where ``header_name`` == value,
    or None if not found."""
    svc = sheets_service()
    col_letter = tab.col_letter(header_name)
    res = svc.spreadsheets().values().get(
        spreadsheetId=tab.doc_id,
        range=f"'{tab.tab_title}'!{col_letter}{start_row}:{col_letter}",
    ).execute()
    col_values = [r[0] if r else "" for r in res.get("values", [])]
    target = str(value)
    for i, v in enumerate(col_values):
        if str(v) == target:
            return start_row + i
    return None


def get_cell(tab: TabHandle, row: int, header_name: str) -> str | None:
    """Read the value of the cell at ``(row, header_name)``."""
    svc = sheets_service()
    col_letter = tab.col_letter(header_name)
    res = svc.spreadsheets().values().get(
        spreadsheetId=tab.doc_id,
        range=f"'{tab.tab_title}'!{col_letter}{row}",
    ).execute()
    vals = res.get("values", [])
    if not vals or not vals[0]:
        return None
    return vals[0][0]


def get_row(tab: TabHandle, row: int) -> dict[str, object]:
    """Read a whole data row as a dict keyed by header name."""
    svc = sheets_service()
    res = svc.spreadsheets().values().get(
        spreadsheetId=tab.doc_id,
        range=f"'{tab.tab_title}'!{row}:{row}",
    ).execute()
    vals = res.get("values", [[]])[0]
    return {h: (vals[i] if i < len(vals) else "") for i, h in enumerate(tab.headers)}


def update_cell(tab: TabHandle, row: int, header_name: str, new_value) -> None:
    """Write a single cell. USER_ENTERED so formulas / numbers are parsed."""
    svc = sheets_service()
    col_letter = tab.col_letter(header_name)
    svc.spreadsheets().values().update(
        spreadsheetId=tab.doc_id,
        range=f"'{tab.tab_title}'!{col_letter}{row}",
        valueInputOption="USER_ENTERED",
        body={"values": [[new_value]]},
    ).execute()


def update_cell_checked(tab: TabHandle, row: int, header_name: str,
                        expected_old, new_value) -> None:
    """Compare-and-swap: only write if the current value matches ``expected_old``.

    Catches stale-data scenarios where another writer modified the cell between
    your read and your intended write. Raises ``StaleDataError`` on mismatch.

    Note: this is racy without locking. It catches *some* concurrent edits, not
    all — use as a best-effort guard, not a transaction.
    """
    actual = get_cell(tab, row, header_name)
    # Normalize None vs "" mismatch — Sheets returns None for empty cells but
    # callers often pass "".
    actual_norm = "" if actual is None else str(actual)
    expected_norm = "" if expected_old is None else str(expected_old)
    if actual_norm != expected_norm:
        raise StaleDataError(row, header_name, expected_old, actual)
    update_cell(tab, row, header_name, new_value)


def batch_update_cells(tab: TabHandle,
                       edits: Iterable[tuple[int, str, object]]) -> dict:
    """Apply many cell updates in a single Sheets API call.

    edits: iterable of ``(row, header_name, new_value)`` tuples.

    Trade-off vs ``update_cell_checked``: 1 API call (vs ~2N) and no
    compare-and-swap. Caller is responsible for verifying preconditions ahead
    of time — typically via a single bulk read through ``gsheet_io.exported_xlsx``.

    Use this when writing 10+ cells at once would otherwise risk HTTP 429
    (Sheets quota: 60 reads/min, 60 writes/min). Retries on 429 internally.

    An empty ``edits`` iterable is a no-op — returns a synthetic
    ``{"totalUpdatedCells": 0, "totalUpdatedRows": 0}`` without an API call.

    Returns the API response dict (``totalUpdatedCells``, ``totalUpdatedRows``, …).
    """
    edits_list = list(edits)
    if not edits_list:
        return {"totalUpdatedCells": 0, "totalUpdatedRows": 0}
    data = []
    for row, header_name, new_value in edits_list:
        col_letter = tab.col_letter(header_name)
        data.append({
            "range": f"'{tab.tab_title}'!{col_letter}{row}",
            "values": [[new_value]],
        })
    svc = sheets_service()
    return _execute_with_retry(
        svc.spreadsheets().values().batchUpdate(
            spreadsheetId=tab.doc_id,
            body={"valueInputOption": "USER_ENTERED", "data": data},
        )
    )


def bulk_set_column(tab: TabHandle, rows: Iterable[int],
                    header_name: str, value) -> dict:
    """Set the same value on the given rows in one column. One API call.

    Convenience wrapper over ``batch_update_cells`` for the common pattern of
    flipping a column (e.g. ``disabled = 'yes'`` across many rows of a module,
    or ``required = 'yes'`` across all "Other, specify" follow-ups).
    """
    return batch_update_cells(tab, [(r, header_name, value) for r in rows])


def append_row(tab: TabHandle, row_dict: Mapping[str, object]) -> int:
    """Append one row at the bottom of the tab. Returns the 1-based landed row."""
    svc = sheets_service()
    values = [row_dict.get(h, "") for h in tab.headers]
    res = svc.spreadsheets().values().append(
        spreadsheetId=tab.doc_id,
        range=f"'{tab.tab_title}'!A1",
        valueInputOption="USER_ENTERED",
        insertDataOption="INSERT_ROWS",
        body={"values": [values]},
    ).execute()
    return _parse_last_appended_row(res)


def insert_row_at(tab: TabHandle, position_row: int,
                  row_dict: Mapping[str, object]) -> None:
    """Insert a new row at 1-based ``position_row``, shifting everything below down.

    Two API calls: insertDimension (shifts rows) + values.update (fills new row).
    Can't fold into one batchUpdate without losing USER_ENTERED parsing on the
    new row's values — important for the rare ``=NOW()``-style formula cell.

    Required for inserting a question inside a `begin group`/`end group` block.
    """
    if position_row < 1:
        raise ValueError(f"position_row must be >= 1, got {position_row}")

    svc = sheets_service()
    svc.spreadsheets().batchUpdate(
        spreadsheetId=tab.doc_id,
        body={
            "requests": [{
                "insertDimension": {
                    "range": {
                        "sheetId": tab.sheet_id,
                        "dimension": "ROWS",
                        "startIndex": position_row - 1,
                        "endIndex": position_row,
                    },
                    "inheritFromBefore": True,
                },
            }],
        },
    ).execute()

    values = [row_dict.get(h, "") for h in tab.headers]
    last_col_letter = _col_idx_to_letter(len(tab.headers) - 1)
    svc.spreadsheets().values().update(
        spreadsheetId=tab.doc_id,
        range=f"'{tab.tab_title}'!A{position_row}:{last_col_letter}{position_row}",
        valueInputOption="USER_ENTERED",
        body={"values": [values]},
    ).execute()


def delete_row(tab: TabHandle, row: int) -> None:
    """Delete a single row (1-based). Rows below shift up."""
    if row < 2:
        raise ValueError(f"row must be >= 2 (header is row 1), got {row}")
    svc = sheets_service()
    svc.spreadsheets().batchUpdate(
        spreadsheetId=tab.doc_id,
        body={
            "requests": [{
                "deleteDimension": {
                    "range": {
                        "sheetId": tab.sheet_id,
                        "dimension": "ROWS",
                        "startIndex": row - 1,  # 0-based, inclusive
                        "endIndex": row,        # 0-based, exclusive
                    },
                },
            }],
        },
    ).execute()


# ---------- Variable reference rewriting ----------

# Columns in the survey tab where ${field_name} references can appear.
REFERENCE_COLUMNS = (
    "relevance", "constraint", "calculation", "label", "hint",
    "choice_filter", "repeat_count", "required", "default",
)
# Multilingual variants: label::English, label:Hindi, hint:Malagasy, etc.
_LABEL_HINT_VARIANT = re.compile(r"^(label|hint)(::|:).+$")


def _all_reference_columns(headers: Iterable[str]) -> list[str]:
    out = []
    for h in headers:
        if h in REFERENCE_COLUMNS or _LABEL_HINT_VARIANT.match(h or ""):
            out.append(h)
    return out


def rename_variable(
    tab: TabHandle,
    old_name: str,
    new_name: str,
    *,
    name_column: str = "name",
) -> dict:
    """Rename a variable in the ``name`` column AND update every ${old_name}
    reference across all known reference columns.

    Single round-trip: pull the whole tab once, plan, write everything in one
    batchUpdate. Returns ``{'name_renamed': N, 'references_updated': M}``.
    """
    svc = sheets_service()

    res = svc.spreadsheets().values().get(
        spreadsheetId=tab.doc_id,
        range=f"'{tab.tab_title}'",
    ).execute()
    grid = res.get("values", [])
    if not grid:
        raise RuntimeError(f"Tab '{tab.tab_title}' is empty")

    fetched_headers = grid[0]

    def col_idx(name: str) -> int:
        try:
            return fetched_headers.index(name)
        except ValueError as e:
            raise ValueError(
                f"Tab '{tab.tab_title}' missing column {name!r} (fetched headers: "
                f"{fetched_headers})"
            ) from e

    name_idx = col_idx(name_column)
    ref_col_indexes = [col_idx(h) for h in _all_reference_columns(fetched_headers)]

    old_ref = "${" + old_name + "}"
    new_ref = "${" + new_name + "}"
    name_renamed = 0
    refs_updated = 0
    updates_a1: list[dict] = []

    for r_idx, row in enumerate(grid[1:], start=2):
        if len(row) < len(fetched_headers):
            row = row + [""] * (len(fetched_headers) - len(row))

        if row[name_idx] == old_name:
            col_letter = _col_idx_to_letter(name_idx)
            updates_a1.append({
                "range": f"'{tab.tab_title}'!{col_letter}{r_idx}",
                "values": [[new_name]],
            })
            name_renamed += 1

        for c_idx in ref_col_indexes:
            cell = row[c_idx]
            if not cell or old_ref not in str(cell):
                continue
            new_cell = str(cell).replace(old_ref, new_ref)
            col_letter = _col_idx_to_letter(c_idx)
            updates_a1.append({
                "range": f"'{tab.tab_title}'!{col_letter}{r_idx}",
                "values": [[new_cell]],
            })
            refs_updated += 1

    if updates_a1:
        svc.spreadsheets().values().batchUpdate(
            spreadsheetId=tab.doc_id,
            body={
                "valueInputOption": "USER_ENTERED",
                "data": updates_a1,
            },
        ).execute()

    return {"name_renamed": name_renamed, "references_updated": refs_updated}


# ---------- Multi-tab convenience: choice list ----------

def add_choice_list(
    doc_id: str,
    list_name: str,
    choices: Sequence[tuple[str, str]] | Sequence[Mapping[str, object]],
    *,
    extra_cols: Mapping[str, str] | None = None,
) -> int:
    """Append a choice list to the 'choices' tab.

    ``choices`` can be a list of ``(id, label)`` tuples OR a list of dicts.
    The XLSForm spec accepts either ``name`` or ``value`` as the choice id
    column — this function detects which the tab uses and writes there.
    Tuple form: first element goes into the id column, second into ``label``.
    Dict form: passed through; aliases ``name``/``value`` to whichever the
    tab has.

    ``extra_cols`` sets a default for some column on every appended row.

    Returns the 1-based row of the LAST appended choice.
    """
    tab = open_tab(doc_id, "choices")

    if "list_name" not in tab.headers or "label" not in tab.headers:
        raise ValueError(
            f"choices tab missing required columns 'list_name' and/or 'label'; "
            f"got headers: {list(tab.headers)}"
        )
    # XLSForm allows either 'name' or 'value' as the choice id column.
    if "name" in tab.headers:
        id_col = "name"
    elif "value" in tab.headers:
        id_col = "value"
    else:
        raise ValueError(
            f"choices tab has neither 'name' nor 'value' column; "
            f"got headers: {list(tab.headers)}"
        )

    extra_cols = dict(extra_cols or {})
    rows_to_append = []
    for c in choices:
        if isinstance(c, tuple):
            cid, label = c
            row = {"list_name": list_name, id_col: cid, "label": label}
        else:
            row = dict(c)
            row.setdefault("list_name", list_name)
            # Alias whichever id key the caller used to whichever the tab needs.
            if id_col == "value" and "name" in row and "value" not in row:
                row["value"] = row.pop("name")
            elif id_col == "name" and "value" in row and "name" not in row:
                row["name"] = row.pop("value")
        for k, v in extra_cols.items():
            row.setdefault(k, v)
        rows_to_append.append([row.get(h, "") for h in tab.headers])

    if not rows_to_append:
        raise ValueError("no choices to append")

    svc = sheets_service()
    res = svc.spreadsheets().values().append(
        spreadsheetId=doc_id,
        range="'choices'!A1",
        valueInputOption="USER_ENTERED",
        insertDataOption="INSERT_ROWS",
        body={"values": rows_to_append},
    ).execute()
    return _parse_last_appended_row(res)


# ---------- Formatting (textFormat.foregroundColor) ----------

def set_text_color(
    tab: TabHandle,
    row: int,
    header_name: str,
    color_rgb: tuple[float, float, float],
) -> None:
    """Set foreground (text) color for a single cell. Components are 0..1 floats."""
    r, g, b = color_rgb
    if not all(0 <= c <= 1 for c in (r, g, b)):
        raise ValueError(f"color components must be 0..1; got {color_rgb}")

    svc = sheets_service()
    col_idx0 = tab.col_idx_0(header_name)
    svc.spreadsheets().batchUpdate(
        spreadsheetId=tab.doc_id,
        body={
            "requests": [{
                "repeatCell": {
                    "range": {
                        "sheetId": tab.sheet_id,
                        "startRowIndex": row - 1,
                        "endRowIndex": row,
                        "startColumnIndex": col_idx0,
                        "endColumnIndex": col_idx0 + 1,
                    },
                    "cell": {
                        "userEnteredFormat": {
                            "textFormat": {
                                "foregroundColor": {
                                    "red": r, "green": g, "blue": b,
                                },
                            },
                        },
                    },
                    "fields": "userEnteredFormat.textFormat.foregroundColor",
                },
            }],
        },
    ).execute()


def get_text_color(tab: TabHandle, row: int, header_name: str
                   ) -> tuple[float, float, float] | None:
    """Read foreground color of a cell. Returns (r, g, b) floats 0..1, or None."""
    svc = sheets_service()
    col_idx0 = tab.col_idx_0(header_name)
    col_letter = _col_idx_to_letter(col_idx0)
    res = svc.spreadsheets().get(
        spreadsheetId=tab.doc_id,
        ranges=[f"'{tab.tab_title}'!{col_letter}{row}"],
        fields="sheets.data.rowData.values.effectiveFormat.textFormat.foregroundColor",
    ).execute()
    try:
        cell = res["sheets"][0]["data"][0]["rowData"][0]["values"][0]
        c = cell["effectiveFormat"]["textFormat"]["foregroundColor"]
    except (KeyError, IndexError):
        return None
    return (c.get("red", 0.0), c.get("green", 0.0), c.get("blue", 0.0))

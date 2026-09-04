#!/usr/bin/env python3
"""
SurveyCTO Form Checker

Validates XLSForm files for common errors:
- Expression syntax errors (unbalanced parentheses, unclosed ${} references, unclosed quotes)
- References to non-existent fields in relevance, choice_filter, calculation, and constraint expressions
- Multi-field XPath dependency cycles across relevance, calculation, and required expressions
- Undefined choice lists
- Markup in dynamic search() choice rows (silently blanks the picker in that language)
- Plain calculate over the `end` metadata field (never populates; use calculate_here)
- Missing or malformed survey-level and section timing instrumentation
- Audio audit anchors that name a missing, invisible or in-repeat field, or use ${} syntax
- Audio audits placed inside a repeat group
- Single whole-survey audio audits, which record nothing after Edit Saved Form
- Implausibly sparse section timing coverage for the questionnaire's size
- Missing required columns
- Typos in field names and labels
- Missing constraint messages
- Integer fields without constraints
- Numeric fields (integer/decimal) without -999 refuse option
- Calculate fields without calculation formulas
- Missing Hindi translations
- Naming convention issues
- select_multiple questions with 'other' option but no specify field
- select_multiple questions with exclusive options (don't know, refuse, nothing, etc.) missing constraints
- Impossible literal values: ${var}=X or selected(${var}, X) where X isn't in var's choice list
- Conditional formatting rules (type-based color coding) are preserved
- Cell formatting (red text for unverified translations) is preserved
- Encryption public_key is a base64 DER payload, not a PEM-wrapped key
- Version formula in settings sheet is evaluated

Usage:
    python surveycto_checker.py <path_to_xlsform.xlsx> [--platform surveycto|odk|kobo]
    python surveycto_checker.py  # checks ai_health_pilot_baseline.xlsx by default
"""

import argparse
import base64
import binascii
import re
import sys
import subprocess
from pathlib import Path

import openpyxl
import pandas as pd


class SurveyCTOChecker:
    """Validates SurveyCTO XLSForm files for common errors."""

    EXPRESSION_COLUMNS = ('relevance', 'calculation', 'constraint', 'choice_filter',
                          'repeat_count', 'default')
    DEPENDENCY_COLUMNS = ('relevance', 'calculation', 'required')
    SELF_REFERENCE_COLUMNS = DEPENDENCY_COLUMNS
    # XPath string functions the ODK docs list but SurveyCTO's JavaRosa parser rejects.
    UNSUPPORTED_ODK_FUNCTIONS = ('starts-with', 'contains',
                                 'substring-before', 'substring-after')
    UPLOAD_SENSITIVE_COLUMNS = (
        'type', 'name', 'relevance', 'calculation', 'constraint',
        'constraint_message', 'required', 'required message',
        'read only', 'repeat_count', 'choice_filter', 'default'
    )
    # Field types SurveyCTO puts on screen. Audio-audit start/end anchors must
    # name one of these: calculate, calculate_here, metadata rows and group
    # boundaries are not visible fields and silently break the audit.
    VISIBLE_FIELD_TYPES = (
        'text', 'integer', 'decimal', 'date', 'time', 'datetime', 'geopoint',
        'geotrace', 'geoshape', 'barcode', 'image', 'audio', 'video', 'file',
        'range', 'rank', 'note', 'comments', 'trigger', 'acknowledge', 'email',
        'enumerator', 'select_one', 'select_multiple',
        'select_one_from_file', 'select_multiple_from_file',
    )
    # Multi-word invisible types whose first token collides with a visible type
    # ('audio audit' would otherwise read as 'audio').
    INVISIBLE_MULTIWORD_TYPES = (
        'audio audit', 'text audit', 'speed violations audit',
        'speed violations count',
    )
    # Above this share of the form's visible fields, a single audio audit is
    # effectively a whole-survey recording.
    AUDIO_AUDIT_MAX_SPAN_SHARE = 0.5
    # Above this many seconds, a single time-based audit is effectively a
    # whole-survey recording.
    AUDIO_AUDIT_MAX_DURATION_SEC = 900

    def __init__(self, file_path, platform='surveycto'):
        self.file_path = Path(file_path)
        self.platform = platform.lower()
        if self.platform not in {'surveycto', 'odk', 'kobo'}:
            raise ValueError("platform must be one of: surveycto, odk, kobo")
        self.survey_df = None
        self.choices_df = None
        self.settings_df = None
        self.errors = []
        self.warnings = []

    def load_form(self):
        """Load the XLSForm file."""
        try:
            self.survey_df = pd.read_excel(self.file_path, sheet_name='survey')
            self.choices_df = pd.read_excel(self.file_path, sheet_name='choices')
            try:
                self.settings_df = pd.read_excel(self.file_path, sheet_name='settings')
            except ValueError:
                self.warnings.append("No 'settings' sheet found (optional)")

            # SurveyCTO uses 'value' for choice identifiers; XLSForm standard uses 'name'.
            # Accept either by renaming 'value' -> 'name' when 'name' is absent.
            if 'name' not in self.choices_df.columns and 'value' in self.choices_df.columns:
                self.choices_df = self.choices_df.rename(columns={'value': 'name'})

            # XLSForm spec is 'constraint message' (space); some teams use 'constraint_message'
            # (underscore). Internally the checker uses the underscore form, so alias the spec
            # name to it when the underscore form isn't present.
            if ('constraint_message' not in self.survey_df.columns
                    and 'constraint message' in self.survey_df.columns):
                self.survey_df = self.survey_df.rename(
                    columns={'constraint message': 'constraint_message'})

            # Filter out disabled rows from survey (preserve original indices for row reporting)
            if 'disabled' in self.survey_df.columns:
                disabled_count = (self.survey_df['disabled'].astype(str).str.lower() == 'yes').sum()
                self.survey_df = self.survey_df[
                    self.survey_df['disabled'].astype(str).str.lower() != 'yes'
                ]
                if disabled_count > 0:
                    print(f"Filtered out {disabled_count} disabled row(s)")

            return True
        except Exception as e:
            self.errors.append(f"Failed to load file: {e}")
            return False

    def check_field_references(self):
        """Check for references to non-existent fields."""
        print("\n=== Checking Field References ===")

        # Get all existing field names
        existing_fields = set(self.survey_df['name'].dropna().astype(str))
        print(f"Found {len(existing_fields)} defined fields")

        # Columns that may contain field references
        # Note: label and hint columns can also contain ${field} references for piping
        reference_columns = ['relevance', 'choice_filter', 'calculation', 'constraint',
                            'constraint_message', 'repeat_count', 'default',
                            'label', 'label:Hindi', 'hint', 'hint:Hindi']

        issues = []

        for idx, row in self.survey_df.iterrows():
            field_name = row.get('name', f'Row {idx}')

            for col in reference_columns:
                if pd.notna(row.get(col)):
                    expression = str(row[col])

                    # Find all ${field_name} references
                    references = re.findall(r'\$\{([^}]+)\}', expression)

                    for ref in references:
                        # Clean up the reference (remove function calls, indexing, etc.)
                        base_ref = re.match(r'^([a-zA-Z_][a-zA-Z0-9_]*)', ref)
                        if base_ref:
                            base_field = base_ref.group(1)
                            if base_field not in existing_fields:
                                issues.append({
                                    'row': idx + 2,  # +2 for Excel row (1-indexed + header)
                                    'field': field_name,
                                    'column': col,
                                    'missing_ref': base_field,
                                    'expression': expression
                                })

        if issues:
            print(f"\n❌ Found {len(issues)} reference(s) to non-existent fields:\n")
            for issue in issues:
                error_msg = (f"  Row {issue['row']}: '{issue['field']}' in column '{issue['column']}'\n"
                           f"    References non-existent field: ${{{issue['missing_ref']}}}\n"
                           f"    Expression: {issue['expression']}\n")
                print(error_msg)
                self.errors.append(error_msg)
        else:
            print("✅ All field references are valid")

        return len(issues) == 0

    def check_choices_field_references(self):
        """Check for references to non-existent fields in choices sheet labels.

        Choice labels can contain ${field_name} references (e.g., ${adult_label_1})
        that are piped from survey fields. This checks that all such references
        point to fields that exist in the survey.
        """
        print("\n=== Checking Choices Sheet Field References ===")

        # Get all existing field names from survey
        existing_fields = set(self.survey_df['name'].dropna().astype(str))

        issues = []

        # Check label columns in choices sheet
        label_columns = [col for col in self.choices_df.columns if col.startswith('label')]

        for idx, row in self.choices_df.iterrows():
            list_name = row.get('list_name', '')
            choice_name = row.get('name', '')

            for col in label_columns:
                if pd.notna(row.get(col)):
                    label = str(row[col])

                    # Find all ${field_name} references
                    references = re.findall(r'\$\{([^}]+)\}', label)

                    for ref in references:
                        # Clean up the reference (remove function calls, indexing, etc.)
                        base_ref = re.match(r'^([a-zA-Z_][a-zA-Z0-9_]*)', ref)
                        if base_ref:
                            base_field = base_ref.group(1)
                            if base_field not in existing_fields:
                                issues.append({
                                    'row': idx + 2,  # +2 for Excel row (1-indexed + header)
                                    'list_name': list_name,
                                    'choice_name': choice_name,
                                    'column': col,
                                    'missing_ref': base_field,
                                    'label': label[:50]
                                })

        if issues:
            # Group by missing field for cleaner output
            missing_fields = set(issue['missing_ref'] for issue in issues)
            print(f"\n❌ Found {len(issues)} reference(s) to {len(missing_fields)} non-existent field(s) in choices:\n")

            for field in sorted(missing_fields):
                field_issues = [i for i in issues if i['missing_ref'] == field]
                lists_affected = set(i['list_name'] for i in field_issues)
                error_msg = (f"  Missing field: ${{{field}}}\n"
                           f"    Used in choice lists: {', '.join(sorted(lists_affected))}\n"
                           f"    Rows: {', '.join(str(i['row']) for i in field_issues[:5])}"
                           f"{'...' if len(field_issues) > 5 else ''}\n")
                print(error_msg)
                self.errors.append(f"Choices sheet references non-existent field: ${{{field}}}")
        else:
            print("✅ All field references in choices are valid")

        return len(issues) == 0

    def check_choice_lists(self):
        """Check for references to undefined choice lists."""
        print("\n=== Checking Choice Lists ===")

        # Get all defined choice lists
        defined_lists = set(self.choices_df['list_name'].dropna().unique())
        print(f"Found {len(defined_lists)} defined choice lists")

        issues = []

        for idx, row in self.survey_df.iterrows():
            field_type = str(row.get('type', ''))
            field_name = row.get('name', f'Row {idx}')

            # Check select_one and select_multiple types
            if field_type.startswith('select_one '):
                list_name = field_type.replace('select_one ', '').strip()
                if list_name and list_name not in defined_lists:
                    issues.append({
                        'row': idx + 2,
                        'field': field_name,
                        'type': field_type,
                        'missing_list': list_name
                    })

            elif field_type.startswith('select_multiple '):
                list_name = field_type.replace('select_multiple ', '').strip()
                if list_name and list_name not in defined_lists:
                    issues.append({
                        'row': idx + 2,
                        'field': field_name,
                        'type': field_type,
                        'missing_list': list_name
                    })

        if issues:
            print(f"\n❌ Found {len(issues)} reference(s) to undefined choice lists:\n")
            for issue in issues:
                error_msg = (f"  Row {issue['row']}: '{issue['field']}' type '{issue['type']}'\n"
                           f"    References undefined list: '{issue['missing_list']}'\n")
                print(error_msg)
                self.errors.append(error_msg)
        else:
            print("✅ All choice lists are defined")

        return len(issues) == 0

    def check_required_columns(self):
        """Check for required columns in survey and choices sheets."""
        print("\n=== Checking Required Columns ===")

        required_survey_cols = ['type', 'name']
        required_choices_cols = ['list_name', 'name', 'label']

        missing_survey = [col for col in required_survey_cols if col not in self.survey_df.columns]
        missing_choices = [col for col in required_choices_cols if col not in self.choices_df.columns]

        if missing_survey:
            error_msg = f"Survey sheet missing required columns: {missing_survey}"
            print(f"❌ {error_msg}")
            self.errors.append(error_msg)

        if missing_choices:
            error_msg = f"Choices sheet missing required columns: {missing_choices}"
            print(f"❌ {error_msg}")
            self.errors.append(error_msg)

        if not missing_survey and not missing_choices:
            print("✅ All required columns present")
            return True

        return False

    def check_duplicate_names(self):
        """Check for duplicate field names."""
        print("\n=== Checking for Duplicate Field Names ===")

        # Get all field names (excluding NaN and structural types)
        field_names = self.survey_df[
            self.survey_df['type'].notna() &
            ~self.survey_df['type'].str.contains('group|repeat', case=False, na=False)
        ]['name'].dropna()

        duplicates = field_names[field_names.duplicated()].unique()

        if len(duplicates) > 0:
            print(f"\n❌ Found {len(duplicates)} duplicate field name(s):\n")
            for dup in duplicates:
                rows = self.survey_df[self.survey_df['name'] == dup].index + 2
                error_msg = f"  Field '{dup}' appears in rows: {list(rows)}"
                print(error_msg)
                self.errors.append(error_msg)
            return False
        else:
            print("✅ No duplicate field names found")
            return True

    def check_required_fields(self):
        """Check that all question fields have required = yes.

        Non-question types (notes, groups, repeats, metadata, calculations) are excluded.
        """
        print("\n=== Checking Required Fields ===")

        # Types that are NOT questions and should be excluded from this check
        non_question_types = {
            'note', 'calculate', 'deviceid', 'subscriberid', 'simserial',
            'phonenumber', 'username', 'start', 'end', 'caseid', 'geopoint',
            'begin group', 'end group', 'begin repeat', 'end repeat',
            'begin_group', 'end_group', 'begin_repeat', 'end_repeat'
        }

        issues = []

        for idx, row in self.survey_df.iterrows():
            field_type = str(row.get('type', '')).strip().lower()
            field_name = row.get('name', f'Row {idx}')
            required_val = str(row.get('required', '')).strip().lower()

            # Skip if no type or if it's a non-question type
            if not field_type or pd.isna(row.get('type')):
                continue

            # Check if it starts with a non-question type (handles "begin group" etc.)
            is_non_question = False
            for nq_type in non_question_types:
                if field_type == nq_type or field_type.startswith(nq_type):
                    is_non_question = True
                    break

            if is_non_question:
                continue

            # This is a question type - check if required = yes
            if required_val != 'yes':
                issues.append({
                    'row': idx + 2,
                    'field': field_name,
                    'type': field_type,
                    'required': required_val if required_val else '(blank)'
                })

        if issues:
            print(f"\n⚠️  Found {len(issues)} question(s) without required=yes:\n")
            for issue in issues:
                warning_msg = (f"  Row {issue['row']}: '{issue['field']}' (type: {issue['type']})\n"
                             f"    required = {issue['required']}\n")
                print(warning_msg)
                self.warnings.append(warning_msg)
        else:
            print("✅ All questions have required=yes")

        return len(issues) == 0

    def check_other_specify_fields(self):
        """Check that 'other (specify)' choices have corresponding specify fields.

        This check looks for choice lists where 'other' requires a follow-up text field.
        It uses multiple heuristics to match various naming conventions.
        """
        print("\n=== Checking 'Other Specify' Fields ===")

        issues = []

        # Get all existing field names for pattern matching
        existing_fields = set(self.survey_df['name'].dropna().astype(str))

        # Find choice lists with 'other' option that requires specification
        # Look for 'other' choices with labels containing "specify" or similar
        other_choices = self.choices_df[
            self.choices_df['name'].astype(str).str.lower() == 'other'
        ]

        # Determine which lists need a specify field based on the label
        lists_needing_specify = set()
        for _, choice_row in other_choices.iterrows():
            list_name = choice_row.get('list_name', '')
            label = str(choice_row.get('label', '')).lower()
            # Check if label indicates specification is needed
            if 'specify' in label or 'बताएं' in label:
                lists_needing_specify.add(list_name)

        print(f"Found {len(lists_needing_specify)} choice list(s) with 'other (specify)' option")

        # Find all fields using these lists
        for idx, row in self.survey_df.iterrows():
            field_type = str(row.get('type', ''))
            field_name = str(row.get('name', ''))

            # Extract list name from select_one/select_multiple
            if field_type.startswith('select_one '):
                list_name = field_type.replace('select_one ', '').strip()
            elif field_type.startswith('select_multiple '):
                list_name = field_type.replace('select_multiple ', '').strip()
            else:
                continue

            # Only check lists that need specify fields
            if list_name not in lists_needing_specify:
                continue

            # Generate possible "other" field name patterns
            # Pattern 1: {field_name}_other (e.g., s4_provider_type_other)
            # Pattern 2: {section}{num}_other (e.g., s4_other from s4_provider_type)
            # Pattern 3: {section}{num}_{first_word}_other (e.g., s4_provider_other)
            # Pattern 4: {field_name}_other_specify
            # Pattern 5: {section}{num}_other_specify (e.g., v1_other_specify)
            # Pattern 6: {field_name}_other_text
            possible_patterns = []

            # Pattern 1: exact match
            possible_patterns.append(f"{field_name}_other")

            # Extract section prefix (e.g., 's4', 'v1', 't3')
            section_match = re.match(r'^([a-z]+\d+[a-z]?)_', field_name)
            if section_match:
                section_prefix = section_match.group(1)
                # Pattern 2: section + _other
                possible_patterns.append(f"{section_prefix}_other")
                # Pattern 5: section + _other_specify
                possible_patterns.append(f"{section_prefix}_other_specify")

                # Pattern 3: section + first part of name + _other
                remaining = field_name[len(section_prefix) + 1:]  # Skip prefix and underscore
                if '_' in remaining:
                    first_part = remaining.split('_')[0]
                    possible_patterns.append(f"{section_prefix}_{first_part}_other")

            # Pattern 4 & 6: variations with _specify or _text suffix
            possible_patterns.append(f"{field_name}_other_specify")
            possible_patterns.append(f"{field_name}_other_text")

            # Check if any pattern matches an existing field
            found_match = False
            for pattern in possible_patterns:
                if pattern in existing_fields:
                    found_match = True
                    break

            if not found_match:
                issues.append({
                    'row': idx + 2,
                    'field': field_name,
                    'tried_patterns': possible_patterns[:3],  # Show first 3 patterns
                    'list': list_name
                })

        if issues:
            print(f"\n⚠️  Found {len(issues)} field(s) with 'other (specify)' choice but no specify field:\n")
            for issue in issues:
                warning_msg = (f"  Row {issue['row']}: '{issue['field']}' (list: '{issue['list']}')\n"
                             f"    Tried patterns: {', '.join(issue['tried_patterns'])}\n")
                print(warning_msg)
                self.warnings.append(warning_msg)
        else:
            print("✅ All 'other (specify)' choices have specify fields")

        return len(issues) == 0

    def check_expression_syntax(self):
        """Check for syntax errors in relevance, calculation, and constraint expressions.

        Validates:
        - Balanced parentheses
        - Properly closed ${} field references
        - Balanced quotes (single and double)
        - SurveyCTO parser-sensitive spaced comparison operators
        - Duplicated boolean operators
        - Malformed quoted selected() choice codes
        - ODK-isms SurveyCTO rejects: '==' equality, starts-with()/contains()/
          substring-before()/substring-after()
        """
        print("\n=== Checking Expression Syntax ===")

        issues = []

        for idx, row in self.survey_df.iterrows():
            field_name = row.get('name', f'Row {idx}')

            for col in self.EXPRESSION_COLUMNS:
                if pd.notna(row.get(col)):
                    expression = str(row[col])
                    syntax_errors = self._check_expression(expression)

                    for error in syntax_errors:
                        issues.append({
                            'row': idx + 2,
                            'field': field_name,
                            'column': col,
                            'error': error,
                            'expression': expression
                        })

        if issues:
            print(f"\n❌ Found {len(issues)} expression syntax error(s):\n")
            for issue in issues:
                error_msg = (f"  Row {issue['row']}: '{issue['field']}' in column '{issue['column']}'\n"
                           f"    {issue['error']}\n"
                           f"    Expression: {issue['expression']}\n")
                print(error_msg)
                self.errors.append(error_msg)
        else:
            print("✅ All expressions have valid syntax")

        return len(issues) == 0

    def _check_expression(self, expression):
        """Check a single expression for syntax errors.

        Returns a list of error messages (empty if no errors).
        """
        errors = []

        # Check 1: Balanced parentheses
        paren_count = 0
        for i, char in enumerate(expression):
            if char == '(':
                paren_count += 1
            elif char == ')':
                paren_count -= 1
                if paren_count < 0:
                    errors.append(f"Unmatched closing parenthesis ')' at position {i}")
                    break

        if paren_count > 0:
            errors.append(f"Unclosed parenthesis - {paren_count} opening '(' without matching ')'")

        # Check 2: Properly closed ${} references
        i = 0
        while i < len(expression):
            if expression[i:i+2] == '${':
                # Find closing brace
                close_pos = expression.find('}', i + 2)
                if close_pos == -1:
                    errors.append(f"Unclosed field reference '${{' at position {i}")
                    break
                i = close_pos + 1
            else:
                i += 1

        # Check 3: Balanced quotes (but handle escaped quotes and mixed usage)
        # SurveyCTO uses single quotes for strings, double quotes when string contains single quotes
        in_single_quote = False
        in_double_quote = False

        for i, char in enumerate(expression):
            if char == "'" and not in_double_quote:
                in_single_quote = not in_single_quote
            elif char == '"' and not in_single_quote:
                in_double_quote = not in_double_quote

        if in_single_quote:
            errors.append("Unclosed single quote '")
        if in_double_quote:
            errors.append('Unclosed double quote "')

        # Check 4: SurveyCTO rejects split comparison operators like ". > = 0".
        if re.search(r'(?:>|<|!)\s+=', expression):
            errors.append("Invalid spaced comparison operator; use >=, <=, or != without spaces")

        # Check 5: duplicated boolean operators such as "or  or" usually come
        # from an accidental deleted term and SurveyCTO rejects them.
        if re.search(r'\b(?:and|or)\b\s+\b(?:and|or)\b', expression, re.IGNORECASE):
            errors.append("Duplicated boolean operator; remove the extra 'and'/'or'")

        # Check 6: malformed selected() choice arguments with an opening quote
        # but no closing quote before the selected() closing parenthesis, e.g.
        # selected(${x}, '-997). The general quote-balance check misses pairs of
        # these errors in the same expression.
        if (re.search(r"\bselected\s*\([^)]*,\s*'[^']*\)", expression)
                or re.search(r'\bselected\s*\([^)]*,\s*"[^"]*\)', expression)):
            errors.append(
                "Malformed quoted selected() choice code; close the quote before ')'"
            )

        # Check 7: '==' is generic ODK/JavaScript equality; SurveyCTO's
        # comparison operator is '=' and the parser rejects '=='.
        if '==' in expression:
            errors.append("SurveyCTO uses '=' for equality, not '=='")

        # Check 8: unsupported ODK string functions. The lookbehind keeps
        # hyphenated lookalikes such as count-selected() from matching.
        # Use substr(string, 0, N) = 'prefix', regex(string, '.*pattern.*'),
        # or selected-at() instead.
        for function_name in self.UNSUPPORTED_ODK_FUNCTIONS:
            if re.search(rf'(?<![\w-]){function_name}\s*\(', expression):
                errors.append(
                    f"Unsupported function '{function_name}()' - SurveyCTO rejects it; "
                    "use substr()/regex()/selected-at() instead"
                )

        return errors

    @staticmethod
    def _has_text(value):
        """Return True when a spreadsheet cell has non-whitespace content."""
        return pd.notna(value) and bool(str(value).strip())

    @staticmethod
    def _find_dependency_cycles(graph):
        """Return multi-field strongly connected components in a dependency graph."""
        next_index = 0
        indices = {}
        lowlinks = {}
        stack = []
        on_stack = set()
        cycles = []

        def visit(node):
            nonlocal next_index
            indices[node] = next_index
            lowlinks[node] = next_index
            next_index += 1
            stack.append(node)
            on_stack.add(node)

            for dependency in graph.get(node, set()):
                if dependency not in indices:
                    visit(dependency)
                    lowlinks[node] = min(lowlinks[node], lowlinks[dependency])
                elif dependency in on_stack:
                    lowlinks[node] = min(lowlinks[node], indices[dependency])

            if lowlinks[node] != indices[node]:
                return

            component = []
            while True:
                member = stack.pop()
                on_stack.remove(member)
                component.append(member)
                if member == node:
                    break
            if len(component) > 1:
                cycles.append(sorted(component))

        for node in graph:
            if node not in indices:
                visit(node)

        return sorted(cycles)

    def check_upload_parser_blockers(self):
        """Catch parser-level issues that SurveyCTO rejects at upload time.

        These are stricter than the local semantic checks above:
        - Spreadsheet formula errors in upload-sensitive XLSForm columns.
        - Spacer rows with no type/name/label but with relevance/calculation/etc.
        - Self-references and multi-field cycles across relevance, calculation,
          and required expressions, which SurveyCTO rejects as XPath dependency
          cycles.
        """
        print("\n=== Checking SurveyCTO Upload Parser Blockers ===")

        issues = []
        existing_fields = set(self.survey_df['name'].dropna().astype(str))
        dependency_graph = {name: set() for name in existing_fields}
        field_rows = {}

        for idx, row in self.survey_df.iterrows():
            field_type = row.get('type', '')
            field_name = row.get('name', '')
            label = row.get('label', '')

            for col in self.UPLOAD_SENSITIVE_COLUMNS:
                value = row.get(col)
                if self._has_text(value) and '#ERROR' in str(value).upper():
                    issues.append({
                        'row': idx + 2,
                        'field': str(field_name).strip() or '(blank row)',
                        'problem': f"Spreadsheet formula error in {col}: {value}",
                    })

            has_identity = any(
                self._has_text(v)
                for v in (field_type, field_name, label)
            )
            active_expressions = [
                col for col in self.EXPRESSION_COLUMNS
                if self._has_text(row.get(col))
            ]
            if not has_identity and active_expressions:
                issues.append({
                    'row': idx + 2,
                    'field': '(blank row)',
                    'problem': (
                        "Row has no type/name/label but has expression(s) in "
                        f"{', '.join(active_expressions)}"
                    ),
                })

            if self._has_text(field_name):
                name = str(field_name).strip()
                field_rows[name] = idx + 2
                for col in self.DEPENDENCY_COLUMNS:
                    expression = row.get(col)
                    if not self._has_text(expression):
                        continue
                    for ref in re.findall(r'\$\{([^}]+)\}', str(expression)):
                        base_ref = re.match(r'^([a-zA-Z_][a-zA-Z0-9_]*)', ref)
                        if base_ref and base_ref.group(1) in existing_fields:
                            dependency_graph[name].add(base_ref.group(1))

                for col in self.SELF_REFERENCE_COLUMNS:
                    expression = row.get(col)
                    if not self._has_text(expression):
                        continue
                    refs = re.findall(r'\$\{([^}]+)\}', str(expression))
                    for ref in refs:
                        base_ref = re.match(r'^([a-zA-Z_][a-zA-Z0-9_]*)', ref)
                        if base_ref and base_ref.group(1) == name:
                            issues.append({
                                'row': idx + 2,
                                'field': name,
                                'problem': (
                                    f"Self-reference in {col}: {str(expression)}"
                                ),
                            })

        for cycle in self._find_dependency_cycles(dependency_graph):
            issues.append({
                'row': min(field_rows.get(name, 0) for name in cycle),
                'field': ', '.join(cycle),
                'problem': (
                    "XPath dependency cycle across relevance/calculation/required: "
                    + ' -> '.join(cycle)
                ),
            })

        if issues:
            print(f"\n❌ Found {len(issues)} SurveyCTO upload parser blocker(s):\n")
            for issue in issues:
                error_msg = (f"  Row {issue['row']}: '{issue['field']}'\n"
                             f"    {issue['problem']}\n")
                print(error_msg)
                self.errors.append(error_msg)
        else:
            print("✅ No SurveyCTO upload parser blockers found")

        return len(issues) == 0

    def check_typos(self):
        """Check for common typos in field names and labels."""
        print("\n=== Checking for Typos ===")

        # Common typos to check for
        typos = [
            ('enumnerator', 'enumerator'),
            ('heatlh', 'health'),
            ('helath', 'health'),
            ('symtoms', 'symptoms'),
            ('sympton', 'symptom'),
            ('speciy', 'specify'),
            ('provder', 'provider'),
            ('repondent', 'respondent'),
            ('hosptial', 'hospital'),
            ('vilalge', 'village'),
            ('distirct', 'district'),
            ('sevrity', 'severity'),
            ('severeity', 'severity'),
        ]

        issues = []

        for idx, row in self.survey_df.iterrows():
            field_name = str(row.get('name', '')).lower()
            label = str(row.get('label', '')).lower()
            field_name_orig = row.get('name', f'Row {idx}')

            for typo, correct in typos:
                # Check field name
                if typo in field_name:
                    issues.append({
                        'row': idx + 2,
                        'field': field_name_orig,
                        'location': 'field name',
                        'typo': typo,
                        'correct': correct
                    })
                # Check label
                if typo in label:
                    issues.append({
                        'row': idx + 2,
                        'field': field_name_orig,
                        'location': 'label',
                        'typo': typo,
                        'correct': correct
                    })

        if issues:
            print(f"\n⚠️  Found {len(issues)} potential typo(s):\n")
            for issue in issues:
                warning_msg = (f"  Row {issue['row']}: '{issue['field']}' in {issue['location']}\n"
                             f"    Found '{issue['typo']}' - should be '{issue['correct']}'?\n")
                print(warning_msg)
                self.warnings.append(warning_msg)
        else:
            print("✅ No common typos found")

        return len(issues) == 0

    def check_missing_constraint_messages(self):
        """Check that fields with constraints have constraint messages."""
        print("\n=== Checking Constraint Messages ===")

        issues = []

        for idx, row in self.survey_df.iterrows():
            field_name = row.get('name', f'Row {idx}')
            constraint = row.get('constraint', '')
            constraint_msg = row.get('constraint_message', '')

            if pd.notna(constraint) and str(constraint).strip():
                if pd.isna(constraint_msg) or not str(constraint_msg).strip():
                    issues.append({
                        'row': idx + 2,
                        'field': field_name,
                        'constraint': str(constraint)
                    })

        if issues:
            print(f"\n⚠️  Found {len(issues)} field(s) with constraint but no message:\n")
            for issue in issues:
                warning_msg = (f"  Row {issue['row']}: '{issue['field']}'\n"
                             f"    constraint: {issue['constraint']}\n"
                             f"    Missing constraint_message\n")
                print(warning_msg)
                self.warnings.append(warning_msg)
        else:
            print("✅ All constraints have messages")

        return len(issues) == 0

    def check_integer_constraints(self):
        """Check that integer fields have constraints."""
        print("\n=== Checking Integer Constraints ===")

        issues = []

        for idx, row in self.survey_df.iterrows():
            field_type = str(row.get('type', '')).strip().lower()
            field_name = row.get('name', f'Row {idx}')
            constraint = row.get('constraint', '')
            label = str(row.get('label', ''))[:50]

            if field_type == 'integer':
                if pd.isna(constraint) or not str(constraint).strip():
                    issues.append({
                        'row': idx + 2,
                        'field': field_name,
                        'label': label
                    })

        if issues:
            print(f"\n⚠️  Found {len(issues)} integer field(s) without constraints:\n")
            for issue in issues:
                warning_msg = (f"  Row {issue['row']}: '{issue['field']}'\n"
                             f"    Label: {issue['label']}\n"
                             f"    Consider adding range validation\n")
                print(warning_msg)
                self.warnings.append(warning_msg)
        else:
            print("✅ All integer fields have constraints")

        return len(issues) == 0

    def check_calculate_fields(self):
        """Check that calculate fields have a calculation formula.

        Fields with type 'calculate' or 'calculate_here' must have a non-empty
        'calculation' column, otherwise they serve no purpose and are likely errors.

        For calculate_here fields (timing checkpoints), the calculation should
        typically be something like once(duration()) or once(format-date-time(now(), ...)).
        """
        print("\n=== Checking Calculate Fields ===")

        issues = []

        for idx, row in self.survey_df.iterrows():
            field_type = str(row.get('type', '')).strip().lower()
            field_name = row.get('name', f'Row {idx}')
            calculation = row.get('calculation', '')

            # Check both 'calculate' and 'calculate_here' types
            if field_type in ['calculate', 'calculate_here']:
                if pd.isna(calculation) or not str(calculation).strip():
                    issues.append({
                        'row': idx + 2,
                        'field': field_name,
                        'type': field_type
                    })

        if issues:
            print(f"\n❌ Found {len(issues)} calculate field(s) with empty calculation:\n")
            for issue in issues:
                error_msg = (f"  Row {issue['row']}: '{issue['field']}' (type: {issue['type']})\n"
                           f"    Missing calculation formula\n")
                print(error_msg)
                self.errors.append(error_msg)
        else:
            print("✅ All calculate fields have formulas")

        return len(issues) == 0

    def check_hindi_translations(self):
        """Check for missing Hindi translations on questions."""
        print("\n=== Checking Hindi Translations ===")

        if 'label:Hindi' not in self.survey_df.columns:
            print("ℹ️  No 'label:Hindi' column found (single language form)")
            return True

        # Question types that need labels
        question_prefixes = ['text', 'integer', 'decimal', 'select_one', 'select_multiple',
                            'geopoint', 'image', 'date', 'time']

        issues = []

        for idx, row in self.survey_df.iterrows():
            field_type = str(row.get('type', '')).lower()
            field_name = row.get('name', f'Row {idx}')
            label = row.get('label', '')
            label_hindi = row.get('label:Hindi', '')

            # Check if this is a question type
            is_question = any(field_type.startswith(prefix) for prefix in question_prefixes)

            if is_question and pd.notna(label) and str(label).strip():
                if pd.isna(label_hindi) or not str(label_hindi).strip():
                    issues.append({
                        'row': idx + 2,
                        'field': field_name,
                        'label': str(label)[:50]
                    })

        if issues:
            print(f"\n⚠️  Found {len(issues)} question(s) missing Hindi translation:\n")
            for issue in issues:
                warning_msg = (f"  Row {issue['row']}: '{issue['field']}'\n"
                             f"    English: {issue['label']}\n")
                print(warning_msg)
                self.warnings.append(warning_msg)
        else:
            print("✅ All questions have Hindi translations")

        return len(issues) == 0

    def check_select_multiple_other(self):
        """Check that select_multiple questions with 'other' option have specify fields.

        Unlike check_other_specify_fields which only checks choices labeled 'other (specify)',
        this checks ALL select_multiple questions that have any 'other' choice option.
        """
        print("\n=== Checking select_multiple 'Other' Fields ===")

        issues = []

        # Get all existing field names for pattern matching
        existing_fields = set(self.survey_df['name'].dropna().astype(str))

        # Find all choice lists that have an 'other' option (any 'other', not just 'other specify')
        lists_with_other = set(
            self.choices_df[
                self.choices_df['name'].astype(str).str.lower() == 'other'
            ]['list_name'].dropna().unique()
        )

        # Find all select_multiple fields using lists with 'other' option
        for idx, row in self.survey_df.iterrows():
            field_type = str(row.get('type', ''))
            field_name = str(row.get('name', ''))

            # Only check select_multiple
            if not field_type.startswith('select_multiple '):
                continue

            list_name = field_type.replace('select_multiple ', '').strip()

            # Only check if this list has an 'other' option
            if list_name not in lists_with_other:
                continue

            # Generate possible "other" field name patterns
            possible_patterns = []

            # Pattern 1: {field_name}_other
            possible_patterns.append(f"{field_name}_other")

            # Extract section prefix (e.g., 's4', 'v1', 't3')
            section_match = re.match(r'^([a-z]+\d+[a-z]?)_', field_name)
            if section_match:
                section_prefix = section_match.group(1)
                # Pattern 2: section + _other
                possible_patterns.append(f"{section_prefix}_other")
                # Pattern 3: section + _other_specify
                possible_patterns.append(f"{section_prefix}_other_specify")

                # Pattern 4: section + first part of name + _other
                remaining = field_name[len(section_prefix) + 1:]
                if '_' in remaining:
                    first_part = remaining.split('_')[0]
                    possible_patterns.append(f"{section_prefix}_{first_part}_other")

            # Additional patterns
            possible_patterns.append(f"{field_name}_other_specify")
            possible_patterns.append(f"{field_name}_other_text")

            # Check if any pattern matches an existing field
            found_match = False
            for pattern in possible_patterns:
                if pattern in existing_fields:
                    found_match = True
                    break

            if not found_match:
                issues.append({
                    'row': idx + 2,
                    'field': field_name,
                    'list': list_name,
                    'tried_patterns': possible_patterns[:3]
                })

        if issues:
            print(f"\n⚠️  Found {len(issues)} select_multiple field(s) with 'other' option but no specify field:\n")
            for issue in issues:
                warning_msg = (f"  Row {issue['row']}: '{issue['field']}' (list: '{issue['list']}')\n"
                             f"    Tried patterns: {', '.join(issue['tried_patterns'])}\n")
                print(warning_msg)
                self.warnings.append(warning_msg)
        else:
            print("✅ All select_multiple fields with 'other' option have specify fields")

        return len(issues) == 0

    def check_select_multiple_exclusive(self):
        """Check that select_multiple questions with exclusive options have constraints.

        Certain options like 'Don't know' (-97), 'Refuse to answer' (-98), 'Nothing',
        'Not sick', etc. should be exclusive - they cannot be selected along with other options.
        This check verifies that select_multiple questions with such options have appropriate
        constraints like: not(selected(., '-97')) or count-selected(.) = 1
        """
        print("\n=== Checking select_multiple Exclusive Options ===")

        issues = []

        # Patterns that indicate exclusive options
        # Use exact match for choice names, substring match for labels
        # Short patterns like 'na', 'dk' only match as exact choice names to avoid false positives
        exact_name_patterns = [
            '-97', '-98', '-99',  # Numeric codes for dk/refuse/na
            'dk', 'na',  # Short codes - only match exactly
            'none', 'nothing',
            'not_sick',
            'dont_know', 'dont_remember',
        ]

        # These patterns can match as substrings in labels
        label_patterns = [
            "don't know", "don't remember",
            'refuse to answer', 'declined to answer',
            'not applicable',
            'none of the above', 'none of these',
        ]

        # Build a map of choice lists to their exclusive options
        list_exclusive_opts = {}
        for list_name in self.choices_df['list_name'].dropna().unique():
            list_choices = self.choices_df[self.choices_df['list_name'] == list_name]
            exclusive_opts = []
            for _, choice_row in list_choices.iterrows():
                choice_name = str(choice_row.get('name', '')).lower()
                choice_label = str(choice_row.get('label', '')).lower() if pd.notna(choice_row.get('label')) else ''

                is_exclusive = False

                # Check exact match for name patterns
                for pattern in exact_name_patterns:
                    if choice_name == pattern:
                        is_exclusive = True
                        break

                # Check substring match for label patterns (more specific phrases)
                if not is_exclusive:
                    for pattern in label_patterns:
                        if pattern in choice_label:
                            is_exclusive = True
                            break

                if is_exclusive:
                    exclusive_opts.append(choice_row.get('name'))

            if exclusive_opts:
                list_exclusive_opts[list_name] = exclusive_opts

        # Check each select_multiple question
        for idx, row in self.survey_df.iterrows():
            field_type = str(row.get('type', ''))
            field_name = str(row.get('name', ''))
            constraint = str(row.get('constraint', '')) if pd.notna(row.get('constraint')) else ''

            if not field_type.startswith('select_multiple '):
                continue

            list_name = field_type.replace('select_multiple ', '').strip()

            # Check if this list has exclusive options
            if list_name not in list_exclusive_opts:
                continue

            exclusive_opts = list_exclusive_opts[list_name]

            # Check if constraint handles each exclusive option
            missing_constraints = []
            for opt in exclusive_opts:
                opt_str = str(opt)
                # Look for patterns like: not(selected(., 'opt')) or count-selected(.) = 1
                # or: selected(., 'opt') ... count-selected
                if opt_str not in constraint:
                    missing_constraints.append(opt_str)

            if missing_constraints:
                issues.append({
                    'row': idx + 2,
                    'field': field_name,
                    'list': list_name,
                    'exclusive_opts': exclusive_opts,
                    'missing': missing_constraints,
                    'current_constraint': constraint if constraint else '(none)'
                })

        if issues:
            print(f"\n⚠️  Found {len(issues)} select_multiple field(s) missing exclusive option constraints:\n")
            for issue in issues:
                warning_msg = (f"  Row {issue['row']}: '{issue['field']}' (list: '{issue['list']}')\n"
                             f"    Exclusive options: {issue['exclusive_opts']}\n"
                             f"    Missing constraint for: {issue['missing']}\n"
                             f"    Current constraint: {issue['current_constraint']}\n"
                             f"    Suggested: (not(selected(., '{issue['missing'][0]}')) or count-selected(.) = 1)\n")
                print(warning_msg)
                self.warnings.append(warning_msg)
        else:
            print("✅ All select_multiple fields with exclusive options have constraints")

        return len(issues) == 0

    def check_impossible_literal_values(self):
        """Flag literal values compared against a select variable that aren't in the choice list.

        Catches bugs like ``constraint: if(selected(${s_l4}, '997'), ...)`` when the
        ``s_l4`` choice list actually has ``-997`` (negative) for "Ne sait pas" -- the
        condition silently never fires. Likewise ``${var} = 5`` when ``var`` is bound to
        a list with only ``1..4`` is dead code.

        For each ``select_one X`` / ``select_multiple X`` field, this scans the row's
        expression columns for references to ``${var}`` paired with a literal RHS (via
        ``=`` / ``!=`` or ``selected()`` / ``count-selected()``) and errors if the
        literal isn't a ``name``/``value`` in ``X``.

        Skip cases (no error):
        - The select's choice list is dynamic (``select_one_from_file foo.csv``,
          ``select_one ${dynamic_list}``, etc.) -- can't validate without the CSV.
        - The RHS is itself a ``${...}`` reference, not a literal.
        - The literal is empty, whitespace, or ``''``.
        - The calling row is disabled (already filtered out by ``load_form``).
        - The select-defining row is disabled (also filtered out, so its var name
          won't appear in the var-to-list map).
        """
        print("\n=== Checking Impossible Literal Values in Expressions ===")

        # Step 1: build var_name -> choice_list_name map from select_one / select_multiple
        # rows. Skip dynamic list types (select_*_from_file, ${...} list names).
        var_to_list = {}
        for _, row in self.survey_df.iterrows():
            field_type = str(row.get('type', '')).strip()
            field_name = row.get('name')
            if pd.isna(field_name) or not str(field_name).strip():
                continue
            name = str(field_name).strip()

            list_name = None
            if field_type.startswith('select_one_from_file ') or \
                    field_type.startswith('select_multiple_from_file '):
                continue  # dynamic external CSV -- skip
            if field_type.startswith('select_one '):
                list_name = field_type[len('select_one '):].strip()
            elif field_type.startswith('select_multiple '):
                list_name = field_type[len('select_multiple '):].strip()
            else:
                continue

            if not list_name or '${' in list_name:
                continue  # blank or dynamic list reference
            var_to_list[name] = list_name

        # Step 2: build choice_list_name -> set of valid literal values (as strings).
        # load_form already aliases the 'value' column to 'name' if needed.
        list_to_values = {}
        for _, choice_row in self.choices_df.iterrows():
            list_name = choice_row.get('list_name')
            value = choice_row.get('name')
            if pd.isna(list_name) or pd.isna(value):
                continue
            list_name = str(list_name).strip()
            value_str = str(value).strip()
            if not list_name or not value_str:
                continue
            list_to_values.setdefault(list_name, set()).add(value_str)

        # Step 3: regex patterns. The literal capture is intentionally permissive:
        # we then verify it's a real literal (not blank, not just whitespace).
        # Each pattern yields (var_name, literal_value) groups.
        # Note: equality patterns require the ${var} on the LEFT side. That misses
        # ``5 = ${var}`` but is consistent with how XLSForms are written in practice.
        eq_quoted_re = re.compile(r"\$\{(\w+)\}\s*(?:=|!=)\s*'([^']*)'")
        eq_double_quoted_re = re.compile(r'\$\{(\w+)\}\s*(?:=|!=)\s*"([^"]*)"')
        eq_unquoted_re = re.compile(r"\$\{(\w+)\}\s*(?:=|!=)\s*(-?\d+(?:\.\d+)?)\b")
        sel_quoted_re = re.compile(
            r"(?:count-)?selected\s*\(\s*(?:\.|\$\{(\w+)\})\s*,\s*'([^']*)'\s*\)"
        )
        sel_double_quoted_re = re.compile(
            r'(?:count-)?selected\s*\(\s*(?:\.|\$\{(\w+)\})\s*,\s*"([^"]*)"\s*\)'
        )
        sel_unquoted_re = re.compile(
            r"(?:count-)?selected\s*\(\s*(?:\.|\$\{(\w+)\})\s*,\s*(-?\d+(?:\.\d+)?)\s*\)"
        )

        check_cols = ('relevance', 'relevant', 'constraint', 'calculation',
                      'choice_filter', 'repeat_count', 'required', 'default')

        issues = []

        for idx, row in self.survey_df.iterrows():
            field_name = row.get('name', f'Row {idx}')
            row_var = str(field_name).strip() if pd.notna(field_name) else ''

            for col in check_cols:
                if col not in self.survey_df.columns:
                    continue
                value = row.get(col)
                if pd.isna(value):
                    continue
                expression = str(value)
                if not expression.strip():
                    continue

                # Collect (var_name_or_None, literal) hits from all patterns.
                # For selected()/count-selected() with `.` as first arg, var_name is
                # None and we substitute the current row's field name (self-reference).
                hits = []
                for m in eq_quoted_re.finditer(expression):
                    hits.append((m.group(1), m.group(2)))
                for m in eq_double_quoted_re.finditer(expression):
                    hits.append((m.group(1), m.group(2)))
                for m in eq_unquoted_re.finditer(expression):
                    hits.append((m.group(1), m.group(2)))
                for m in sel_quoted_re.finditer(expression):
                    hits.append((m.group(1) or row_var, m.group(2)))
                for m in sel_double_quoted_re.finditer(expression):
                    hits.append((m.group(1) or row_var, m.group(2)))
                for m in sel_unquoted_re.finditer(expression):
                    hits.append((m.group(1) or row_var, m.group(2)))

                for var_name, literal in hits:
                    if not var_name:
                        continue
                    # Skip non-literal RHS: empty / whitespace.
                    literal_str = literal.strip()
                    if not literal_str:
                        continue
                    # Skip ${...} that snuck through (regexes don't match these, but
                    # defensive).
                    if literal_str.startswith('${'):
                        continue
                    # Only check variables that are bound to a (static) choice list.
                    if var_name not in var_to_list:
                        continue
                    list_name = var_to_list[var_name]
                    if list_name not in list_to_values:
                        continue  # list undefined -- different check flags this
                    valid_values = list_to_values[list_name]
                    if literal_str in valid_values:
                        continue  # all good

                    # Also try a numeric-normalised comparison (e.g. '5' vs '5.0') to
                    # avoid false positives.
                    matched = False
                    try:
                        lit_num = float(literal_str)
                        for v in valid_values:
                            try:
                                if float(v) == lit_num:
                                    matched = True
                                    break
                            except (TypeError, ValueError):
                                continue
                    except (TypeError, ValueError):
                        pass
                    if matched:
                        continue

                    issues.append({
                        'row': idx + 2,
                        'field': field_name if pd.notna(field_name) else f'Row {idx}',
                        'column': col,
                        'var': var_name,
                        'list': list_name,
                        'literal': literal_str,
                        'valid_values': sorted(valid_values, key=lambda s: (len(s), s)),
                    })

        if issues:
            print(f"\n❌ Found {len(issues)} impossible literal value reference(s):\n")
            for issue in issues:
                valid = issue['valid_values']
                if len(valid) > 10:
                    valid_display = ','.join(valid[:10]) + f',... ({len(valid)} total)'
                else:
                    valid_display = ','.join(valid)
                error_msg = (
                    f"  Row {issue['row']} [{issue['field']}] — {issue['column']} "
                    f"references impossible value '{issue['literal']}' for "
                    f"${{{issue['var']}}} (choice list '{issue['list']}' has: "
                    f"{valid_display})\n"
                )
                print(error_msg)
                self.errors.append(error_msg)
        else:
            print("✅ All literal comparisons reference valid choice values")

        return len(issues) == 0

    def check_numeric_refuse_option(self):
        """Check that numeric fields (integer/decimal) have -999 refuse option.

        All numeric fields should allow respondents to refuse answering by entering -999.
        The constraint should include 'or . = -999' or similar pattern.
        """
        print("\n=== Checking Numeric Refuse Option (-999) ===")

        issues = []

        for idx, row in self.survey_df.iterrows():
            field_type = str(row.get('type', '')).strip().lower()
            field_name = row.get('name', f'Row {idx}')
            constraint = str(row.get('constraint', '')) if pd.notna(row.get('constraint')) else ''

            # Only check integer and decimal types
            if field_type not in ['integer', 'decimal']:
                continue

            # Check if -999 is in the constraint
            if '-999' not in constraint:
                issues.append({
                    'row': idx + 2,
                    'field': field_name,
                    'type': field_type,
                    'constraint': constraint if constraint else '(no constraint)'
                })

        if issues:
            print(f"\n⚠️  Found {len(issues)} numeric field(s) without -999 refuse option:\n")
            for issue in issues:
                warning_msg = (f"  Row {issue['row']}: '{issue['field']}' (type: {issue['type']})\n"
                             f"    Constraint: {issue['constraint']}\n"
                             f"    Add: or . = -999\n")
                print(warning_msg)
                self.warnings.append(warning_msg)
        else:
            print("✅ All numeric fields have -999 refuse option")

        return len(issues) == 0

    def check_blank_names(self):
        """Check that all rows with a type also have a non-blank name.

        SurveyCTO requires every field (including metadata types like subscriberid)
        to have a non-whitespace name. A space or empty name causes upload errors like:
        "Question or group with no name [row : N]"
        """
        print("\n=== Checking for Blank/Missing Names ===")

        issues = []

        for idx, row in self.survey_df.iterrows():
            field_type = row.get('type', '')
            field_name = row.get('name', '')

            # Skip rows with no type
            if pd.isna(field_type) or not str(field_type).strip():
                continue

            # Check if name is missing, NaN, or whitespace-only
            if pd.isna(field_name) or not str(field_name).strip():
                issues.append({
                    'row': idx + 2,
                    'type': str(field_type).strip(),
                    'name': repr(field_name) if not pd.isna(field_name) else '(empty)'
                })

        if issues:
            print(f"\n❌ Found {len(issues)} row(s) with type but blank/missing name:\n")
            for issue in issues:
                error_msg = (f"  Row {issue['row']}: type='{issue['type']}', name={issue['name']}\n"
                           f"    Every row with a type must have a non-blank name\n")
                print(error_msg)
                self.errors.append(error_msg)
        else:
            print("✅ All typed rows have names")

        return len(issues) == 0

    def check_empty_groups(self):
        """Check that groups and repeats have at least one enabled child.

        After filtering disabled rows, a group may have no children left.
        SurveyCTO rejects these with: "Group has no children! Group: <name>"
        This check handles nested groups correctly.
        """
        print("\n=== Checking for Empty Groups ===")

        issues = []

        # Build a list of (type, name) tuples preserving order
        rows = []
        for idx, row in self.survey_df.iterrows():
            field_type = str(row.get('type', '')).strip().lower() if pd.notna(row.get('type')) else ''
            field_name = str(row.get('name', '')).strip() if pd.notna(row.get('name')) else ''
            rows.append((idx + 2, field_type, field_name))  # Excel row, type, name

        # Use a stack to track group nesting and child counts
        # Each stack entry: (excel_row, group_name, child_count)
        stack = []

        for excel_row, field_type, field_name in rows:
            if field_type in ('begin group', 'begin_group', 'begin repeat', 'begin_repeat'):
                stack.append((excel_row, field_name, 0))
            elif field_type in ('end group', 'end_group', 'end repeat', 'end_repeat'):
                if stack:
                    begin_row, group_name, child_count = stack.pop()
                    if child_count == 0:
                        issues.append({
                            'row': begin_row,
                            'name': group_name,
                            'end_row': excel_row
                        })
                    # The group itself counts as a child of its parent
                    if stack:
                        parent = stack[-1]
                        stack[-1] = (parent[0], parent[1], parent[2] + 1)
            else:
                # Non-structural row — increment child count of current group
                if stack and field_type:
                    parent = stack[-1]
                    stack[-1] = (parent[0], parent[1], parent[2] + 1)

        if issues:
            print(f"\n❌ Found {len(issues)} empty group(s)/repeat(s) (no enabled children):\n")
            for issue in issues:
                error_msg = (f"  Row {issue['row']}: group '{issue['name']}' "
                           f"(ends at row {issue['end_row']}) has no enabled children\n"
                           f"    SurveyCTO will reject this. Remove the group or ensure it has content.\n")
                print(error_msg)
                self.errors.append(error_msg)
        else:
            print("✅ All groups/repeats have enabled children")

        return len(issues) == 0

    def check_naming_conventions(self):
        """Check for naming convention issues in field names."""
        print("\n=== Checking Naming Conventions ===")

        issues = []

        for idx, row in self.survey_df.iterrows():
            field_name = str(row.get('name', ''))
            field_type = str(row.get('type', '')).lower()

            if not field_name or pd.isna(row.get('name')):
                continue

            # Skip metadata types
            if field_type in ['deviceid', 'subscriberid', 'simserial', 'phonenumber',
                             'username', 'start', 'end', 'caseid']:
                continue

            # Check for camelCase
            if re.search(r'[a-z][A-Z]', field_name):
                issues.append({
                    'row': idx + 2,
                    'field': field_name,
                    'issue': 'camelCase detected (use snake_case)'
                })

            # Check for dots in name (unusual, should use underscores)
            if '.' in field_name:
                issues.append({
                    'row': idx + 2,
                    'field': field_name,
                    'issue': 'dot in name (use underscores for consistency)'
                })

            # Check for spaces
            if ' ' in field_name:
                issues.append({
                    'row': idx + 2,
                    'field': field_name,
                    'issue': 'space in name (use underscores)'
                })

            # Check for uppercase
            if field_name != field_name.lower():
                issues.append({
                    'row': idx + 2,
                    'field': field_name,
                    'issue': 'uppercase letters (use lowercase)'
                })

        if issues:
            print(f"\n⚠️  Found {len(issues)} naming convention issue(s):\n")
            for issue in issues:
                warning_msg = f"  Row {issue['row']}: '{issue['field']}' - {issue['issue']}\n"
                print(warning_msg)
                self.warnings.append(warning_msg)
        else:
            print("✅ All field names follow conventions")

        return len(issues) == 0

    def check_conditional_formatting(self):
        """Check that conditional formatting rules are preserved in the survey sheet.

        The survey sheet uses type-based color coding to highlight different question types.
        These rules make the form easier to read and maintain. If they're removed (e.g., by
        re-saving with pandas), this check will fail.
        """
        print("\n=== Checking Conditional Formatting Rules ===")

        try:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb['survey']
        except Exception as e:
            print(f"⚠️  Could not load file with openpyxl: {e}")
            return True  # Don't fail if we can't check

        # Get all conditional formatting rules
        cf_rules = ws.conditional_formatting._cf_rules

        if not cf_rules:
            error_msg = ("❌ No conditional formatting rules found in survey sheet!\n"
                        "   The type-based color coding has been removed.\n"
                        "   Restore from a backup or re-apply formatting from template.")
            print(error_msg)
            self.errors.append(error_msg)
            wb.close()
            return False

        # Extract all formulas from the rules
        all_formulas = []
        for _, rules in cf_rules.items():
            for rule in rules:
                if rule.formula:
                    all_formulas.extend(rule.formula)

        # Expected formatting rules - these are the key type-based highlights
        expected_patterns = [
            ('begin group', '$A1="begin group"'),
            ('end group', '$A1="end group"'),
            ('begin repeat', '$A1="begin repeat"'),
            ('end repeat', '$A1="end repeat"'),
            ('text', '$A1="text"'),
            ('integer', '$A1="integer"'),
            ('decimal', '$A1="decimal"'),
            ('note', '$A1="note"'),
            ('calculate', 'calculate'),  # May be in OR with calculate_here
            ('select_one/select_multiple', 'select_'),  # Complex formula with LEFT()
            ('disabled rows', '$P1="yes"'),  # Strikethrough for disabled
            ('metadata fields', 'username'),  # Part of OR for metadata types
        ]

        missing_rules = []
        found_rules = []

        for rule_name, pattern in expected_patterns:
            found = any(pattern.lower() in formula.lower() for formula in all_formulas)
            if found:
                found_rules.append(rule_name)
            else:
                missing_rules.append(rule_name)

        # Count total rules
        total_rules = sum(len(rules) for rules in cf_rules.values())
        print(f"  Found {total_rules} conditional formatting rule(s)")
        print(f"  Type-based rules verified: {len(found_rules)}/{len(expected_patterns)}")

        if missing_rules:
            # Only error if many rules are missing (suggests major formatting loss)
            if len(missing_rules) > len(expected_patterns) // 2:
                error_msg = (f"❌ Many conditional formatting rules are missing!\n"
                            f"   Missing: {', '.join(missing_rules)}\n"
                            f"   The type-based color coding may have been damaged.\n"
                            f"   Consider restoring from backup.")
                print(error_msg)
                self.errors.append(error_msg)
                wb.close()
                return False
            else:
                warning_msg = (f"⚠️  Some conditional formatting rules may be missing:\n"
                              f"   {', '.join(missing_rules)}")
                print(warning_msg)
                self.warnings.append(warning_msg)
        else:
            print("✅ All expected conditional formatting rules are present")

        wb.close()
        return len(missing_rules) <= len(expected_patterns) // 2

    def check_formatting_preserved(self):
        """Check that cell formatting (especially red text for unverified translations) is preserved.

        Red text in Hindi columns indicates unverified translations that need review.
        If formatting is accidentally removed (e.g., by using pandas to save), this check will fail.
        """
        print("\n=== Checking Cell Formatting ===")

        # Reference file with known good formatting
        reference_file = self.file_path.parent / 'backups' / 'ai_health_pilot_baseline_backup_review_SAFE.xlsx'

        try:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb['survey']
        except Exception as e:
            print(f"⚠️  Could not load file with openpyxl: {e}")
            return True  # Don't fail if we can't check

        # Count formatted cells (non-black font colors)
        def count_formatted_cells(worksheet):
            formatted = 0
            red_text = 0
            for row in worksheet.iter_rows(min_row=1, max_row=min(worksheet.max_row, 300)):
                for cell in row:
                    if cell.font and cell.font.color:
                        color = cell.font.color
                        if color.type == 'rgb' and color.rgb:
                            rgb = color.rgb.upper()
                            if rgb not in ['00000000', 'FF000000', None, '']:
                                formatted += 1
                                # Check for red (FFFF0000 or 00FF0000)
                                if rgb in ['FFFF0000', '00FF0000'] or rgb.startswith('FFFF00'):
                                    red_text += 1
            return formatted, red_text

        current_formatted, current_red = count_formatted_cells(ws)
        wb.close()

        print(f"  Current file: {current_formatted} formatted cells, {current_red} red text cells")

        # Compare against reference if available
        if reference_file.exists():
            try:
                ref_wb = openpyxl.load_workbook(reference_file)
                ref_ws = ref_wb['survey']
                ref_formatted, ref_red = count_formatted_cells(ref_ws)
                ref_wb.close()

                print(f"  Reference file: {ref_formatted} formatted cells, {ref_red} red text cells")

                # Check if formatting has been significantly reduced
                if current_formatted < ref_formatted * 0.5:
                    error_msg = (f"  ❌ Formatting may have been removed!\n"
                                f"     Current: {current_formatted} formatted cells\n"
                                f"     Reference: {ref_formatted} formatted cells\n"
                                f"     Red text (unverified translations) may have been lost.\n"
                                f"     Restore from backup: {reference_file.name}")
                    print(error_msg)
                    self.errors.append(error_msg)
                    return False

                if current_red < ref_red * 0.5 and ref_red > 5:
                    warning_msg = (f"  ⚠️  Red text count reduced significantly\n"
                                  f"     Current: {current_red}, Reference: {ref_red}\n"
                                  f"     This may indicate translations were verified OR formatting was lost.")
                    print(warning_msg)
                    self.warnings.append(warning_msg)

            except Exception as e:
                print(f"  ⚠️  Could not load reference file: {e}")

        else:
            print(f"  ℹ️  No reference file found at {reference_file}")
            # Still check that there's SOME formatting if this is a Hindi survey
            if 'label:Hindi' in self.survey_df.columns and current_formatted == 0:
                warning_msg = ("  ⚠️  No formatted cells found in a multi-language survey.\n"
                             "     Red text should mark unverified Hindi translations.")
                print(warning_msg)
                self.warnings.append(warning_msg)

        print("✅ Formatting check complete")
        return True

    def check_public_key_format(self):
        """Validate SurveyCTO's settings.public_key representation."""
        print("\n=== Checking Encryption Public Key ===")

        if self.settings_df is None or 'public_key' not in self.settings_df.columns:
            print("  ℹ️  No public_key setting found")
            return True

        keys = [
            str(value).strip()
            for value in self.settings_df['public_key']
            if pd.notna(value) and str(value).strip()
        ]
        if not keys:
            print("  ℹ️  No public encryption key configured")
            return True

        issues = []
        for key in keys:
            if '-----BEGIN' in key or '-----END' in key or re.search(r'\s', key):
                issues.append(
                    "settings.public_key must contain only the single-line base64 "
                    "DER payload, without PEM headers, footers, or whitespace"
                )
                continue

            try:
                decoded = base64.b64decode(key, validate=True)
            except (binascii.Error, ValueError):
                issues.append("settings.public_key is not valid base64")
                continue

            if not decoded or decoded[0] != 0x30:
                issues.append(
                    "settings.public_key does not decode to a DER SEQUENCE"
                )

        if issues:
            for issue in dict.fromkeys(issues):
                print(f"  ❌ {issue}")
                self.errors.append(issue)
            return False

        print("✅ Encryption public_key is valid single-line base64 DER")
        return True

    def check_version_formula(self):
        """Check that the version formula in settings has been evaluated.

        The settings sheet has a version formula that generates a timestamp (YYMMDDHHmm).
        If the formula hasn't been evaluated (no cached value), we run recalc_excel.sh
        to open Excel and force recalculation.
        """
        print("\n=== Checking Version Formula ===")

        try:
            # Load with data_only=True to get cached values
            wb_data = openpyxl.load_workbook(self.file_path, data_only=True)
            if 'settings' not in wb_data.sheetnames:
                print("  ℹ️  No settings sheet found")
                wb_data.close()
                return True

            settings = wb_data['settings']
            cached_value = settings['C2'].value
            wb_data.close()

            # Also check the formula itself
            wb_formula = openpyxl.load_workbook(self.file_path)
            formula = wb_formula['settings']['C2'].value
            wb_formula.close()

            # Check if it's a formula
            is_formula = formula and str(formula).startswith('=')

            if is_formula:
                print(f"  Formula: {str(formula)[:60]}...")

                if cached_value:
                    print(f"  ✅ Cached value: {cached_value}")
                    return True
                else:
                    print("  ⚠️  Version formula has not been evaluated (no cached value)")
                    print("  Attempting to recalculate via Excel...")

                    # Try to run recalc_excel.sh
                    recalc_script = Path(__file__).parent / 'recalc_excel.sh'
                    if recalc_script.exists():
                        try:
                            result = subprocess.run(
                                [str(recalc_script), str(self.file_path.absolute())],
                                capture_output=True,
                                text=True,
                                timeout=60
                            )
                            if result.returncode == 0:
                                print("  ✅ Recalculated formulas via Excel")

                                # Verify the cached value is now present
                                wb_verify = openpyxl.load_workbook(self.file_path, data_only=True)
                                new_cached = wb_verify['settings']['C2'].value
                                wb_verify.close()

                                if new_cached:
                                    print(f"  ✅ Version is now: {new_cached}")
                                    return True
                                else:
                                    # Excel may not have cached the value in a way openpyxl can read
                                    # But the formula is preserved, so Excel will calculate it when opened
                                    print("  ℹ️  Cached value not readable by openpyxl, but formula preserved")
                                    print("  ℹ️  Excel will calculate the version when the file is opened")
                                    return True
                            else:
                                print(f"  ⚠️  Excel recalculation failed: {result.stderr}")
                                self.warnings.append("Could not recalculate version formula")
                                return False
                        except subprocess.TimeoutExpired:
                            print("  ⚠️  Excel recalculation timed out")
                            self.warnings.append("Version formula recalculation timed out")
                            return False
                        except Exception as e:
                            print(f"  ⚠️  Could not run recalc script: {e}")
                            self.warnings.append(f"Could not recalculate version: {e}")
                            return False
                    else:
                        print(f"  ⚠️  recalc_excel.sh not found at {recalc_script}")
                        self.warnings.append("Version formula not evaluated, recalc script not found")
                        return False
            else:
                # It's a static value, not a formula
                if formula:
                    print(f"  Version (static): {formula}")
                else:
                    print("  ⚠️  No version set in settings")
                    self.warnings.append("No version set in settings sheet")
                return True

        except Exception as e:
            print(f"  ⚠️  Could not check version: {e}")
            return True  # Don't fail the whole check for this

    def check_dynamic_choice_labels(self):
        """Dynamic search() choice rows must hold bare column names, not markup.

        SurveyCTO reads a dynamic choice row's value/label cells as the names of
        columns in the attached data source, and it does so in EVERY language
        column. Wrapping a translation in <i>...</i> (or red-font markup, or a
        stray space) makes the first and last comma-separated tokens something
        like "<i>respondent_name" and "interview_date</i>", which match no
        column, so the picker silently renders blank in that language only.
        Caught live in Aug 2026: a Bengali endline showed only one of three
        columns, leaving enumerators unable to tell two same-ID people apart.
        """
        print("\n=== Checking Dynamic (search) Choice Labels ===")
        if self.survey_df is None or self.choices_df is None:
            return True

        search_lists = set()
        app_col = self._col(self.survey_df, 'appearance')
        type_col = self._col(self.survey_df, 'type')
        if app_col is None or type_col is None:
            return True
        for _, row in self.survey_df.iterrows():
            app = str(row.get(app_col) or '')
            if 'search(' not in app:
                continue
            tp = str(row.get(type_col) or '')
            m = re.match(r'select_(?:one|multiple)\s+(\S+)', tp.strip())
            if m:
                search_lists.add(m.group(1))
        if not search_lists:
            print("  ℹ️  No search()-backed choice lists")
            return True

        list_col = self._col(self.choices_df, 'list_name')
        if list_col is None:
            return True
        target_cols = [c for c in self.choices_df.columns
                       if str(c).lower() == 'value' or str(c).lower() == 'name'
                       or str(c).lower().startswith('label')]
        found = False
        for idx, row in self.choices_df.iterrows():
            if str(row.get(list_col) or '').strip() not in search_lists:
                continue
            for col in target_cols:
                cell = row.get(col)
                if cell is None or str(cell).strip() == '' or str(cell) == 'nan':
                    continue
                text = str(cell)
                if re.search(r'<[^>]+>', text):
                    found = True
                    msg = (f"Row {idx + 2}: dynamic search() choice list "
                           f"'{row.get(list_col)}' has markup in column '{col}': "
                           f"{text!r}\n    These cells name data-source COLUMNS in "
                           f"every language. Strip the markup or the picker "
                           f"renders blank in this language.")
                    print(f"  ❌ {msg}")
                    self.errors.append(msg)
        if not found:
            print(f"  ✅ {len(search_lists)} search()-backed list(s) hold bare column names")
        return not found

    def check_end_metadata_arithmetic(self):
        """A plain calculate over the `end` metadata field never populates.

        `end` is stamped when the form is finalised, and a plain `calculate` is
        not re-evaluated at that moment, so an interview-duration field written
        as round((decimal-date-time(${end_time}) - ...) * 86400, 0) exports
        EMPTY. Confirmed against real submissions in Aug 2026. Use
        `calculate_here` with `once(duration())` on the form's last row instead.
        """
        print("\n=== Checking Duration Calculations ===")
        if self.survey_df is None:
            return True
        type_col = self._col(self.survey_df, 'type')
        name_col = self._col(self.survey_df, 'name')
        calc_col = self._col(self.survey_df, 'calculation')
        if type_col is None or calc_col is None:
            return True

        end_fields = {str(r.get(name_col)).strip()
                      for _, r in self.survey_df.iterrows()
                      if str(r.get(type_col) or '').strip() == 'end'
                      and str(r.get(name_col) or '').strip() not in ('', 'nan')}
        if not end_fields:
            print("  ℹ️  No `end` metadata field declared")
            return True

        hits = 0
        for idx, row in self.survey_df.iterrows():
            if str(row.get(type_col) or '').strip() != 'calculate':
                continue
            calc = str(row.get(calc_col) or '')
            if not any('${%s}' % f in calc for f in end_fields):
                continue
            hits += 1
            msg = (f"Row {idx + 2}: '{row.get(name_col)}' is a plain `calculate` "
                   f"referencing the `end` metadata field. `end` is only stamped "
                   f"at finalisation and plain calculates are not re-evaluated "
                   f"then, so this exports EMPTY. Use `calculate_here` with "
                   f"`once(duration())` on the last row of the form.")
            print(f"  ⚠️  {msg}")
            self.warnings.append(msg)
        if not hits:
            print("  ✅ No plain calculate depends on the `end` metadata field")
        return True

    def check_timing_instrumentation(self):
        """Require survey boundaries and reasonable section timing coverage.

        New forms use the canonical five-field section bundle documented in
        references/timing.md. Existing K2 forms may instead use adjacent
        duration_*/time_* cumulative checkpoint pairs. Structural omissions
        are errors. Coverage below a deliberately conservative size heuristic
        is a warning because group structure cannot identify semantic sections
        perfectly.
        """
        print("\n=== Checking Timing Instrumentation ===")
        if self.platform != 'surveycto':
            print(f"  ℹ️  SurveyCTO calculate_here timing checks skipped for "
                  f"--platform {self.platform}")
            return True
        if self.survey_df is None:
            return True

        type_col = self._col(self.survey_df, 'type')
        name_col = self._col(self.survey_df, 'name')
        calc_col = self._col(self.survey_df, 'calculation')
        relevance_col = self._col(self.survey_df, 'relevance')
        appearance_col = self._col(self.survey_df, 'appearance')
        if type_col is None or name_col is None:
            return True

        def cell(value):
            if value is None or pd.isna(value):
                return ''
            return str(value).strip()

        rows = []
        fields = {}
        for position, (idx, row) in enumerate(self.survey_df.iterrows()):
            item = {
                'position': position,
                'excel_row': idx + 2,
                'type': cell(row.get(type_col)).lower(),
                'name': cell(row.get(name_col)),
                'calculation': cell(row.get(calc_col)) if calc_col else '',
                'relevance': cell(row.get(relevance_col)) if relevance_col else '',
                'appearance': cell(row.get(appearance_col)).lower()
                    if appearance_col else '',
            }
            rows.append(item)
            if item['name']:
                fields[item['name'].lower()] = item

        structure_stack = []
        for item in rows:
            item['group_path'] = tuple(structure_stack)
            if item['type'] in {'begin group', 'begin repeat'}:
                structure_stack.append((item['type'], item['name']))
            elif item['type'] in {'end group', 'end repeat'} and structure_stack:
                structure_stack.pop()

        structure_bounds = {}
        open_structures = []
        for item in rows:
            if item['type'] in {'begin group', 'begin repeat'}:
                open_structures.append(item)
            elif item['type'] in {'end group', 'end repeat'} and open_structures:
                begin_item = open_structures.pop()
                structure_bounds.setdefault(begin_item['name'].lower(), []).append(
                    (begin_item['position'], item['position']))

        question_prefixes = {
            'text', 'integer', 'decimal', 'date', 'time', 'datetime',
            'geopoint', 'geotrace', 'geoshape', 'barcode', 'image', 'audio',
            'video', 'file', 'range', 'rank', 'select_one', 'select_multiple',
            'enumerator', 'email', 'acknowledge'}

        def is_question(item):
            prefix = item['type'].split()[0] if item['type'] else ''
            return prefix in question_prefixes

        respondent_rows = [item for item in rows if is_question(item)]

        duration_re = re.compile(
            r'^\s*once\s*\(\s*duration\s*\(\s*\)\s*\)\s*$', re.I)
        clock_re = re.compile(
            r'''^\s*once\s*\(\s*format-date-time\s*\(\s*now\s*\(\s*\)'''
            r'''\s*,\s*(["'])(.+?)\1\s*\)\s*\)\s*$''', re.I | re.S)
        canonical_clock_re = re.compile(
            r'''^\s*once\s*\(\s*format-date-time\s*\(\s*now\s*\(\s*\)'''
            r'''\s*,\s*(["'])%Y-%m-%dT%H:%M:%S\1\s*\)\s*\)\s*$''', re.I)

        def is_duration_checkpoint(item):
            return (item and item['type'] == 'calculate_here'
                    and duration_re.match(item['calculation']))

        def is_clock_checkpoint(item, canonical=False):
            pattern = canonical_clock_re if canonical else clock_re
            if not item or item['type'] != 'calculate_here':
                return False
            match = pattern.match(item['calculation'])
            if not match or canonical:
                return bool(match)
            date_format = match.group(2)
            has_year = any(token in date_format for token in ('%Y', '%y'))
            has_month = any(token in date_format for token in ('%m', '%b', '%B', '%j'))
            has_day = any(token in date_format for token in ('%d', '%e', '%j'))
            has_time = (any(token in date_format for token in ('%H', '%I'))
                        and '%M' in date_format)
            return has_year and has_month and has_day and has_time

        def is_survey_clock(name, boundary):
            tokens = set(re.split(r'[_\-\s]+', name.lower()))
            boundary_tokens = {'start', 'begin'} if boundary == 'start' else {'end', 'finish'}
            return (bool(tokens & {'survey', 'interview'})
                    and bool(tokens & boundary_tokens)
                    and bool(tokens & {'time', 'timestamp'}))

        def add_error(message):
            print(f"  ❌ {message}")
            self.errors.append(message)

        def add_warning(message):
            print(f"  ⚠️  {message}")
            self.warnings.append(message)

        metadata_types = {item['type'] for item in rows}
        for metadata_type in ('start', 'end'):
            if metadata_type not in metadata_types:
                add_error(f"Missing metadata `{metadata_type}` row. Keep the standard "
                          f"SurveyCTO {metadata_type} metadata as well as timing fields.")

        chosen_clocks = {}
        for boundary in ('start', 'end'):
            canonical_name = f'survey_{boundary}_time'
            candidates = [item for item in rows
                          if is_survey_clock(item['name'], boundary)]
            chosen = fields.get(canonical_name)
            if chosen is None and candidates:
                chosen = candidates[0]
            chosen_clocks[boundary] = chosen
            if chosen is None:
                add_error(f"Missing survey-level {boundary} timestamp. Add "
                          f"`{canonical_name}` with calculate_here and "
                          f"once(format-date-time(now(), ...)).")
            elif not is_clock_checkpoint(chosen):
                add_error(f"Row {chosen['excel_row']}: survey-level {boundary} "
                          f"timestamp '{chosen['name']}' must be a calculate_here "
                          f"using once(format-date-time(now(), ...)).")
            elif chosen['group_path'] or chosen['relevance']:
                add_error(f"Row {chosen['excel_row']}: survey-level {boundary} "
                          f"timestamp '{chosen['name']}' must be outside all groups "
                          "and repeats and have no relevance.")

        if respondent_rows:
            first_question = respondent_rows[0]['position']
            last_question = respondent_rows[-1]['position']
            start_clock = chosen_clocks['start']
            end_clock = chosen_clocks['end']
            if start_clock is not None and start_clock['position'] >= first_question:
                add_error(f"Row {start_clock['excel_row']}: survey-level start "
                          f"timestamp '{start_clock['name']}' must precede the first "
                          "respondent-input field.")
            if end_clock is not None and end_clock['position'] <= last_question:
                add_error(f"Row {end_clock['excel_row']}: survey-level end timestamp "
                          f"'{end_clock['name']}' must follow the last "
                          "respondent-input field.")

        canonical_survey_names = (
            'survey_start_elapsed_sec', 'survey_start_time',
            'survey_end_elapsed_sec', 'survey_end_time', 'overall_duration_sec')
        present_survey_names = [name for name in canonical_survey_names if name in fields]
        canonical_survey_complete = len(present_survey_names) == len(canonical_survey_names)
        if present_survey_names and not canonical_survey_complete:
            missing = [name for name in canonical_survey_names if name not in fields]
            add_error("Incomplete canonical survey timing bundle. Missing: "
                      + ", ".join(missing) + ".")

        def validate_guarded_duration(item, start_name, end_name):
            calculation = item['calculation']
            compact = re.sub(r'\s+', '', calculation.lower())
            start_ref = f'${{{start_name}}}'.lower()
            end_ref = f'${{{end_name}}}'.lower()
            guards_present = (f'string-length({start_ref})' in compact
                              and f'string-length({end_ref})' in compact)
            subtraction = f'{end_ref}-{start_ref}' in compact
            rounded = f'round({end_ref}-{start_ref},0)' in compact
            blank_fallback = (",'')" in compact or ',"")' in compact)
            guarded = ('if(' in compact and guards_present and subtraction
                       and rounded and blank_fallback)
            if item['type'] != 'calculate' or not guarded:
                add_error(f"Row {item['excel_row']}: '{item['name']}' must be a "
                          f"guarded calculate subtracting ${{{start_name}}} from "
                          f"${{{end_name}}}, with string-length checks for both.")

        if canonical_survey_complete:
            for name in ('survey_start_elapsed_sec', 'survey_end_elapsed_sec'):
                item = fields[name]
                if not is_duration_checkpoint(item):
                    add_error(f"Row {item['excel_row']}: '{item['name']}' must be a "
                              f"calculate_here using once(duration()).")
            for name in ('survey_start_time', 'survey_end_time'):
                item = fields[name]
                if not is_clock_checkpoint(item, canonical=True):
                    add_error(f"Row {item['excel_row']}: '{item['name']}' must be a "
                              f"calculate_here using "
                              f"once(format-date-time(now(), "
                              f"'%Y-%m-%dT%H:%M:%S')).")
            validate_guarded_duration(
                fields['overall_duration_sec'], 'survey_start_elapsed_sec',
                'survey_end_elapsed_sec')
            positions = [fields[name]['position'] for name in canonical_survey_names]
            if positions != sorted(positions):
                add_error("Canonical survey timing fields are out of order. Expected "
                          + " then ".join(canonical_survey_names) + ".")
            nested = [name for name in canonical_survey_names
                      if fields[name]['group_path']]
            if nested:
                add_error("Survey-level timing fields must be outside all groups and "
                          "repeats. Affected fields: " + ", ".join(nested) + ".")
            relevant = [name for name in canonical_survey_names
                        if fields[name]['relevance']]
            if relevant:
                add_error("Survey-level timing fields must not have relevance. "
                          "Affected fields: " + ", ".join(relevant) + ".")
            if respondent_rows:
                first_question = respondent_rows[0]['position']
                last_question = respondent_rows[-1]['position']
                if max(fields[name]['position']
                       for name in canonical_survey_names[:2]) >= first_question:
                    add_error("Survey start timing fields must precede the first "
                              "respondent-input field.")
                if min(fields[name]['position']
                       for name in canonical_survey_names[2:]) <= last_question:
                    add_error("Survey end timing fields and overall duration must "
                              "follow the last respondent-input field.")

        overall_present = canonical_survey_complete
        if not overall_present and chosen_clocks['end'] is not None:
            end_position = chosen_clocks['end']['position']
            typed_before = [item for item in rows[:end_position] if item['type']]
            previous = typed_before[-1] if typed_before else None
            overall_present = bool(previous and is_duration_checkpoint(previous))
        suffixes = (
            'start_elapsed_sec', 'start_time', 'end_elapsed_sec',
            'end_time', 'duration_sec')
        section_parts = {}
        suffix_re = re.compile(
            r'^(.+)_(start_elapsed_sec|start_time|end_elapsed_sec|end_time|duration_sec)$',
            re.I)
        for item in rows:
            match = suffix_re.match(item['name'])
            if not match:
                continue
            slug, suffix = match.group(1).lower(), match.group(2).lower()
            if slug in {'survey', 'overall'} or item['name'].lower() == 'overall_duration_sec':
                continue
            section_parts.setdefault(slug, {})[suffix] = item

        canonical_sections = 0
        for slug, parts in sorted(section_parts.items()):
            strong_intent = (
                'start_elapsed_sec' in parts
                or 'end_elapsed_sec' in parts
                or any(parts[suffix]['type'] == 'calculate_here'
                       for suffix in ('start_time', 'end_time') if suffix in parts))
            if not strong_intent:
                continue
            missing = [f'{slug}_{suffix}' for suffix in suffixes
                       if suffix not in parts]
            if missing:
                add_error(f"Incomplete timing bundle for section '{slug}'. Missing: "
                          + ", ".join(missing) + ".")
                continue

            canonical_sections += 1
            for suffix in ('start_elapsed_sec', 'end_elapsed_sec'):
                item = parts[suffix]
                if not is_duration_checkpoint(item):
                    add_error(f"Row {item['excel_row']}: '{item['name']}' must be a "
                              f"calculate_here using once(duration()).")
            for suffix in ('start_time', 'end_time'):
                item = parts[suffix]
                if not is_clock_checkpoint(item, canonical=True):
                    add_error(f"Row {item['excel_row']}: '{item['name']}' must be a "
                              f"calculate_here using "
                              f"once(format-date-time(now(), "
                              f"'%Y-%m-%dT%H:%M:%S')).")
            validate_guarded_duration(parts['duration_sec'],
                                      f'{slug}_start_elapsed_sec',
                                      f'{slug}_end_elapsed_sec')
            positions = [parts[suffix]['position'] for suffix in suffixes]
            if positions != sorted(positions):
                add_error(f"Timing fields for section '{slug}' are out of order. "
                          "Start checkpoints must precede end checkpoints and duration.")
            relevances = {parts[suffix]['relevance'] for suffix in suffixes}
            if len(relevances) > 1:
                add_error(f"Timing fields for section '{slug}' have inconsistent "
                          "relevance. Put them in the same section group or give all "
                          "five fields identical relevance.")

            group_paths = {parts[suffix]['group_path'] for suffix in suffixes}
            if len(group_paths) > 1:
                add_error(f"Timing fields for section '{slug}' must share the same "
                          "group path. Put all five fields inside the section group.")
            else:
                section_path = next(iter(group_paths))
                latest_start = max(parts[suffix]['position'] for suffix in (
                    'start_elapsed_sec', 'start_time'))
                earliest_end = min(parts[suffix]['position'] for suffix in (
                    'end_elapsed_sec', 'end_time', 'duration_sec'))
                questions_between = [
                    item for item in respondent_rows
                    if latest_start < item['position'] < earliest_end]
                if not questions_between:
                    add_error(f"Timing fields for section '{slug}' must have at "
                              "least one respondent-input field between the start "
                              "and end boundaries.")
                if (section_path
                        and section_path[-1][1].lower() == slug):
                    section_questions = [
                        item for item in respondent_rows
                        if item['group_path'][:len(section_path)] == section_path]
                    if section_questions:
                        first_question = section_questions[0]['position']
                        last_question = section_questions[-1]['position']
                        if latest_start >= first_question or earliest_end <= last_question:
                            add_error(f"Timing fields for section '{slug}' must bracket "
                                      "its respondent-input fields: start checkpoints "
                                      "before the first question and end checkpoints "
                                      "and duration after the last question.")

        typed_rows = [item for item in rows if item['type']]
        previous_typed = {item['position']: typed_rows[i - 1] if i else None
                          for i, item in enumerate(typed_rows)}
        survey_overall_positions = set()
        for boundary in ('start', 'end'):
            chosen = chosen_clocks[boundary]
            if chosen is not None:
                previous = previous_typed.get(chosen['position'])
                if previous and is_duration_checkpoint(previous):
                    survey_overall_positions.add(previous['position'])

        legacy_parts = {}
        for item in typed_rows:
            name = item['name'].lower()
            if item['type'] != 'calculate_here':
                continue
            if name.startswith('duration_'):
                if item['position'] in survey_overall_positions:
                    continue
                legacy_parts.setdefault(name[len('duration_'):], {})['duration'] = item
            elif name.startswith('time_'):
                if is_survey_clock(name, 'start') or is_survey_clock(name, 'end'):
                    continue
                legacy_parts.setdefault(name[len('time_'):], {})['time'] = item

        legacy_pairs = 0
        valid_legacy_parts = []
        typed_positions = {item['position']: i for i, item in enumerate(typed_rows)}
        for slug, parts in sorted(legacy_parts.items()):
            missing = [part for part in ('duration', 'time') if part not in parts]
            if missing:
                add_error(f"Incomplete legacy timing pair '{slug}'. Missing "
                          + ", ".join(f'{part}_{slug}' for part in missing) + ".")
                continue
            duration_item = parts['duration']
            time_item = parts['time']
            valid = True
            if not is_duration_checkpoint(duration_item):
                add_error(f"Row {duration_item['excel_row']}: legacy checkpoint "
                          f"'{duration_item['name']}' must use once(duration()).")
                valid = False
            if not is_clock_checkpoint(time_item):
                add_error(f"Row {time_item['excel_row']}: legacy checkpoint "
                          f"'{time_item['name']}' must use "
                          "once(format-date-time(now(), ...)) with a real date and "
                          "time format.")
                valid = False
            duration_typed_position = typed_positions[duration_item['position']]
            time_typed_position = typed_positions[time_item['position']]
            if time_typed_position != duration_typed_position + 1:
                add_error(f"Legacy timing pair '{slug}' must be adjacent and ordered "
                          "duration first, then timestamp.")
                valid = False
            if duration_item['relevance'] != time_item['relevance']:
                add_error(f"Rows {duration_item['excel_row']}-{time_item['excel_row']}: "
                          f"legacy timing pair '{duration_item['name']}'/"
                          f"'{time_item['name']}' has inconsistent relevance.")
                valid = False
            matching_structures = structure_bounds.get(slug, [])
            if matching_structures:
                if not any(end_position < duration_item['position']
                           for _, end_position in matching_structures):
                    add_error(f"Legacy timing pair '{slug}' must appear after the "
                              "matching group or repeat ends.")
                    valid = False
            elif not any(item['position'] < duration_item['position']
                         for item in respondent_rows):
                add_error(f"Legacy timing pair '{slug}' appears before every "
                          "respondent-input field and cannot time a completed section.")
                valid = False
            if valid:
                legacy_pairs += 1
                valid_legacy_parts.append(parts)

        if not overall_present and chosen_clocks['end'] is not None:
            end_position = chosen_clocks['end']['position']
            for parts in valid_legacy_parts:
                duration_item = parts['duration']
                time_item = parts['time']
                if not (is_duration_checkpoint(duration_item)
                        and is_clock_checkpoint(time_item)
                        and time_item['position'] < end_position):
                    continue
                later_questions = [
                    item for item in respondent_rows
                    if time_item['position'] < item['position'] < end_position]
                if not later_questions:
                    overall_present = True
                    break

        if not overall_present:
            add_error("Missing overall duration. Use the canonical guarded "
                      "`overall_duration_sec`; an established K2 form may instead "
                      "use its final complete cumulative duration/timestamp pair, "
                      "provided no respondent-input fields follow it.")

        question_count = 0
        candidate_groups = 0
        group_stack = []
        for item in rows:
            if item['type'] == 'begin group':
                group_stack.append({'appearance': item['appearance'], 'questions': 0})
                continue
            if item['type'] == 'end group':
                if group_stack:
                    group = group_stack.pop()
                    if ('field-list' not in group['appearance']
                            and group['questions'] >= 3):
                        candidate_groups += 1
                continue
            if is_question(item):
                question_count += 1
                if group_stack:
                    group_stack[-1]['questions'] += 1

        timing_units = canonical_sections + legacy_pairs
        expected_units = 0
        if question_count:
            expected_units = max(
                1, (question_count + 39) // 40, (candidate_groups + 1) // 2)

        if question_count and timing_units == 0:
            add_error(f"No section timing was found for {question_count} question "
                      "fields. Add a canonical timing bundle to every substantive "
                      "section, or complete the established K2 checkpoint pairs.")
        elif timing_units < expected_units:
            add_warning(
                f"Sparse section timing coverage: found {timing_units} timing unit"
                f"{'s' if timing_units != 1 else ''}; expected at least "
                f"{expected_units} for {question_count} question fields and "
                f"{candidate_groups} substantive-group candidates. The heuristic "
                "uses one unit per 40 questions or one per two candidate groups, "
                "whichever is larger. Review the timing inventory; technical and "
                "field-list groups do not need their own timers.")

        if not self.errors and not self.warnings:
            print(f"  ✅ Survey boundaries and {timing_units} section timing "
                  f"unit{'s' if timing_units != 1 else ''} look complete")
        return not self.errors

    def _is_visible_field_type(self, norm_type):
        """True when a normalised type string is a field SurveyCTO shows on screen."""
        if not norm_type or norm_type in self.INVISIBLE_MULTIWORD_TYPES:
            return False
        prefix = norm_type.split()[0]
        return prefix in self.VISIBLE_FIELD_TYPES

    def check_audio_audits(self):
        """Validate `audio audit` anchors, placement and span.

        SurveyCTO does not resume an audio audit when a partially completed
        form is reopened with Edit Saved Form. A single whole-survey audit
        anchored at consent therefore records nothing for any interview the
        enumerator parks and resumes. Measured on three long AI Health baseline
        forms (3 Sep 2026): 23 to 45 percent of consenting interviews had no
        recording at all, and 97 to 100 percent of those showed more than five
        minutes of wall-clock time unexplained by active form time. Short forms
        of 4 to 16 minutes captured 100 percent. The fix is to split a long
        form into several field-anchored segments, one per section: see
        references/form-patterns.md, "Audio audits: segment long recordings".

        Also enforces the documented anchor restrictions: anchors must be
        visible fields named bare (no ${...}), neither the anchors nor the
        audit row itself may sit inside a repeat group.
        """
        print("\n=== Checking Audio Audits ===")
        if self.survey_df is None:
            return True
        type_col = self._col(self.survey_df, 'type')
        name_col = self._col(self.survey_df, 'name')
        app_col = self._col(self.survey_df, 'appearance')
        if type_col is None or name_col is None:
            return True

        def cell(value):
            if value is None or pd.isna(value):
                return ''
            return str(value).strip()

        rows = []
        fields = {}
        repeat_depth = 0
        for position, (idx, row) in enumerate(self.survey_df.iterrows()):
            norm_type = ' '.join(cell(row.get(type_col)).lower().split())
            if norm_type == 'end repeat' and repeat_depth:
                repeat_depth -= 1
            item = {
                'position': position,
                'excel_row': idx + 2,
                'type': norm_type,
                'name': cell(row.get(name_col)),
                'appearance': cell(row.get(app_col)) if app_col else '',
                'in_repeat': repeat_depth > 0,
                'visible': self._is_visible_field_type(norm_type),
            }
            if norm_type == 'begin repeat':
                repeat_depth += 1
            rows.append(item)
            if item['name']:
                fields.setdefault(item['name'].lower(), item)

        audits = [item for item in rows if item['type'] == 'audio audit']
        if not audits:
            print("  ℹ️  No audio audit field")
            return True

        ok = True

        def add_error(message):
            nonlocal ok
            ok = False
            print(f"  ❌ {message}")
            self.errors.append(message)

        def add_warning(message):
            print(f"  ⚠️  {message}")
            self.warnings.append(message)

        param_re = re.compile(r'^([A-Za-z]+)\s*=\s*(.*)$')
        time_value_re = re.compile(r'^\d+(\s*-\s*\d+)?$')
        reference_re = re.compile(r'^\$\{(.+?)\}$')

        for audit in audits:
            label = f"Row {audit['excel_row']}: audio audit '{audit['name']}'"
            if audit['in_repeat']:
                add_error(f"{label} sits inside a repeat group. An audio audit "
                          f"does not function inside a repeat. Move the row to "
                          f"the top level of the survey.")

            params = {}
            for chunk in audit['appearance'].split(';'):
                chunk = chunk.strip()
                if not chunk:
                    continue
                match = param_re.match(chunk)
                if match:
                    params[match.group(1).lower()] = match.group(2).strip()
            audit['params'] = params

            if '${' in audit['appearance']:
                add_error(f"{label} uses ${{...}} syntax in its appearance: "
                          f"{audit['appearance']!r}. Audio-audit anchors take the "
                          f"bare field name, e.g. s=audio_consent;d=gps.")

            anchors = {}
            field_anchor_keys = set()
            for key in ('s', 'd'):
                value = params.get(key)
                if value is None or time_value_re.match(value):
                    continue
                field_anchor_keys.add(key)
                match = reference_re.match(value)
                reference = match.group(1).strip() if match else value
                target = fields.get(reference.lower())
                if target is None:
                    add_error(f"{label} anchors {key}={value} on field "
                              f"'{reference}', which does not exist in the survey.")
                elif not target['visible']:
                    add_error(f"{label} anchors {key}={value} on '{reference}', "
                              f"a `{target['type']}` row. Anchors must be visible "
                              f"fields: calculate, calculate_here and group "
                              f"boundaries do not work.")
                elif target['in_repeat']:
                    add_error(f"{label} anchors {key}={value} on '{reference}', "
                              f"which sits inside a repeat group. Audio-audit "
                              f"anchors cannot be inside a repeat.")
                else:
                    anchors[key] = target
            audit['anchors'] = anchors
            audit['field_anchor_keys'] = field_anchor_keys

            if 'p' in params and field_anchor_keys:
                keys = ', '.join(f"{k}=" for k in ('s', 'd')
                                 if k in field_anchor_keys)
                add_warning(f"{label} combines p={params['p']} with a "
                            f"field-anchored {keys}. Mutual exclusion applies to "
                            f"time-based (`p=`) audits: at most one fires per "
                            f"submission, and a p=100 audit blocks every other "
                            f"time-based audit. Drop `p=` so the row is "
                            f"unambiguously question-based. Question-based audits "
                            f"are not mutually exclusive, so every segment fires.")

        visible_positions = [item['position'] for item in rows if item['visible']]
        if len(audits) == 1 and visible_positions:
            audit = audits[0]
            params = audit['params']
            anchors = audit['anchors']
            resume_note = (
                "SurveyCTO does not resume an audio audit after a form is "
                "reopened with Edit Saved Form, so an interview the enumerator "
                "parks and resumes records nothing. On three long AI Health "
                "baseline forms this lost 23 to 45 percent of interviews. Split "
                "the form into several field-anchored segments, one per section "
                "(see references/form-patterns.md, \"Audio audits: segment long "
                "recordings\").")
            if audit['field_anchor_keys'] <= set(anchors) and audit['field_anchor_keys']:
                start_position = anchors['s']['position'] if 's' in anchors else -1
                end_position = (anchors['d']['position'] if 'd' in anchors
                                else rows[-1]['position'])
                span = [p for p in visible_positions
                        if start_position <= p <= end_position]
                share = len(span) / len(visible_positions)
                if share > self.AUDIO_AUDIT_MAX_SPAN_SHARE:
                    add_warning(
                        f"Row {audit['excel_row']}: audio audit "
                        f"'{audit['name']}' is the form's only audio audit and "
                        f"spans {len(span)} of {len(visible_positions)} visible "
                        f"fields ({share:.0%}). {resume_note}")
            elif not audit['field_anchor_keys']:
                duration = params.get('d', '')
                if (re.fullmatch(r'\d+', duration)
                        and int(duration) > self.AUDIO_AUDIT_MAX_DURATION_SEC):
                    add_warning(
                        f"Row {audit['excel_row']}: audio audit "
                        f"'{audit['name']}' is the form's only audio audit and "
                        f"records d={duration} seconds in one stretch. "
                        f"{resume_note}")

        if ok:
            print(f"  ✅ {len(audits)} audio audit row(s) have valid anchors "
                  f"and placement")
        return ok

    def _col(self, df, name):
        """Case-insensitive column lookup; also tolerates space/underscore."""
        want = name.lower().replace('_', ' ')
        for c in df.columns:
            if str(c).lower().replace('_', ' ') == want:
                return c
        return None

    def run_all_checks(self):
        """Run all validation checks."""
        print(f"\n{'='*60}")
        print(f"SurveyCTO Form Checker")
        print(f"File: {self.file_path}")
        print(f"{'='*60}")

        if not self.load_form():
            return False

        results = []
        results.append(self.check_required_columns())
        results.append(self.check_blank_names())
        results.append(self.check_duplicate_names())
        results.append(self.check_empty_groups())
        results.append(self.check_expression_syntax())
        results.append(self.check_upload_parser_blockers())
        results.append(self.check_field_references())
        results.append(self.check_choices_field_references())
        results.append(self.check_choice_lists())
        results.append(self.check_dynamic_choice_labels())
        results.append(self.check_end_metadata_arithmetic())
        results.append(self.check_timing_instrumentation())
        results.append(self.check_audio_audits())
        results.append(self.check_other_specify_fields())
        results.append(self.check_select_multiple_other())
        results.append(self.check_select_multiple_exclusive())
        results.append(self.check_impossible_literal_values())
        results.append(self.check_required_fields())
        results.append(self.check_typos())
        results.append(self.check_missing_constraint_messages())
        results.append(self.check_integer_constraints())
        results.append(self.check_numeric_refuse_option())
        results.append(self.check_calculate_fields())
        results.append(self.check_hindi_translations())
        results.append(self.check_naming_conventions())
        results.append(self.check_conditional_formatting())
        results.append(self.check_formatting_preserved())
        results.append(self.check_public_key_format())
        results.append(self.check_version_formula())

        # Print summary
        print(f"\n{'='*60}")
        print("SUMMARY")
        print(f"{'='*60}")

        if self.errors:
            print(f"\n❌ {len(self.errors)} ERROR(S) FOUND:")
            for i, error in enumerate(self.errors, 1):
                print(f"{i}. {error.strip()}")

        if self.warnings:
            print(f"\n⚠️  {len(self.warnings)} WARNING(S):")
            for i, warning in enumerate(self.warnings, 1):
                print(f"{i}. {warning.strip()}")

        if not self.errors and not self.warnings:
            print("\n✅ All checks passed! Form looks good.")
            return True
        elif not self.errors:
            print("\n✅ No errors found (but there are warnings to review)")
            return True
        else:
            print("\n❌ Form has errors that need to be fixed")
            return False


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(description="Validate an XLSForm questionnaire")
    parser.add_argument('file_path', nargs='?', default='ai_health_pilot_baseline.xlsx')
    parser.add_argument('--platform', choices=('surveycto', 'odk', 'kobo'),
                        default='surveycto',
                        help='target platform; SurveyCTO enables calculate_here timing checks')
    args = parser.parse_args()
    file_path = args.file_path
    if len(sys.argv) == 1:
        print(f"No file specified, using default: {file_path}")

    if not Path(file_path).exists():
        print(f"Error: File not found: {file_path}")
        sys.exit(1)

    checker = SurveyCTOChecker(file_path, platform=args.platform)
    success = checker.run_all_checks()

    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""
Convert SurveyCTO/XLSForm questionnaires to human-readable text format.

Usage:
    python surveycto_to_txt.py input.xlsx [output.txt]
    python surveycto_to_txt.py input.xlsx --no-names --no-relevance
    python surveycto_to_txt.py input.xlsx --language Hindi
    python surveycto_to_txt.py input.xlsx --language all

If output is not specified, creates input_questions.txt in the same directory.
"""

import sys
import os
import re
import argparse
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def find_column_index(headers, column_name):
    """Find the index of a column by name (case-insensitive)."""
    for i, header in enumerate(headers):
        if header and str(header).lower() == column_name.lower():
            return i
    return None


def find_label_column_index(headers):
    """Find the best label column - prefers 'label' over language-specific ones."""
    # First look for exact 'label'
    for i, header in enumerate(headers):
        if header and str(header).lower() == 'label':
            return i

    # Then look for any label column (label::English, label:Hindi, etc.)
    for i, header in enumerate(headers):
        if header and str(header).lower().startswith('label'):
            # Skip 'label:data' columns
            if 'data' not in str(header).lower():
                return i

    return None


def get_all_label_columns(headers):
    """Find all label columns and return list of (index, language_name) tuples.

    Returns columns in order found. The bare 'label' column gets language name
    from its position (first language), language-specific columns use their
    suffix (e.g. 'label:Hindi' -> 'Hindi', 'label::English' -> 'English').
    """
    columns = []
    for i, header in enumerate(headers):
        if header and str(header).lower().startswith('label'):
            h = str(header)
            h_lower = h.lower()
            # Skip 'label:data' columns
            if 'data' in h_lower:
                continue
            # Extract language name
            if h_lower == 'label':
                columns.append((i, None))  # default/first language
            elif '::' in h:
                lang = h.split('::', 1)[1].strip()
                columns.append((i, lang))
            elif ':' in h:
                lang = h.split(':', 1)[1].strip()
                columns.append((i, lang))
    return columns


def find_label_columns_for_language(headers, language):
    """Find label column indices based on language selection.

    Args:
        headers: List of column headers
        language: None for default (first column), a language name, or 'all'

    Returns:
        List of (index, language_name) tuples
    """
    all_cols = get_all_label_columns(headers)
    if not all_cols:
        return []

    if language is None:
        # Default: return just the first label column
        return [all_cols[0]]

    if language.lower() == 'all':
        return all_cols

    # Match specific language (case-insensitive)
    for idx, lang_name in all_cols:
        if lang_name and lang_name.lower() == language.lower():
            return [(idx, lang_name)]

    # Also check if the bare 'label' column is the requested language
    # (e.g., if 'label' is English and user asks for 'english')
    # Fall back to returning nothing if no match
    return []


def load_choices(wb, language=None):
    """Load choices from the choices sheet into a dictionary.

    Args:
        wb: openpyxl workbook
        language: None for default, language name, or 'all'
    """
    choices_dict = {}

    if 'choices' not in wb.sheetnames:
        return choices_dict

    ws = wb['choices']
    headers = [cell.value for cell in ws[1]]

    # Find required columns
    list_name_idx = find_column_index(headers, 'list_name')
    name_idx = find_column_index(headers, 'name')
    if name_idx is None:
        name_idx = find_column_index(headers, 'value')

    if list_name_idx is None or name_idx is None:
        return choices_dict

    # Find label columns based on language selection
    label_cols = find_label_columns_for_language(headers, language)
    if not label_cols:
        # Fallback to default
        label_idx = find_label_column_index(headers)
        if label_idx is None:
            return choices_dict
        label_cols = [(label_idx, None)]

    # Build dictionary of choices by list_name
    for row in ws.iter_rows(min_row=2, values_only=True):
        list_name = row[list_name_idx] if list_name_idx < len(row) else None
        name = row[name_idx] if name_idx < len(row) else None

        if list_name and name is not None:
            # Collect labels from all selected columns
            labels = []
            for col_idx, lang_name in label_cols:
                val = row[col_idx] if col_idx < len(row) else None
                if val:
                    labels.append(str(val))
            combined_label = ' / '.join(labels) if labels else str(name)

            if list_name not in choices_dict:
                choices_dict[list_name] = []
            choices_dict[list_name].append({
                'name': str(name),
                'label': combined_label
            })

    return choices_dict


def strip_html_tags(text):
    """Remove HTML tags from text."""
    if text is None:
        return text
    # Remove HTML tags
    clean = re.sub(r'<[^>]+>', '', str(text))
    # Clean up extra whitespace
    clean = re.sub(r'\s+', ' ', clean).strip()
    return clean


def import_survey_questions(survey_file_path, language=None):
    """
    Import survey questions from an XLSForm Excel file.

    Args:
        survey_file_path: Path to the XLSForm Excel file
        language: None for default (first label column), language name, or 'all'

    Returns:
        Tuple of (questions list, choices dict)
    """
    wb = load_workbook(survey_file_path, read_only=True, data_only=True)

    if 'survey' not in wb.sheetnames:
        raise ValueError("Excel file must have a 'survey' sheet")

    # Load choices first (need to reopen without read_only for choices)
    wb.close()
    wb = load_workbook(survey_file_path, data_only=True)
    choices_dict = load_choices(wb, language=language)

    ws = wb['survey']

    # Get headers from first row
    headers = [cell.value for cell in ws[1]]

    # Find required columns
    name_idx = find_column_index(headers, 'name')
    type_idx = find_column_index(headers, 'type')
    relevance_idx = find_column_index(headers, 'relevance')
    calculation_idx = find_column_index(headers, 'calculation')
    disabled_idx = find_column_index(headers, 'disabled')

    # Find label columns based on language selection
    label_cols = find_label_columns_for_language(headers, language)
    if not label_cols:
        # Fallback to default
        fallback_idx = find_label_column_index(headers)
        if fallback_idx is not None:
            label_cols = [(fallback_idx, None)]

    if name_idx is None:
        raise ValueError("Survey sheet must have a 'name' column")
    if type_idx is None:
        raise ValueError("Survey sheet must have a 'type' column")
    if not label_cols:
        raise ValueError("Survey sheet must have a 'label' column")

    # Get max column index we need
    all_indices = [name_idx, type_idx, relevance_idx, calculation_idx, disabled_idx]
    all_indices.extend([idx for idx, _ in label_cols])
    indices = [i for i in all_indices if i is not None]
    max_idx = max(indices)

    questions = []

    # Process data rows (skip header)
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Ensure row has enough columns
        if len(row) <= max_idx:
            row = list(row) + [None] * (max_idx + 1 - len(row))

        name = row[name_idx]
        type_val = row[type_idx]
        relevance = row[relevance_idx] if relevance_idx is not None else None
        calculation = row[calculation_idx] if calculation_idx is not None else None
        disabled = row[disabled_idx] if disabled_idx is not None else None

        # Combine labels from selected columns
        labels = []
        for col_idx, lang_name in label_cols:
            val = row[col_idx]
            if val:
                labels.append(str(val))
        label = ' / '.join(labels) if labels else None

        # Skip disabled fields
        if disabled and str(disabled).lower() == 'yes':
            continue

        # Skip rows without names
        if not name:
            continue

        # Skip duration-related calculations
        if type_val == 'calculate' and calculation and 'duration' in str(calculation).lower():
            continue

        # Include fields with label OR calculate fields with calculation
        if not (label or (type_val == 'calculate' and calculation)):
            continue

        # Parse type into type and choices
        q_type = None
        choice_list = None
        if type_val:
            parts = str(type_val).split(' ', 1)
            q_type = parts[0]
            choice_list = parts[1] if len(parts) > 1 else None

        # Get choices if this is a select question
        choices = None
        if q_type in ('select_one', 'select_multiple') and choice_list:
            choices = choices_dict.get(choice_list, [])

        questions.append({
            'name': name,
            'type': q_type,
            'type_full': type_val,
            'choice_list': choice_list,
            'label': label,
            'relevance': relevance,
            'calculation': calculation,
            'choices': choices
        })

    wb.close()
    return questions


def convert_survey_to_txt(
    input_file,
    output_file=None,
    include_names=True,
    include_relevance=True,
    include_choices=True,
    strip_html=True,
    language=None
):
    """
    Convert an XLSForm survey to a human-readable text file.

    Args:
        input_file: Path to the XLSForm Excel file
        output_file: Path for output text file (optional)
        include_names: Whether to include variable names in brackets
        include_relevance: Whether to include relevance conditions
        include_choices: Whether to include choice options
        strip_html: Whether to remove HTML tags from labels
        language: None for default (first label column), language name, or 'all'

    Returns:
        Tuple of (output path, question count)
    """
    # Determine output file path
    if output_file is None:
        base_name = Path(input_file).stem
        output_dir = Path(input_file).parent
        # Add language suffix to output filename
        if language and language.lower() != 'all':
            output_file = str(output_dir / f"{base_name}_questions_{language.lower()}.txt")
        elif language and language.lower() == 'all':
            output_file = str(output_dir / f"{base_name}_questions_all_languages.txt")
        else:
            output_file = str(output_dir / f"{base_name}_questions.txt")

    # Import survey questions
    questions = import_survey_questions(input_file, language=language)

    # Get base filename for header
    base_name = Path(input_file).stem

    # Build output content
    lines = []

    # Header
    header = f"Survey Questions from: {base_name}"
    lines.append(header)
    lines.append("=" * len(header))
    lines.append("")

    # Track question count
    question_count = len(questions)

    # Process each question
    for q in questions:
        label = q['label']
        q_type = q['type'] or ''
        choice_list = q['choice_list'] or ''

        # Check if this is a begin group
        if q_type == 'begin' and choice_list == 'group':
            # Format as subheading with blank line before
            if strip_html and label:
                label = strip_html_tags(label)
            lines.append("")
            lines.append(f"## {label}")

        elif q_type == 'begin' and choice_list == 'repeat':
            # Format repeat groups similarly
            if strip_html and label:
                label = strip_html_tags(label)
            lines.append("")
            lines.append(f"### [REPEAT] {label}")

        elif q_type == 'end':
            # Skip end group/repeat markers
            continue

        elif q_type == 'note' and q['name'] and q['name'].endswith('_header'):
            # Notes ending with _header are section headers
            if strip_html and label:
                label = strip_html_tags(label)
            lines.append("")
            lines.append(f"## {label}")

        elif q_type == 'note':
            # Regular notes without variable names
            if strip_html and label:
                label = strip_html_tags(label)
            if label:
                lines.append(f"  {label.rstrip()}")

        elif q_type == 'calculate':
            # Format calculation fields
            if include_names:
                lines.append(f"• [{q['name']}] (calculate): {q['calculation']}")
            else:
                lines.append(f"• (calculate): {q['calculation']}")

        else:
            # Format as regular bullet point
            if strip_html and label:
                label = strip_html_tags(label)

            line_parts = []

            # Add name
            if include_names and q['name']:
                line_parts.append(f"• [{q['name']}]")
            else:
                line_parts.append("•")

            # Add relevance condition if present
            if include_relevance and q['relevance']:
                line_parts.append(f"(If: {q['relevance']})")

            # Build the line
            line = " ".join(line_parts)
            if label:
                line += f": {label.rstrip()}"

            lines.append(line)

            # Add choices for select questions
            if include_choices and q['choices']:
                for choice in q['choices']:
                    choice_label = choice['label']
                    if choice_label:
                        # Strip newlines from choice labels
                        clean_choice = str(choice_label).replace('\n', ' ').strip()
                        lines.append(f"    - {clean_choice}")

    # Write to file
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))

    return output_file, question_count


def main():
    parser = argparse.ArgumentParser(
        description='Convert SurveyCTO/XLSForm questionnaires to text format',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    python surveycto_to_txt.py survey.xlsx
    python surveycto_to_txt.py survey.xlsx output.txt
    python surveycto_to_txt.py survey.xlsx --no-names
    python surveycto_to_txt.py survey.xlsx --no-relevance
    python surveycto_to_txt.py survey.xlsx --no-choices
    python surveycto_to_txt.py survey.xlsx --keep-html
    python surveycto_to_txt.py survey.xlsx --language Hindi
    python surveycto_to_txt.py survey.xlsx --language all
        """
    )

    parser.add_argument(
        'input',
        help='Input XLSForm Excel file (.xlsx)'
    )

    parser.add_argument(
        'output',
        nargs='?',
        default=None,
        help='Output text file (default: input_questions.txt)'
    )

    parser.add_argument(
        '--no-names',
        action='store_true',
        help='Exclude variable names from output'
    )

    parser.add_argument(
        '--no-relevance',
        action='store_true',
        help='Exclude relevance conditions from output'
    )

    parser.add_argument(
        '--no-choices',
        action='store_true',
        help='Exclude choice options from output'
    )

    parser.add_argument(
        '--keep-html',
        action='store_true',
        help='Keep HTML tags in labels (default: strip them)'
    )

    parser.add_argument(
        '--language',
        default=None,
        help='Language to export: a language name (e.g. "Hindi"), or "all" for all languages side by side. Default: first/primary language only.'
    )

    args = parser.parse_args()

    # Validate input file
    if not os.path.exists(args.input):
        print(f"Error: Input file not found: {args.input}", file=sys.stderr)
        sys.exit(1)

    if not args.input.endswith(('.xlsx', '.xls')):
        print(f"Error: Input file must be an Excel file (.xlsx or .xls)", file=sys.stderr)
        sys.exit(1)

    try:
        output_path, question_count = convert_survey_to_txt(
            args.input,
            args.output,
            include_names=not args.no_names,
            include_relevance=not args.no_relevance,
            include_choices=not args.no_choices,
            strip_html=not args.keep_html,
            language=args.language
        )
        print(f"✓ Extracted {question_count} questions")
        print(f"✓ Saved to: {output_path}")
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()

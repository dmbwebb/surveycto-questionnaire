"""Validate a SurveyCTO server dataset definition (.xml) before upload.

This helper supports the dataset-authoring workflow described in
``references/dataset-validation.md``. SurveyCTO's interactive console enforces a
set of rules when a dataset is created or edited, and the server rejects an
uploaded definition (or a publishing configuration) that breaks them. An agent
authoring dataset XML by hand cannot see those rules, so it tends to produce
definitions that look plausible but fail on upload or silently misbehave once a
form starts publishing into them.

This script re-implements those rules so the agent can validate and self-correct
locally, with no server round-trip. It works in two layers:

  1. Dataset-only checks (standard library ``xml.etree`` only): element order and
     required children, enumerations, id/title/idFormatOptions/caseManagementOptions
     value rules, fieldNames rules, and the field-map / data-link rules.
  2. Form cross-reference checks (``openpyxl``): when the referenced form
     ``.xlsx`` files are supplied, each form is parsed into a field list of
     ``{name, type, repeated}`` matching how the server builds its publishing
     field list, and the field map is checked against the real fields (every
     mapped field exists, joining field exists and is inside a repeat group for
     long format, repeated fields carry the ``*`` suffix, etc.).

Usage::

    python validate_dataset.py my_dataset.xml
    python validate_dataset.py my_dataset.xml --form household.xlsx --form roster.xlsx
    python validate_dataset.py my_dataset.xml --json

Findings are reported in four tiers:

  - ``error``        the server (or its XSD) rejects this on upload, or a
                     publishing rule that breaks data collection. Must be fixed.
  - ``warning``      the server accepts the upload but behavior is degraded or a
                     console-level publishing rule will reject the config later.
  - ``recommendation`` best practice / console convention; not enforced.
  - ``cannot_verify`` a real rule that needs a live server (form deployment,
                     uniqueness in stored data, license, id collisions); surfaced
                     so a clean result is not mistaken for a guarantee.

The process exits non-zero when any ``error`` is present.

Standard library only, plus ``openpyxl`` (already a skill dependency) and only
when ``--form`` files are supplied. No network access; nothing is uploaded.

This script does NOT validate the XLSForm itself (use the XLSForm tooling for
that), and it cannot replace a real upload: the rules below are derived from the
SurveyCTO server source and can drift if that source changes. See the rule
manifest in the module body for the source citations behind each rule.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import zipfile
from collections import Counter
from pathlib import Path
from typing import Optional

# A parsed field-map entry: (formField, datasetField, updateLogicAction). The
# form and dataset fields are strings or None (a non-string JSON value is coerced
# to None and reported as a shape error); the action is always a string.
FieldMapEntry = tuple[Optional[str], Optional[str], str]

# ---------------------------------------------------------------------------
# Rule manifest (source citations, for re-derivation on drift)
#
# Derived from the SurveyCTO server source as of 2026-06. Key sources:
#   - dataset.xsd: scto-commons-utils .../datasets/xml/dataset.xsd
#       element order, required children, enumerations.
#   - DatasetValidationUtils.java: scto-server-modules .../dataset/utils/
#       idFormatOptions (prefix/suffix alphanumeric <=10, numberOfDigits 4..8),
#       caseManagementOptions (displayMode, table-view must include id).
#   - DatasetServiceImpl.java: scto-server-modules .../dataset/service/
#       id/title/type rules, joiningField-in-fieldMap on definition import,
#       uniqueRecordField forced to "id" for CASES/ENUMERATORS.
#   - DatasetManagerImpl.java: scto-server .../datasets/manager/
#       duplicate field mapping, joiningField REPLACE + maps-to-uniqueRecordField,
#       standard column sets, max field-name length, reserved "rowId".
#   - XFormManagerImpl.java: scto-server .../forms/manager/
#       long-format publishing requirements and the publishing field list.
#       validateLongFormatPublishingRequirements calls FieldInfoUtils.asBase on the
#       relevance field with no null guard, so a long-format link missing the
#       <relevanceField> element (deserialized to null) NPEs on the next form upload.
#   - DataUtilsLongFormat.java / DataUtilsWideFormat.java: scto-commons-utils
#       .../datasets/: long format strips the '*' from both sides (removeWildcards),
#       so the datasetField carries no '*'; wide format expands a repeat into
#       numbered columns only when BOTH sides carry '*'.
#   - DatasetUtils.java / ServerConstants.java: standard column constants.
# Ticket coverage: SCTO-15028 (enumerator columns), SCTO-15073 (joiningField in
# field map), SCTO-15074 (same field mapped twice), SCTO-15224 (long-format
# relevanceField element required; '*' on the dataset field by format).
# ---------------------------------------------------------------------------

# Element order enforced by the XSD xs:sequence for <definition> and <dataLink>.
DEFINITION_ORDER = [
    "id", "title", "datasetType", "fieldNames", "formLinks", "dataLinks",
    "caseManagementOptions", "idFormatOptions", "discriminator",
    "uniqueRecordField", "allowOfflineUpdates",
]
DEFINITION_REQUIRED = ["id", "title", "datasetType"]

DATALINK_ORDER = [
    "dataLinkClass", "dataLinkType", "dataLinkState", "dataLinkFormat",
    "linkObjectId", "fieldMap", "joiningField", "relevanceField",
    "isAutoConfigured", "publishPartialData",
]
DATALINK_REQUIRED = ["dataLinkClass", "dataLinkType", "linkObjectId"]

# xs:all blocks: order-independent, but only these children are allowed and each
# may appear at most once.
CASE_MGMT_REQUIRED = ["displayMode", "showFinalizedSentWhenTree", "showColumnsWhenTable"]
CASE_MGMT_ALLOWED = {"displayMode", "showFinalizedSentWhenTree", "showColumnsWhenTable",
                     "otherUserCode", "entryMode", "enumeratorDatasetId"}
ID_FORMAT_ALLOWED = {"prefix", "suffix", "numberOfDigits", "allowCapitalLetters"}

DATASET_TYPES = {"SERVER", "CLIENT", "REPORT"}
DISCRIMINATORS = {"CASES", "ENUMERATORS", "DATA"}
DATALINK_CLASSES = {"FORM", "FUSION_TABLE", "SPREADSHEET"}
DATALINK_TYPES = {"INCOMING", "OUTGOING"}
DATALINK_STATES = {"ENABLED", "DISABLED"}
ENTRY_MODES = {"LIST", "ENTRY", "SCAN"}
UPDATE_LOGIC_ACTIONS = {"REPLACE", "ADD_TO_NUMERIC_VALUE", "CONCATENATE_TO_TEXT"}

DATALINK_FORMAT_WIDE = 0
DATALINK_FORMAT_LONG = 1

# Standard column sets the console creates. id/name (enum) and id/label/formids
# (cases) are functionally required; the rest are conventional.
ENUM_REQUIRED_COLUMNS = ["id", "name"]
CASES_REQUIRED_COLUMNS = ["id", "label", "formids"]
CASES_STANDARD_COLUMNS = ["id", "label", "formids", "users", "roles", "sortby", "enumerators"]

RESERVED_FIELD_NAMES = {"rowId"}
MAX_FIELD_NAME_LENGTH = 60

# Metadata fields the server UNCONDITIONALLY appends to a form's incoming
# publishing field list (FormUtils.addMetadataFields). They are valid field-map
# sources even though they are not rows in the survey sheet.
METADATA_FIELDS = [
    ("SubmissionDate", "datetime"),
    ("formdef_version", "text"),
    ("review_quality", "text"),
    ("KEY", "text"),
]
# Fields available only in some publishing contexts or only when the form itself
# declares them: formdef_id appears in outgoing-publishing summaries, not the
# incoming form-to-dataset feed; review_status only for reviewed submissions;
# instanceID/instanceName are meta rows a form may or may not define. When a
# field map names one of these and it is not found in the supplied form, warn
# rather than hard-error, since availability cannot be determined offline.
CONDITIONAL_META_FIELD_NAMES = {"formdef_id", "review_status", "instanceID", "instanceName"}

# XLSForm survey `type` leading token -> publishing field type string, mirroring
# the server's discoverElementType / dataType mapping. Notes are typed 'note' (not
# publishable); group and repeat containers are structural, not fields.
XLSFORM_TYPE_MAP = {
    "text": "text",
    "integer": "integer",
    "decimal": "decimal",
    "range": "integer",
    "date": "date",
    "time": "time",
    "datetime": "datetime",
    "dateTime": "datetime",
    "select_one": "select_one",
    "select_multiple": "select_multiple",
    "select_one_from_file": "select_one",
    "select_multiple_from_file": "select_multiple",
    "rank": "text",
    "calculate": "text",
    "geopoint": "geopoint",
    "geoshape": "geoshape",
    "geotrace": "geotrace",
    "barcode": "barcode",
    "image": "image",
    "audio": "audio",
    "background-audio": "audio",
    "video": "video",
    "file": "file",
    "acknowledge": "text",
    "hidden": "text",
    "username": "text",
    "phonenumber": "text",
    "deviceid": "text",
    "subscriberid": "text",
    "simserial": "text",
    "start": "datetime",
    "end": "datetime",
    "today": "date",
    "audit": "binary",
    "comments": "text",
    "sensor_statistic": "text",
    "sensor_stream": "text",
}


# ---------------------------------------------------------------------------
# Findings
# ---------------------------------------------------------------------------

ERROR = "error"
WARNING = "warning"
RECOMMENDATION = "recommendation"
CANNOT_VERIFY = "cannot_verify"

_SEVERITY_ORDER = {ERROR: 0, WARNING: 1, RECOMMENDATION: 2, CANNOT_VERIFY: 3}


class Finding:
    def __init__(self, severity: str, rule: str, message: str,
                 location: str = "", fix: str = "") -> None:
        self.severity = severity
        self.rule = rule
        self.message = message
        self.location = location
        self.fix = fix

    def as_dict(self) -> dict:
        return {
            "severity": self.severity,
            "rule": self.rule,
            "message": self.message,
            "location": self.location,
            "fix": self.fix,
        }


class Report:
    def __init__(self) -> None:
        self.findings: list[Finding] = []

    def add(self, severity: str, rule: str, message: str,
            location: str = "", fix: str = "") -> None:
        self.findings.append(Finding(severity, rule, message, location, fix))

    def error(self, rule: str, message: str, location: str = "", fix: str = "") -> None:
        self.add(ERROR, rule, message, location, fix)

    def warning(self, rule: str, message: str, location: str = "", fix: str = "") -> None:
        self.add(WARNING, rule, message, location, fix)

    def recommend(self, rule: str, message: str, location: str = "", fix: str = "") -> None:
        self.add(RECOMMENDATION, rule, message, location, fix)

    def cannot_verify(self, rule: str, message: str, location: str = "", fix: str = "") -> None:
        self.add(CANNOT_VERIFY, rule, message, location, fix)

    @property
    def has_errors(self) -> bool:
        return any(f.severity == ERROR for f in self.findings)

    def counts(self) -> dict[str, int]:
        out = {ERROR: 0, WARNING: 0, RECOMMENDATION: 0, CANNOT_VERIFY: 0}
        for f in self.findings:
            out[f.severity] = out.get(f.severity, 0) + 1
        return out


# ---------------------------------------------------------------------------
# Form field extraction (openpyxl)
# ---------------------------------------------------------------------------

class FormField:
    def __init__(self, name: str, ftype: str, repeat_path: tuple = (), metadata: bool = False) -> None:
        self.name = name
        self.type = ftype
        # Names of the repeat groups enclosing this field, outermost first. A
        # field is repeated when this is non-empty; the path is what long-format
        # scope validation compares against the joining field's repeat path.
        self.repeat_path = tuple(repeat_path)
        self.metadata = metadata

    @property
    def repeated(self) -> bool:
        return bool(self.repeat_path)

    def as_dict(self) -> dict:
        return {"name": self.name, "type": self.type,
                "repeated": self.repeated, "metadata": self.metadata}


class FormInfo:
    """A parsed form: its publishable fields plus its declared form_id (from the
    settings sheet), used to match a dataLink's linkObjectId to the real form."""

    def __init__(self, fields: list, form_id: Optional[str] = None) -> None:
        self.fields = fields
        self.form_id = form_id


def _xlsform_type(raw_type: str) -> str:
    token = raw_type.strip().split()[0] if raw_type.strip() else ""
    # `select_one listname`, `select_multiple listname`, and the *_from_file
    # variants collapse on their leading token.
    if token in ("select_one", "select_multiple"):
        return XLSFORM_TYPE_MAP[token]
    return XLSFORM_TYPE_MAP.get(token, XLSFORM_TYPE_MAP.get(raw_type.strip(), "text"))


def extract_form_fields(xlsx_path: str) -> list[FormField]:
    """Parse an XLSForm survey sheet into the field list the server publishes.

    Mirrors the server's publishing field list: group and repeat containers are
    not fields, notes are retained but typed 'note' (not publishable), every other named row is a field, a field
    is ``repeated`` when any ancestor row is a ``begin repeat`` (nested groups do
    not make a field repeated), and the standard metadata fields are appended.
    Field names are the bare leaf ``name`` value.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - exercised via hard import in tests
        raise RuntimeError(
            "openpyxl is required to cross-check field maps against form files. "
            "Install it (pip install openpyxl) or run without --form."
        ) from exc

    _guard_xlsx_zip_bomb(xlsx_path)
    wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    try:
        if "survey" not in wb.sheetnames:
            raise ValueError(f"{xlsx_path}: no 'survey' worksheet; not an XLSForm.")
        ws = wb["survey"]

        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            return _with_metadata([])

        header_map: dict[str, int] = {}
        for idx, cell in enumerate(header):
            if cell is None:
                continue
            key = str(cell).strip().lower()
            # First occurrence wins; ignore translated/duplicate columns.
            header_map.setdefault(key, idx)

        if "type" not in header_map or "name" not in header_map:
            raise ValueError(f"{xlsx_path}: survey sheet missing 'type' or 'name' column.")
        type_idx = header_map["type"]
        name_idx = header_map["name"]

        fields: list[FormField] = []
        # Stack of (kind, name) for the open containers above the current row,
        # where kind is "group" or "repeat".
        stack: list[tuple] = []

        for row in rows:
            raw_type = _cell(row, type_idx)
            name = _cell(row, name_idx)
            token = raw_type.strip().lower()

            if token in ("begin group", "begin_group"):
                stack.append(("group", name.strip()))
                continue
            if token in ("begin repeat", "begin_repeat"):
                stack.append(("repeat", name.strip()))
                continue
            if token in ("end group", "end_group", "end repeat", "end_repeat"):
                if stack:
                    stack.pop()
                continue
            if not raw_type.strip():
                continue
            if not name.strip():
                continue

            repeat_path = tuple(n for kind, n in stack if kind == "repeat")
            # Notes are retained (typed "note") rather than dropped: they are not
            # offered as publishable fields, but the server keeps them in the
            # element tree, so a field map that names one should warn (publishes
            # nothing) rather than report a missing field.
            ftype = "note" if token == "note" else _xlsform_type(raw_type)
            fields.append(FormField(name.strip(), ftype, repeat_path, metadata=False))
    finally:
        wb.close()
    return _with_metadata(fields)


def load_form(xlsx_path: str) -> FormInfo:
    """Parse a form into a FormInfo (publishable fields plus its declared
    form_id from the settings sheet)."""
    fields = extract_form_fields(xlsx_path)
    return FormInfo(fields, _read_form_id(xlsx_path))


def _read_form_id(xlsx_path: str) -> Optional[str]:
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover
        return None
    try:
        wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001 - form_id is best-effort
        return None
    try:
        if "settings" not in wb.sheetnames:
            return None
        rows = wb["settings"].iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            return None
        cols = {str(c).strip().lower(): i for i, c in enumerate(header) if c is not None}
        if "form_id" not in cols:
            return None
        idx = cols["form_id"]
        for row in rows:
            if idx < len(row) and row[idx] not in (None, ""):
                return str(row[idx]).strip()
        return None
    finally:
        wb.close()


# A single XLSForm worksheet entry should never legitimately decompress to this
# much; a member larger than this is treated as a zip bomb and refused.
MAX_XLSX_MEMBER_BYTES = 100 * 1024 * 1024


def _guard_xlsx_zip_bomb(xlsx_path: str) -> None:
    try:
        with zipfile.ZipFile(xlsx_path) as zf:
            for info in zf.infolist():
                if info.file_size > MAX_XLSX_MEMBER_BYTES:
                    raise ValueError(
                        f"{xlsx_path}: entry {info.filename!r} decompresses to "
                        f"{info.file_size:,} bytes; refusing to load (possible zip bomb)."
                    )
    except zipfile.BadZipFile as exc:
        raise ValueError(f"{xlsx_path}: not a valid .xlsx (zip) file.") from exc


def _with_metadata(fields: list[FormField]) -> list[FormField]:
    # The server treats the four metadata names as always-valid by membership, so
    # append them unconditionally; appended last, they win in a by-name lookup
    # even when a survey field shares the name.
    meta = [FormField(n, t, (), metadata=True) for n, t in METADATA_FIELDS]
    return fields + meta


def _cell(row: tuple, idx: int) -> str:
    if idx >= len(row):
        return ""
    val = row[idx]
    return "" if val is None else str(val)


# ---------------------------------------------------------------------------
# Dataset XML parsing
# ---------------------------------------------------------------------------

class DataLink:
    def __init__(self, index: int) -> None:
        self.index = index
        self.children: list[str] = []
        self.link_class: Optional[str] = None
        self.link_type: Optional[str] = None
        self.link_state: Optional[str] = None
        self.link_format: Optional[str] = None
        self.link_object_id: Optional[str] = None
        self.field_map_raw: Optional[str] = None
        self.joining_field: Optional[str] = None
        self.relevance_field: Optional[str] = None
        self.is_auto_configured_raw: Optional[str] = None
        self.publish_partial_data_raw: Optional[str] = None
        # Parsed field map as list of (formField, datasetField, updateLogicAction).
        self.field_map: Optional[list[FieldMapEntry]] = None
        self.field_map_error: Optional[str] = None

    @property
    def is_long_format(self) -> bool:
        try:
            return self.link_format is not None and int(self.link_format) == DATALINK_FORMAT_LONG
        except (TypeError, ValueError):
            return False


class Dataset:
    def __init__(self) -> None:
        self.definition_children: list[str] = []
        self.id: Optional[str] = None
        self.title: Optional[str] = None
        self.dataset_type: Optional[str] = None
        self.field_names_raw: Optional[str] = None
        self.field_names: list[str] = []
        self.discriminator: Optional[str] = None
        self.unique_record_field: Optional[str] = None
        self.allow_offline_updates: Optional[str] = None
        self.form_links: list[str] = []
        self.data_links: list[DataLink] = []
        self.case_mgmt: Optional[dict] = None
        self.id_format: Optional[dict] = None
        self.root_children: list[str] = []
        self.instance_present: bool = False
        self.instance_version: Optional[str] = None

    @property
    def has_long_format_link(self) -> bool:
        return any(dl.is_long_format for dl in self.data_links)


def _localname(tag: str) -> str:
    return tag.rsplit("}", 1)[-1] if "}" in tag else tag


def _text(el) -> str:
    return (el.text or "").strip() if el is not None else ""


def parse_field_map(raw: str) -> list[FieldMapEntry]:
    """Parse a fieldMap into [(formField, datasetField, updateLogicAction)].

    Accepts both the modern array form
    ``[{"formField":..,"datasetField":..,"updateLogicAction":..}]`` and the
    legacy object form ``{"formField":"datasetField"}``. A non-string form or
    dataset field is coerced to None so downstream string operations are safe; the
    None is then flagged as a shape error.
    """
    def _str_or_none(v: object) -> Optional[str]:
        return v if isinstance(v, str) else None

    data = json.loads(raw)
    out: list[FieldMapEntry] = []
    if isinstance(data, dict):
        for k, v in data.items():
            ds = v if isinstance(v, str) else (v.get("datasetField") if isinstance(v, dict) else None)
            action = v.get("updateLogicAction") if isinstance(v, dict) else None
            out.append((_str_or_none(k), _str_or_none(ds), str(action) if action else "REPLACE"))
    elif isinstance(data, list):
        for entry in data:
            if not isinstance(entry, dict):
                raise ValueError("field map array entries must be JSON objects")
            action = entry.get("updateLogicAction")
            out.append((_str_or_none(entry.get("formField")),
                        _str_or_none(entry.get("datasetField")),
                        str(action) if action else "REPLACE"))
    else:
        raise ValueError("field map must be a JSON array or object")
    return out


def _guard_doctype(data: bytes) -> None:
    # Defense against XXE / billion-laughs entity expansion. A DTD can only define
    # entities in the prolog, but rather than reason about where the prolog ends
    # (comments, processing instructions, and CDATA can each embed a literal
    # "<dataset" that confuses a prolog scanner, and each became a bypass), reject
    # any DOCTYPE anywhere in the file. A SurveyCTO dataset definition never
    # legitimately contains the literal string "<!doctype", so a whole-file scan
    # is bypass-proof and does not false-reject real definitions.
    if b"<!doctype" in data.lower():
        raise ValueError("dataset definition must not contain a DOCTYPE declaration")


def parse_dataset(xml_path: str) -> Dataset:
    import xml.etree.ElementTree as ET

    raw = Path(xml_path).read_bytes()
    _guard_doctype(raw)
    root = ET.fromstring(raw)
    if _localname(root.tag) != "dataset":
        raise ValueError(f"root element is <{_localname(root.tag)}>, expected <dataset>")

    ds = Dataset()
    definition = None
    for child in root:
        name = _localname(child.tag)
        ds.root_children.append(name)
        if name == "definition" and definition is None:
            definition = child
        elif name == "instance":
            ds.instance_present = True
            ds.instance_version = None
            for ic in child:
                if _localname(ic.tag) == "version":
                    ds.instance_version = _text(ic)
    if definition is None:
        raise ValueError("missing <definition> element")

    for child in definition:
        name = _localname(child.tag)
        ds.definition_children.append(name)
        if name == "id":
            ds.id = _text(child)
        elif name == "title":
            ds.title = _text(child)
        elif name == "datasetType":
            ds.dataset_type = _text(child)
        elif name == "fieldNames":
            ds.field_names_raw = _text(child)
            ds.field_names = [c.strip() for c in ds.field_names_raw.split(",") if c.strip()] if ds.field_names_raw else []
        elif name == "discriminator":
            ds.discriminator = _text(child)
        elif name == "uniqueRecordField":
            ds.unique_record_field = _text(child)
        elif name == "allowOfflineUpdates":
            ds.allow_offline_updates = _text(child)
        elif name == "formLinks":
            for fl in child:
                for fid in fl:
                    if _localname(fid.tag) == "formId":
                        ds.form_links.append(_text(fid))
        elif name == "dataLinks":
            idx = 0
            for dl_el in child:
                if _localname(dl_el.tag) != "dataLink":
                    continue
                ds.data_links.append(_parse_data_link(dl_el, idx))
                idx += 1
        elif name == "caseManagementOptions":
            ds.case_mgmt = _parse_block(child)
        elif name == "idFormatOptions":
            ds.id_format = _parse_block(child)
    return ds


def _parse_data_link(dl_el, idx: int) -> DataLink:
    dl = DataLink(idx)
    for c in dl_el:
        name = _localname(c.tag)
        dl.children.append(name)
        if name == "dataLinkClass":
            dl.link_class = _text(c)
        elif name == "dataLinkType":
            dl.link_type = _text(c)
        elif name == "dataLinkState":
            dl.link_state = _text(c)
        elif name == "dataLinkFormat":
            dl.link_format = _text(c)
        elif name == "linkObjectId":
            dl.link_object_id = _text(c)
        elif name == "fieldMap":
            dl.field_map_raw = _text(c)
        elif name == "joiningField":
            dl.joining_field = _text(c)
        elif name == "relevanceField":
            dl.relevance_field = _text(c)
        elif name == "isAutoConfigured":
            dl.is_auto_configured_raw = _text(c)
        elif name == "publishPartialData":
            dl.publish_partial_data_raw = _text(c)
    if dl.field_map_raw:
        try:
            dl.field_map = parse_field_map(dl.field_map_raw)
        except Exception as exc:  # noqa: BLE001 - reported as a finding
            dl.field_map_error = str(exc)
    return dl


def _parse_block(el) -> dict:
    block: dict = {"_children": []}
    for c in el:
        name = _localname(c.tag)
        block["_children"].append(name)
        if name == "showColumnsWhenTable":
            cols = [_text(cn) for cn in c if _localname(cn.tag) == "columnNames"]
            block[name] = cols
        else:
            block[name] = _text(c)
    return block


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

_ID_RE = re.compile(r"^[0-9a-z_-]+$")
# Latin alphanumeric, used by the server's otherUserCode check (isLatinAlphanumeric).
_ALNUM_RE = re.compile(r"^[A-Za-z0-9]+$")


def _is_unicode_alnum(value: str) -> bool:
    """Mirror Apache Commons StringUtils.isAlphanumeric: non-empty, and every
    character is a Unicode letter or digit. The server applies this to the
    idFormatOptions prefix and suffix, so an accented prefix like 'Énu' is valid.
    Uses isdecimal (Unicode Nd, matching Java Character.isDigit) rather than
    isdigit, which would also accept superscripts the server rejects.
    """
    return bool(value) and all(c.isalpha() or c.isdecimal() for c in value)


# xs:boolean lexical space: exactly these, case-sensitive. The server validates
# the definition against the XSD first, so any other value (TRUE, True, yes) is
# rejected before the value logic runs.
XSD_BOOLEAN_VALID = {"true", "false", "1", "0"}
XSD_BOOLEAN_TRUE = {"true", "1"}


def _xsd_boolean_is_true(text: Optional[str]) -> bool:
    return (text or "").strip() in XSD_BOOLEAN_TRUE


def _check_sequence(children: list[str], order: list[str], report: Report,
                    rule: str, location: str) -> None:
    """Flag children not in the canonical order, and unknown children."""
    rank = {name: i for i, name in enumerate(order)}
    last_rank = -1
    # The highest-ranked element seen so far; an out-of-order element belongs
    # before this one. Updated only on in-order elements so the message and fix
    # name the right anchor even when several elements are out of order.
    high_name = None
    seen: set = set()
    for name in children:
        if name not in rank:
            report.error(rule, f"Unexpected element <{name}> in {location}; "
                         f"the schema does not allow it here.", location)
            continue
        # Every element in these sequences is maxOccurs=1, so a repeat is rejected.
        if name in seen:
            report.error(rule, f"<{name}> appears more than once in {location}; "
                         "the schema allows it at most once.", location)
            continue
        seen.add(name)
        if rank[name] < last_rank:
            report.error(
                rule,
                f"<{name}> appears after <{high_name}> in {location}, but the "
                f"schema requires this order: {', '.join(order)}.",
                location,
                fix=f"Move <{name}> before <{high_name}>.",
            )
        else:
            last_rank = rank[name]
            high_name = name


def _validate_block_children(children: list, allowed: set, label: str,
                             loc: str, report: Report) -> None:
    """An xs:all option block (idFormatOptions, caseManagementOptions) accepts only
    a fixed set of children, each at most once; anything else is a schema error."""
    reported: set = set()
    for name in children:
        if name in reported:
            continue
        if name not in allowed:
            report.error("block-unexpected-child",
                         f"<{name}> is not a valid child of <{label}>; the schema does not "
                         "allow it here.", loc)
        elif children.count(name) > 1:
            report.error("block-duplicate-child",
                         f"<{name}> appears more than once in <{label}>; the schema allows it "
                         "at most once.", loc)
        reported.add(name)


def validate_dataset(ds: Dataset, forms: dict, report: Report) -> None:
    _validate_structure(ds, report)
    _validate_identity(ds, report)
    _validate_type_and_discriminator(ds, report)
    _validate_xsd_booleans(ds, report)
    _validate_field_names(ds, report)
    _validate_id_format(ds, report)
    _validate_case_mgmt(ds, report)
    _validate_standard_columns(ds, report)
    _validate_unique_record_field(ds, report)
    _validate_data_links(ds, forms, report)
    _verify_offline_only(ds, report)


def _validate_structure(ds: Dataset, report: Report) -> None:
    _check_sequence(ds.definition_children, DEFINITION_ORDER, report,
                    "definition-order", "<definition>")
    for req in DEFINITION_REQUIRED:
        if req not in ds.definition_children:
            report.error("definition-required",
                         f"<definition> is missing the required <{req}> element.",
                         "<definition>")
    # The import path dereferences getFormLinks()/getDataLinks() unconditionally,
    # so omitting either element makes the server fail the upload with an NPE,
    # even though the schema marks them optional. Always include them, empty if
    # unused.
    for req in ("formLinks", "dataLinks"):
        if req not in ds.definition_children:
            report.error("definition-formlinks-required",
                         f"Omitting <{req}> makes the server fail the import with an internal "
                         f"error. Include an empty <{req}/> even when it is unused.",
                         "<definition>", fix=f"Add <{req}/>.")
    # The <dataset> root allows only <definition> and <instance>; <instance>, when
    # present, requires a non-empty <version>.
    for name in ds.root_children:
        if name not in ("definition", "instance"):
            report.error("root-unexpected-element",
                         f"Unexpected element <{name}> under <dataset>; only <definition> and "
                         "<instance> are allowed.", "<dataset>")
    if ds.instance_present and not (ds.instance_version or "").strip():
        report.error("instance-version-required",
                     "<instance> must contain a non-empty <version>.", "<instance>")


def _validate_xsd_booleans(ds: Dataset, report: Report) -> None:
    """Boolean-typed elements must use the xs:boolean lexical space (true/false/
    1/0, case-sensitive); the server rejects anything else (TRUE, yes) at XSD
    validation, before any value logic."""
    checks = [("allowOfflineUpdates", ds.allow_offline_updates, "definition/allowOfflineUpdates")]
    if ds.case_mgmt is not None and "showFinalizedSentWhenTree" in ds.case_mgmt:
        checks.append(("showFinalizedSentWhenTree", ds.case_mgmt.get("showFinalizedSentWhenTree"),
                       "definition/caseManagementOptions"))
    if ds.id_format is not None and "allowCapitalLetters" in ds.id_format:
        checks.append(("allowCapitalLetters", ds.id_format.get("allowCapitalLetters"),
                       "definition/idFormatOptions"))
    for dl in ds.data_links:
        if dl.is_auto_configured_raw is not None:
            checks.append((f"isAutoConfigured (dataLink[{dl.index}])", dl.is_auto_configured_raw,
                           f"dataLink[{dl.index}]"))
        if dl.publish_partial_data_raw is not None:
            checks.append((f"publishPartialData (dataLink[{dl.index}])", dl.publish_partial_data_raw,
                           f"dataLink[{dl.index}]"))
    for name, value, loc in checks:
        # value is None only when the element is absent; a present-but-empty
        # element ("") is itself an invalid xs:boolean and must be flagged.
        if value is not None and value.strip() not in XSD_BOOLEAN_VALID:
            report.error("xsd-boolean-lexical",
                         f"<{name.split()[0]}> is {value!r}; an xs:boolean must be one of "
                         "true, false, 1, 0 (lowercase). The server rejects other values.", loc)


def _validate_identity(ds: Dataset, report: Report) -> None:
    if not ds.id:
        report.error("id-blank", "Please specify an ID for the dataset.", "definition/id")
    else:
        # The server strips a leading "formid." (implicit-dataset prefix) before
        # validating the rest case-insensitively against [0-9a-z_-].
        to_check = ds.id
        if to_check.lower().startswith("formid."):
            to_check = to_check[len("formid."):]
        if not _ID_RE.match(to_check.lower()):
            report.error(
                "id-chars",
                "The ID can only contain numbers, letters, dashes and underscores. "
                f"Got: {ds.id!r}.",
                "definition/id",
            )
        # The server reaches the _qc check only for SERVER datasets: CLIENT is
        # rejected earlier as unsupported, and REPORT datasets are allowed to end
        # in _qc. So flag it only for SERVER (the only type the agent authors).
        if ds.id.lower().endswith("_qc") and ds.dataset_type == "SERVER":
            report.error("id-qc-suffix",
                         "The ID can not end with '_qc'. Please correct the ID and try again.",
                         "definition/id")
    if not ds.title:
        report.error("title-blank", "No title has been specified for this dataset.",
                     "definition/title")


def _validate_type_and_discriminator(ds: Dataset, report: Report) -> None:
    if ds.dataset_type is None:
        report.error("type-missing", "Please specify a dataset type (<datasetType>).",
                     "definition/datasetType")
    elif ds.dataset_type not in DATASET_TYPES:
        report.error("type-enum",
                     f"<datasetType> is {ds.dataset_type!r}; must be one of "
                     f"{', '.join(sorted(DATASET_TYPES))}.", "definition/datasetType")
    elif ds.dataset_type == "CLIENT":
        report.error("type-client",
                     "I'm sorry but we no longer support desktop datasets. Use SERVER.",
                     "definition/datasetType",
                     fix="Set <datasetType>SERVER</datasetType>.")
    elif ds.dataset_type == "REPORT":
        report.error("type-report",
                     "You cannot create a quality checks report dataset like this. Please use "
                     "the relevant action from inside the dataset action list. Use SERVER.",
                     "definition/datasetType",
                     fix="Set <datasetType>SERVER</datasetType>.")

    if ds.discriminator is not None and ds.discriminator not in DISCRIMINATORS:
        report.error("discriminator-enum",
                     f"<discriminator> is {ds.discriminator!r}; must be one of "
                     f"{', '.join(sorted(DISCRIMINATORS))}.", "definition/discriminator")

    # The server infers the discriminator from the option blocks present and
    # ignores a conflicting <discriminator>. Flag the mismatch so the author is
    # not surprised that, e.g., a definition with <idFormatOptions> is treated as
    # an enumerator dataset regardless of what <discriminator> says.
    eff = effective_discriminator(ds)
    if ds.discriminator and ds.discriminator in DISCRIMINATORS and ds.discriminator != eff:
        if eff == "CASES":
            reason = "<caseManagementOptions> is present, which forces CASES"
        elif eff == "ENUMERATORS":
            reason = "<idFormatOptions> is present, which forces ENUMERATORS"
        else:
            reason = "no option block is present"
        report.warning("discriminator-inferred",
                       f"The declared <discriminator>{ds.discriminator}</discriminator> is "
                       f"overridden: {reason}, so the server treats this as a {eff} dataset.",
                       "definition/discriminator")


def _db_column_key(base: str) -> str:
    # The server safens a field name to a DB column by replacing every character
    # outside [A-Za-z0-9_] with '_' and comparing case-insensitively; two field
    # names that collapse to the same key collide on import.
    return re.sub(r"[^A-Za-z0-9_]", "_", base.strip()).lower()


def _validate_field_names(ds: Dataset, report: Report) -> None:
    # The server CSV-parses <fieldNames> (quotes honored), then rejects any field
    # name containing a comma. Our split is naive, so detect quoted commas here.
    if ds.field_names_raw and '"' in ds.field_names_raw:
        import csv as _csv
        try:
            parsed = next(_csv.reader([ds.field_names_raw]))
            for f in parsed:
                if "," in f:
                    report.error("field-comma",
                                 f"Dataset field names cannot contain commas. Conflicting field: "
                                 f"{f.strip()!r}.", "definition/fieldNames")
        except Exception:  # noqa: BLE001 - best-effort
            pass
    seen_keys: dict[str, str] = {}
    for col in ds.field_names:
        base = col[:-1] if col.endswith("*") else col
        # The import path does not reject the reserved name 'rowId' (that check is
        # not on the import path), so it is a warning; the '*' is part of the name
        # for the exact comparison, so 'rowId*' is not the reserved name.
        if col in RESERVED_FIELD_NAMES:
            report.warning("field-reserved",
                           f"The field name {col!r} is reserved; rename it to avoid problems.",
                           "definition/fieldNames")
        # The length limit applies to the raw token, including any '*' suffix.
        if len(col) > MAX_FIELD_NAME_LENGTH:
            report.error("field-too-long",
                         f"Dataset field names cannot be longer than {MAX_FIELD_NAME_LENGTH} "
                         f"characters. Conflicting field: {col!r}.", "definition/fieldNames")
        key = _db_column_key(base)
        if not key:
            continue
        if key in seen_keys:
            first = seen_keys[key]
            detail = (f"Column {base!r} appears more than once" if first == base
                      else f"Columns {first!r} and {base!r} collapse to the same database column "
                           f"name {key!r}")
            report.error("field-column-conflict",
                         f"{detail} in <fieldNames>; the server rejects the conflict on import.",
                         "definition/fieldNames")
        else:
            seen_keys[key] = base


def effective_discriminator(ds: Dataset) -> str:
    """The discriminator the server infers on import (getDatasetDiscriminator).

    The presence of an option block overrides the declared <discriminator>:
    <caseManagementOptions> forces CASES; otherwise <idFormatOptions> forces
    ENUMERATORS; otherwise the declared <discriminator> (or DATA) is used.
    """
    if ds.case_mgmt is not None or ds.discriminator == "CASES":
        return "CASES"
    if ds.id_format is not None or ds.discriminator == "ENUMERATORS":
        return "ENUMERATORS"
    return "DATA"


def _is_enumerators(ds: Dataset) -> bool:
    return effective_discriminator(ds) == "ENUMERATORS"


def _is_cases(ds: Dataset) -> bool:
    return effective_discriminator(ds) == "CASES"


def _validate_id_format(ds: Dataset, report: Report) -> None:
    fmt = ds.id_format
    if fmt is None:
        # The server defaults idFormatOptions (6 digits, no prefix/suffix) when an
        # enumerator dataset omits the block, so this is a warning, not a rejection.
        if _is_enumerators(ds):
            report.warning("idformat-default-enum",
                           "This enumerator dataset has no <idFormatOptions>; the server will "
                           "default to 6 digits with no prefix or suffix. Add <idFormatOptions> "
                           "to control the generated ID format.", "definition/idFormatOptions")
        return

    children = fmt.get("_children", [])
    # Schema-level checks apply regardless of discriminator: only the known
    # children are allowed (no duplicates), and numberOfDigits is a required
    # integer (a present-but-empty <numberOfDigits/> is an invalid xs:integer).
    _validate_block_children(children, ID_FORMAT_ALLOWED, "idFormatOptions",
                             "definition/idFormatOptions", report)
    digits = fmt.get("numberOfDigits")
    if "numberOfDigits" not in children:
        report.error("idformat-digits-required",
                     "<idFormatOptions> must contain <numberOfDigits>.",
                     "definition/idFormatOptions")
        digits_int = None
    else:
        try:
            digits_int = int(digits if isinstance(digits, str) else "")
        except ValueError:
            digits_int = None
            report.error("idformat-digits-number",
                         f"Number of digits should be a whole number. Got: {digits!r}.",
                         "definition/idFormatOptions")

    # idFormatOptions present normally forces ENUMERATORS; the only way it is not
    # an enumerator dataset is when caseManagementOptions is also present (CASES
    # wins), in which case the server ignores the idFormatOptions value rules.
    if not _is_enumerators(ds):
        report.recommend("idformat-ignored",
                         "<idFormatOptions> is ignored because <caseManagementOptions> makes "
                         "this a cases dataset.", "definition/idFormatOptions")
        return

    prefix = fmt.get("prefix", "") or ""
    suffix = fmt.get("suffix", "") or ""
    if prefix and (len(prefix) > 10 or not _is_unicode_alnum(prefix)):
        report.error("idformat-prefix",
                     "Prefix can contain alphanumeric characters only and shouldn't "
                     f"exceed 10 characters. Got: {prefix!r}.", "definition/idFormatOptions")
    if suffix and (len(suffix) > 10 or not _is_unicode_alnum(suffix)):
        report.error("idformat-suffix",
                     "Suffix can contain alphanumeric characters only and shouldn't "
                     f"exceed 10 characters. Got: {suffix!r}.", "definition/idFormatOptions")
    if digits_int is not None and (digits_int < 4 or digits_int > 8):
        report.error("idformat-digits-range",
                     "Number of digits can't be less than 4 or higher than 8. "
                     f"Got: {digits_int}.", "definition/idFormatOptions")


def _validate_case_mgmt(ds: Dataset, report: Report) -> None:
    cm = ds.case_mgmt
    if cm is None:
        # The server defaults caseManagementOptions (tree display) when a cases
        # dataset omits the block, so this is a warning, not a rejection.
        if _is_cases(ds):
            report.warning("casemgmt-default",
                           "This cases dataset has no <caseManagementOptions>; the server will "
                           "default to tree display. Add <caseManagementOptions> to control the "
                           "display mode and entry mode.", "definition/caseManagementOptions")
        return

    children = cm.get("_children", [])
    _validate_block_children(children, CASE_MGMT_ALLOWED, "caseManagementOptions",
                             "definition/caseManagementOptions", report)
    for req in CASE_MGMT_REQUIRED:
        if req not in children:
            report.error("casemgmt-required-child",
                         f"<caseManagementOptions> is missing the required <{req}> element.",
                         "definition/caseManagementOptions")

    display = cm.get("displayMode", "")
    if not display:
        report.error("casemgmt-displaymode-blank",
                     "Please select whether this cases dataset should be rendered as a "
                     "tree or as a table (<displayMode>).", "definition/caseManagementOptions")
    elif display == "tree":
        pass
    elif display == "table":
        cols = cm.get("showColumnsWhenTable")
        if not cols:
            report.error("casemgmt-table-empty",
                         "Please select which columns to display when rendered as a table "
                         "(<showColumnsWhenTable> with <columnNames> children).",
                         "definition/caseManagementOptions")
        elif "id" not in cols:
            report.error("casemgmt-table-id",
                         "The 'id' column should be included in the list of columns to "
                         "display (<showColumnsWhenTable>).", "definition/caseManagementOptions")
    else:
        report.error("casemgmt-displaymode-enum",
                     f"Unknown UI option for cases dataset: {display!r}. Use 'tree' or 'table'.",
                     "definition/caseManagementOptions")

    entry = cm.get("entryMode")
    if entry and entry not in ENTRY_MODES:
        report.error("casemgmt-entrymode-enum",
                     f"<entryMode> is {entry!r}; must be one of {', '.join(sorted(ENTRY_MODES))}.",
                     "definition/caseManagementOptions")

    other = cm.get("otherUserCode")
    if other and not _ALNUM_RE.match(other):
        report.error("casemgmt-otherusercode",
                     "'Other user code' field should be latin alphanumeric (letters and "
                     f"digits only). Got: {other!r}.", "definition/caseManagementOptions")

    if cm.get("enumeratorDatasetId"):
        report.cannot_verify("casemgmt-enum-dataset",
                             f"Confirm the linked enumerator dataset {cm.get('enumeratorDatasetId')!r} "
                             "exists and is accessible.", "definition/caseManagementOptions")


def _base_columns(ds: Dataset) -> set[str]:
    return {c[:-1] if c.endswith("*") else c for c in ds.field_names}


def _validate_standard_columns(ds: Dataset, report: Report) -> None:
    if _is_enumerators(ds):
        cols = _base_columns(ds)
        for req in ENUM_REQUIRED_COLUMNS:
            if req not in cols:
                report.warning("enum-required-column",
                               f"Enumerator datasets need the {req!r} column; the upload "
                               "succeeds, but rows without it are rejected when enumerator data "
                               "is inserted.", "definition/fieldNames")
        if "users" not in cols:
            report.warning("enum-users-column",
                           "Enumerator datasets normally include a 'users' column. Without it "
                           "the dataset silently loses per-user enumerator filtering, "
                           "auto-selection, and the manager-code prompt.",
                           "definition/fieldNames",
                           fix="Add 'users' after 'name': id,name,users,...")
    if _is_cases(ds):
        cols = _base_columns(ds)
        for req in CASES_REQUIRED_COLUMNS:
            if req not in cols:
                report.warning("cases-required-column",
                               f"Cases datasets need the {req!r} column; the upload succeeds, but "
                               "the case list fails to render in Collect without it.",
                               "definition/fieldNames")
        for conv in ("users", "roles", "sortby", "enumerators"):
            if conv not in cols:
                report.recommend("cases-standard-column",
                                 f"Cases datasets normally include {conv!r}. Include the full "
                                 f"standard set ({','.join(CASES_STANDARD_COLUMNS)}) to match "
                                 "the console and keep filtering available.",
                                 "definition/fieldNames")


def _looks_like_cases(ds: Dataset) -> bool:
    # Mirror DatasetUtils.looksLikeCasesDataset: a cases discriminator, the literal
    # id "cases", or a field set of >=6 columns covering the standard cases
    # signature. The server uses this (not just the inferred discriminator) when
    # deciding to force uniqueRecordField to 'id'.
    if _is_cases(ds) or (ds.id or "") == "cases":
        return True
    cols = _base_columns(ds)
    signature = {"id", "label", "formids", "users", "roles", "sortby"}
    return len(ds.field_names) >= 6 and signature.issubset(cols)


def _urf_forced_to_id(ds: Dataset) -> bool:
    # The server forces uniqueRecordField to 'id' for enumerator and cases-like
    # datasets before validating the supplied value.
    return _is_enumerators(ds) or _looks_like_cases(ds)


def _validate_unique_record_field(ds: Dataset, report: Report) -> None:
    urf = ds.unique_record_field
    forced_id = _urf_forced_to_id(ds)
    if not urf:
        if forced_id:
            report.warning("urf-missing",
                           "Cases and enumerator datasets use 'id' as the unique record field. "
                           "Add <uniqueRecordField>id</uniqueRecordField>.",
                           "definition/uniqueRecordField")
        return
    if forced_id:
        # The server overwrites whatever is supplied with 'id'; it does not check
        # the supplied value against the field list, so do not require it there.
        if urf != "id":
            report.warning("urf-not-id",
                           "Cases and enumerator datasets force the unique record field to "
                           f"'id'; the supplied {urf!r} is ignored.", "definition/uniqueRecordField")
        return
    # For every other (DATA) dataset, a non-blank uniqueRecordField must be one of
    # the columns in <fieldNames>, or the server rejects the upload. This applies
    # to long-format datasets too: the unique record field is the dataset COLUMN
    # the joining field maps into, not the bare form field. (No long-format
    # exemption exists in the server.)
    if urf not in _base_columns(ds):
        if not ds.field_names:
            detail = "but <fieldNames> is empty, so no column matches"
        else:
            detail = "but it is not one of the columns in <fieldNames>"
        report.error("urf-not-a-column",
                     f'Sorry, the field "{urf}" doesn\'t exist in the dataset: the unique record '
                     f"field must be an existing dataset column, {detail}. For long format, use "
                     "the dataset column the joining field maps into (not the bare form field).",
                     "definition/uniqueRecordField")


def _validate_data_links(ds: Dataset, forms: dict, report: Report) -> None:
    incoming_forms = [dl.link_object_id for dl in ds.data_links
                      if dl.link_class == "FORM" and dl.link_type == "INCOMING"
                      and dl.link_object_id]
    # When the definition references more than one distinct form but only one form
    # file is supplied, the single-form fallback in _resolve_form would match that
    # one file against every link, cross-referencing the wrong fields. Disable the
    # fallback in that ambiguous case.
    allow_single_fallback = len(set(incoming_forms)) <= 1
    for dl in ds.data_links:
        loc = f"dataLink[{dl.index}]"
        _check_sequence(dl.children, DATALINK_ORDER, report, "datalink-order", loc)
        for req in DATALINK_REQUIRED:
            if req not in dl.children:
                report.error("datalink-required",
                             f"{loc} is missing the required <{req}> element.", loc)

        if dl.link_class is not None and dl.link_class not in DATALINK_CLASSES:
            report.error("datalink-class-enum",
                         f"{loc}: <dataLinkClass> is {dl.link_class!r}; must be one of "
                         f"{', '.join(sorted(DATALINK_CLASSES))}.", loc)
        if dl.link_type is not None and dl.link_type not in DATALINK_TYPES:
            report.error("datalink-type-enum",
                         f"{loc}: <dataLinkType> is {dl.link_type!r}; must be INCOMING or OUTGOING.",
                         loc)
        if dl.link_state is not None and dl.link_state not in DATALINK_STATES:
            report.error("datalink-state-enum",
                         f"{loc}: <dataLinkState> is {dl.link_state!r}; must be ENABLED or DISABLED "
                         "(an empty element is rejected too).", loc)
        if dl.link_format is not None:
            # A present-but-empty <dataLinkFormat/> is an invalid xs:integer too.
            if not re.match(r"^[+-]?\d+$", dl.link_format):
                report.error("datalink-format-integer",
                             f"{loc}: <dataLinkFormat> is {dl.link_format!r}; must be an integer "
                             "(0 for wide, 1 for long).", loc)
            elif int(dl.link_format) not in (DATALINK_FORMAT_WIDE, DATALINK_FORMAT_LONG):
                report.recommend("datalink-format-range",
                                 f"{loc}: <dataLinkFormat> is {dl.link_format}; only 0 (wide) and 1 "
                                 "(long) are defined. The server silently treats other values as "
                                 "wide.", loc)

        if dl.field_map_error:
            report.error("fieldmap-json",
                         f"{loc}: the field map JSON is not valid: {dl.field_map_error}.",
                         loc)
            continue

        # Outgoing and cloud links are configured through the console, not a
        # dataset definition import, so the incoming-publishing rules below (field
        # map, joining field, unique record field) do not apply to them.
        if dl.link_type == "OUTGOING" or dl.link_class in ("SPREADSHEET", "FUSION_TABLE"):
            report.warning("outgoing-link-console-only",
                           f"{loc}: outgoing / cloud publishing links (OUTGOING, SPREADSHEET, "
                           "FUSION_TABLE) are set up in the console, not via a dataset definition "
                           "import. This link will not be created by uploading the definition.", loc)
            continue

        form_obj = _resolve_form(dl.link_object_id, forms, allow_single_fallback)

        # Validate the field map even when it is absent: the long-format and
        # joining-field rules must still fire for an incoming link with no <fieldMap>.
        _validate_field_map(ds, dl, report, loc)
        if dl.field_map:
            if form_obj is not None:
                _cross_reference_form(ds, dl, form_obj.fields, report, loc)
            elif forms:
                report.recommend("form-not-supplied",
                                 f"{loc}: no form file was supplied for {dl.link_object_id!r}, so "
                                 "field names in the map were not checked against the form.", loc)

        # An incoming FORM link with no field map publishes nothing; the console
        # rejects it ("Please select at least one field"). It imports but is inert.
        if dl.link_class == "FORM" and not dl.field_map:
            report.error("fieldmap-empty",
                         f"{loc}: the field map is empty, so this link publishes nothing. "
                         "Please select at least one field.", loc)

        # The linkObjectId must be the deployed form's form_id, not its file name.
        if (form_obj is not None and form_obj.form_id and dl.link_object_id
                and form_obj.form_id != dl.link_object_id):
            report.warning("linkobject-formid-mismatch",
                           f"{loc}: <linkObjectId> is {dl.link_object_id!r}, but the supplied "
                           f"form's form_id is {form_obj.form_id!r}. The link must reference the "
                           "deployed form's ID (the form_id in its settings sheet), not the file "
                           "name.", loc, fix=f"Set <linkObjectId>{form_obj.form_id}</linkObjectId>.")

    # The streaming license is a dataset-level capability; the server warns once
    # when a definition with any dataLinks is imported without it.
    if ds.data_links:
        report.cannot_verify("streaming-license",
                             "Confirm the subscription supports publishing into datasets, "
                             "otherwise the import succeeds but the configured publishing will not "
                             "happen.", "definition/dataLinks")

    # A second incoming FORM link to the same form is rejected by the console.
    for f in sorted(name for name, n in Counter(incoming_forms).items() if n > 1):
        report.warning("duplicate-form-link",
                       f"More than one incoming FORM link targets {f!r}. A form can publish to a "
                       "dataset through only one link; the console rejects the duplicate.",
                       "definition/dataLinks")


def _validate_field_map(ds: Dataset, dl: DataLink, report: Report, loc: str) -> None:
    field_map = dl.field_map or []
    form_fields = [t[0] for t in field_map]
    dataset_fields = [t[1] for t in field_map]
    # The server forces the unique record field to 'id' for cases/enumerator
    # datasets, so the joining and mapping checks use that effective value rather
    # than the (possibly absent or ignored) declared <uniqueRecordField>.
    effective_urf = "id" if _urf_forced_to_id(ds) else ds.unique_record_field

    if dl.is_long_format and not dl.joining_field:
        report.error("long-format-requires-joining",
                     f"{loc}: long-format publishing (dataLinkFormat 1) requires a <joiningField> "
                     "from the repeat group to identify unique records.", loc)

    # A long-format incoming FORM link must include the <relevanceField> ELEMENT,
    # even empty. <relevanceField> is a long-standing schema element that the
    # console always writes, so emitting it (empty when unused) imports cleanly on
    # every server. Omitting it is the problem: on older servers (before the
    # FieldInfoUtils null-safety fix in scto-commons 3.0.2 / SCTO-15201) the import
    # succeeds, but validateLongFormatPublishingRequirements then reads the null
    # relevance field and throws a NullPointerException that fails the next upload
    # of the linked form. A present-but-empty <relevanceField></relevanceField>
    # deserializes to "" and is safe everywhere. Wide-format links never reach that
    # validation, so this applies to long format only.
    if (dl.is_long_format and dl.link_class == "FORM"
            and "relevanceField" not in dl.children):
        report.error("long-format-relevance-required",
                     f"{loc}: this long-format link has no <relevanceField> element. The console "
                     "always writes it (empty when unused); include it so the link imports cleanly "
                     "on every server. Omitting it makes older servers fail the next upload of the "
                     "linked form.", loc,
                     fix="Add <relevanceField></relevanceField> after <joiningField> (leave it empty when unused).")

    if any(f is None for f in form_fields) or any(d is None for d in dataset_fields):
        report.error("fieldmap-shape",
                     f"{loc}: every field map entry needs a string form field and a string "
                     "dataset field.", loc)

    # SCTO-15074: a form field or dataset field must not be mapped twice.
    _flag_duplicates(form_fields, report, loc, "form field")
    _flag_duplicates(dataset_fields, report, loc, "dataset field")

    # The dataset column a field publishes into is also subject to the 60-char
    # limit; the server rejects an over-long destination on import.
    for df in dataset_fields:
        if df is None:
            continue
        df_base = df[:-1] if df.endswith("*") else df
        if len(df_base) > MAX_FIELD_NAME_LENGTH:
            report.error("fieldmap-dataset-field-too-long",
                         f"{loc}: dataset field names cannot be longer than "
                         f"{MAX_FIELD_NAME_LENGTH} characters. Conflicting field: {df_base!r}.", loc)

    # Repeat-suffix ('*') rules between the form field, the dataset field, and the
    # publishing format. The '*' marks a field inside a repeat group.
    #   - Long format writes one dataset row per repeat instance into a single
    #     column, so the datasetField carries NO '*': the console omits it and the
    #     publish path strips it (DataUtilsLongFormat.removeWildcards). A '*' here is
    #     non-canonical.
    #   - Wide format expands a repeated field into numbered dataset columns, which
    #     the server does only when the datasetField carries a matching '*'
    #     (DataUtilsWideFormat expands only when both sides are repeated); a repeated
    #     formField mapped to a datasetField without '*' publishes nothing.
    for ff, df, _a in field_map:
        if ff is None or df is None:
            continue
        ff_star = ff.endswith("*")
        df_star = df.endswith("*")
        if dl.is_long_format and df_star:
            report.warning("fieldmap-long-dataset-suffix-extra",
                           f"{loc}: the dataset field {df!r} has a '*' suffix, but a long-format "
                           "link takes no '*' on the dataset field (one row per repeat instance, a "
                           f"single column). Keep the '*' on the form field {ff!r} and remove it "
                           "from the dataset field to match what the console produces.", loc)
        elif not dl.is_long_format and ff_star and not df_star:
            report.warning("fieldmap-wide-dataset-suffix-missing",
                           f"{loc}: {ff!r} is a repeated field but its dataset field {df!r} has no "
                           "'*'. In wide format the '*' must be on both the form field and the "
                           "dataset field to expand the repeat into numbered columns; without it "
                           "this field publishes nothing.", loc)

    # updateLogicAction enum.
    for ff, _df, action in field_map:
        if action not in UPDATE_LOGIC_ACTIONS:
            report.error("fieldmap-action-enum",
                         f"{loc}: updateLogicAction {action!r} for {ff!r} is not valid; use one "
                         f"of {', '.join(sorted(UPDATE_LOGIC_ACTIONS))}.", loc)

    joining = dl.joining_field
    if joining:
        # SCTO-15073: the joining field must be present in the field map.
        if joining not in form_fields:
            base_match = any(
                (ff or "").rstrip("*") == joining.rstrip("*") for ff in form_fields
            )
            if base_match:
                report.error("joining-field-suffix",
                             f"{loc}: the joining field {joining!r} is not in the field map; a "
                             "field with the same base name is, but the repeat '*' suffix differs. "
                             "Match the suffix on both.", loc)
            else:
                report.error("joining-field-in-map",
                             f"{loc}: at least one joining field is not included in the field map "
                             f"({joining}).", loc,
                             fix=f"Add a field map entry whose formField is {joining!r}.")
        else:
            # Joining field's update logic must be REPLACE (console publishing rule).
            for ff, _df, action in field_map:
                if ff == joining and action != "REPLACE":
                    report.error("joining-field-replace",
                                 f"{loc}: the joining field {joining!r} must use the default "
                                 "REPLACE update option; the console rejects the publishing "
                                 "configuration otherwise.", loc)

        # The joining field must merge on the dataset's unique record column. The
        # interactive console enforces this for both formats (the field map stores
        # the column the joining field maps into).
        if effective_urf:
            mapped = [(_df or "").rstrip("*") for ff, _df, _a in field_map if ff == joining]
            if mapped and mapped[0] != effective_urf:
                report.error("joining-merges-on-urf",
                             f"{loc}: the joining field maps to {mapped[0]!r}, but the dataset's "
                             f"unique record field is {effective_urf!r}. Please merge on the dataset "
                             f"field {effective_urf}, so that the appropriate row can be updated.", loc)

    # The unique record field must be mapped by some entry (console publishing rule).
    # An entirely empty field map is already reported by fieldmap-empty, so only
    # flag a non-empty map that omits the unique record column.
    if effective_urf and dl.link_type == "INCOMING" and field_map:
        if not any((df or "").rstrip("*") == effective_urf for df in dataset_fields):
            report.error("urf-not-mapped",
                         f"{loc}: because the {effective_urf} field is the unique ID field for this "
                         "dataset, you must publish a form field into that dataset field and select "
                         "it as the form field to identify unique records.", loc)


def _flag_duplicates(values: list, report: Report, loc: str, label: str) -> None:
    seen: set = set()
    dupes: set = set()
    for v in values:
        if v is None:
            continue
        if v in seen:
            dupes.add(v)
        seen.add(v)
    for v in sorted(dupes):
        report.error("fieldmap-duplicate",
                     f"{loc}: the {label} {v!r} is mapped more than once. You can't map the same "
                     "field twice; the publishing configuration is rejected when it is saved or "
                     "edited in the console.", loc)


def _resolve_form(link_object_id: Optional[str], forms: dict,
                  allow_single_fallback: bool = True) -> Optional["FormInfo"]:
    if not link_object_id or not forms:
        return None
    # Prefer matching the form's declared form_id (the deployed link target).
    for fi in forms.values():
        if fi.form_id and fi.form_id == link_object_id:
            return fi
    # Then the file stem (common when the file is named after the form id).
    if link_object_id in forms:
        return forms[link_object_id]
    # Finally, a single supplied form matches regardless of id, to ease the common
    # one-form case where the file name differs from the form id. Skipped when the
    # definition references several distinct forms (the match would be ambiguous).
    if allow_single_fallback and len(forms) == 1:
        return next(iter(forms.values()))
    return None


def _base_field_name(name: str) -> str:
    # Strip a trailing repeat wildcard and any namespace prefix (the server's
    # findElementByName ignores a "prefix:" before the name).
    base = name[:-1] if name.endswith("*") else name
    if ":" in base:
        base = base.split(":", 1)[1]
    return base


def _cross_reference_form(ds: Dataset, dl: DataLink, form_fields: list,
                          report: Report, loc: str) -> None:
    field_map = dl.field_map or []
    by_name = {f.name: f for f in form_fields}

    joining = dl.joining_field
    jbase = _base_field_name(joining) if joining else None
    jfield = by_name.get(jbase) if jbase else None
    # Repeat groups enclosing the joining field. In long format, every published
    # field must sit in this same set of repeats (or fewer): a field inside a
    # repeat that does not also enclose the joining field does not qualify.
    j_repeats = set(jfield.repeat_path) if jfield else set()
    mapped_form_bases = {_base_field_name(ff) for ff, _df, _a in field_map if ff}

    for ff, _df, _a in field_map:
        if ff is None:
            continue
        base = _base_field_name(ff)
        field = by_name.get(base)
        if field is None:
            if base in CONDITIONAL_META_FIELD_NAMES:
                report.warning("fieldmap-conditional-meta",
                               f"{loc}: {base!r} is a metadata field that is only available in "
                               "certain publishing contexts or when the form declares it; it is "
                               "not in this form. Confirm it is published before relying on it.",
                               loc)
            else:
                report.error("fieldmap-form-field-missing",
                             f"{loc}: the form field {base!r} in the field map does not exist in "
                             "the form.", loc)
            continue
        if field.type == "note":
            report.warning("fieldmap-note",
                           f"{loc}: {base!r} is a note in the form; notes hold no data, so mapping "
                           "one publishes nothing.", loc)
            continue
        if field.repeated and not ff.endswith("*"):
            report.error("fieldmap-repeat-suffix-missing",
                         f"{loc}: {base!r} is inside a repeat group in the form, so its form field "
                         "must carry a '*' suffix (in wide format the dataset field carries a "
                         "matching '*' too; in long format the dataset field has no '*').", loc)
        if not field.repeated and ff.endswith("*"):
            if dl.is_long_format:
                # The server strips wildcards in long format, so this is not a
                # rejection, but the '*' is misleading on a non-repeated field.
                report.warning("fieldmap-repeat-suffix-extra",
                               f"{loc}: {ff!r} has a '*' suffix but {base!r} is not inside a "
                               "repeat group; remove the suffix to avoid confusion.", loc)
            else:
                report.error("fieldmap-repeat-suffix-extra",
                             f"{loc}: {ff!r} has a '*' suffix but {base!r} is not inside a repeat "
                             "group in the form.", loc)
        # Long-format scope: a published field in a sibling repeat does not qualify.
        if dl.is_long_format and jfield is not None and not field.metadata:
            if not set(field.repeat_path).issubset(j_repeats):
                report.error("long-format-field-scope",
                             f"{loc}: in long format, every published field must be in the same "
                             f"repeat instance as the joining field {jbase!r}, in a parent group, "
                             f"or outside all groups. {base!r} is in a different repeat group and "
                             "does not qualify.", loc)

    if joining:
        if jfield is None:
            report.error("joining-field-form-missing",
                         f"{loc}: the joining field {jbase!r} is not part of the form definition.",
                         loc)
        elif dl.is_long_format and not jfield.repeated:
            report.error("joining-field-not-repeat",
                         f"{loc}: long-format publishing requires the joining field {jbase!r} to "
                         "be inside a repeat group in the form.", loc)

    if dl.relevance_field:
        rbase = _base_field_name(dl.relevance_field)
        rfield = by_name.get(rbase) if rbase else None
        if rbase and rfield is None:
            report.warning("relevance-field-missing",
                           f"{loc}: the relevance field {rbase!r} is not part of the form "
                           "definition.", loc)
        # The server scope-checks the relevance field only when it is also a
        # published (mapped) field.
        elif (dl.is_long_format and jfield is not None and rfield is not None
              and rbase in mapped_form_bases
              and not set(rfield.repeat_path).issubset(j_repeats)):
            report.error("long-format-relevance-scope",
                         f"{loc}: in long format, the relevance field {rbase!r} must be in the "
                         f"same repeat instance as the joining field {jbase!r}, in a parent group, "
                         "or outside all groups.", loc)


def _verify_offline_only(ds: Dataset, report: Report) -> None:
    # One deduplicated reminder for every form the definition references: the
    # forms in <formLinks> and the incoming FORM dataLink targets must all be
    # deployed, and each linkObjectId must match the deployed form's form_id.
    referenced_forms = list(dict.fromkeys(
        ds.form_links + [dl.link_object_id for dl in ds.data_links
                         if dl.link_class == "FORM" and dl.link_type == "INCOMING" and dl.link_object_id]))
    if referenced_forms:
        report.cannot_verify("forms-deployed",
                             "Deploy the referenced forms before uploading; each must already exist "
                             "on the server and its form_id must match the reference: "
                             + ", ".join(referenced_forms) + ".", "definition")
    if ds.id:
        report.cannot_verify("id-collision",
                             f"Confirm the dataset id {ds.id!r} is not already used by another "
                             "dataset or reserved by a form.", "definition/id")
    if _xsd_boolean_is_true(ds.allow_offline_updates):
        # The server resolves the unique record field to 'id' for cases/enumerator
        # datasets before this check, so it only rejects DATA datasets that enable
        # offline updates without a unique record field.
        if not ds.unique_record_field and not _urf_forced_to_id(ds):
            report.error("offline-requires-urf",
                         "You can't enable offline updates for a dataset without a unique record "
                         "field. The server rejects this on upload.",
                         "definition/allowOfflineUpdates",
                         fix="Add <uniqueRecordField> or set <allowOfflineUpdates>false</allowOfflineUpdates>.")
        report.cannot_verify("offline-license",
                             "Confirm the subscription supports offline updates.",
                             "definition/allowOfflineUpdates")


# ---------------------------------------------------------------------------
# CLI / output
# ---------------------------------------------------------------------------

def _load_forms(form_paths: list[str], report: Report) -> dict:
    forms: dict = {}
    for p in form_paths:
        stem = Path(p).stem
        try:
            forms[stem] = load_form(p)
        except Exception as exc:  # noqa: BLE001 - reported as a finding
            report.error("form-parse", f"Could not parse form {p!r}: {exc}", p)
    return forms


def run(xml_path: str, form_paths: list[str]) -> Report:
    report = Report()
    forms = _load_forms(form_paths, report)
    try:
        ds = parse_dataset(xml_path)
    except Exception as exc:  # noqa: BLE001 - top-level parse failure
        report.error("xml-parse", f"Could not parse dataset XML: {exc}", xml_path)
        return report
    validate_dataset(ds, forms, report)
    return report


_SEVERITY_LABEL = {
    ERROR: "ERROR",
    WARNING: "WARNING",
    RECOMMENDATION: "RECOMMENDATION",
    CANNOT_VERIFY: "CANNOT VERIFY OFFLINE",
}


def format_text(report: Report) -> str:
    lines: list[str] = []
    # Errors, warnings, and recommendations are the actionable tiers, shown per
    # finding. The cannot-verify items are server-side preconditions, collected
    # into one deduplicated pre-upload checklist below instead of one tagged line
    # each (which is repetitive and conflates them with the actionable tiers).
    actionable = [f for f in report.findings if f.severity != CANNOT_VERIFY]
    for f in sorted(actionable, key=lambda f: _SEVERITY_ORDER[f.severity]):
        head = f"[{_SEVERITY_LABEL[f.severity]}] {f.rule}"
        if f.location:
            head += f" ({f.location})"
        lines.append(head)
        lines.append(f"    {f.message}")
        if f.fix:
            lines.append(f"    fix: {f.fix}")

    cannot_verify = [f for f in report.findings if f.severity == CANNOT_VERIFY]
    if cannot_verify:
        if lines:
            lines.append("")
        lines.append("Before uploading, confirm on the server (cannot be checked offline):")
        for f in cannot_verify:
            lines.append(f"  - {f.message}")

    counts = report.counts()
    lines.append("")
    lines.append(
        f"Summary: {counts[ERROR]} error(s), {counts[WARNING]} warning(s), "
        f"{counts[RECOMMENDATION]} recommendation(s), "
        f"{counts[CANNOT_VERIFY]} server-side item(s) to confirm."
    )
    if not report.has_errors:
        if counts[WARNING] or counts[RECOMMENDATION]:
            lines.append("No blocking errors found. Review the warnings and recommendations above.")
        else:
            lines.append("No blocking errors found.")
    return "\n".join(lines)


def format_json(report: Report) -> str:
    return json.dumps({
        "ok": not report.has_errors,
        "counts": report.counts(),
        "findings": [f.as_dict() for f in report.findings],
    }, indent=2)


def main(argv: Optional[list[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description="Validate a SurveyCTO server dataset definition (.xml) before upload.")
    parser.add_argument("dataset", help="Path to the dataset definition .xml file.")
    parser.add_argument("--form", action="append", default=[], metavar="FORM.xlsx",
                        help="Referenced form XLSForm to cross-check the field map against. "
                             "Repeatable.")
    parser.add_argument("--json", action="store_true", help="Emit findings as JSON.")
    args = parser.parse_args(argv)

    report = run(args.dataset, args.form)
    print(format_json(report) if args.json else format_text(report))
    return 1 if report.has_errors else 0


if __name__ == "__main__":
    sys.exit(main())

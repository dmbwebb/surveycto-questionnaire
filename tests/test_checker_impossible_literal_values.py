"""Tests for SurveyCTOChecker.check_impossible_literal_values.

Builds a minimal in-memory XLSForm, writes it to a tempfile, runs the new check,
and asserts the expected errors fire (and only those).
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pytest


SCRIPTS = Path(__file__).resolve().parents[1] / "scripts"
sys.path.insert(0, str(SCRIPTS))

from surveycto_checker import SurveyCTOChecker  # noqa: E402


def _build_form(tmp_path: Path) -> Path:
    """Write a minimal XLSForm with a controlled set of impossible-literal cases."""
    wb = openpyxl.Workbook()

    # --- survey sheet ---
    survey = wb.active
    survey.title = "survey"
    survey.append([
        "type", "name", "label", "relevance", "constraint", "calculation",
        "disabled",
    ])
    # Define a select_one bound to choice list 'mylist' (1,2,3,-997).
    survey.append(["select_one mylist", "foo", "Foo?", "", "", "", ""])
    # BAD: equality against literal '997' — choice list has -997, not 997.
    survey.append(["text", "bad_eq", "Bad eq", "${foo} = 997", "", "", ""])
    # GOOD: equality against the actual literal '-997'.
    survey.append(["text", "good_eq", "Good eq", "${foo} = -997", "", "", ""])
    # GOOD: selected() with quoted '-997'.
    survey.append(["text", "good_sel", "Good sel",
                   "selected(${foo}, '-997')", "", "", ""])
    # SKIP: RHS is another ${var} — not a literal.
    survey.append(["text", "skip_var_rhs", "Skip var RHS",
                   "${foo} = ${some_other_var}", "", "", ""])
    # SKIP: variable bound to dynamic list (select_one_from_file).
    survey.append(["select_one_from_file ext.csv", "dyn", "Dynamic", "", "", "", ""])
    survey.append(["text", "skip_dyn", "Skip dyn", "${dyn} = 'whatever'", "", "", ""])
    # SKIP: disabled row — should not be checked.
    survey.append(["text", "disabled_row", "Disabled", "${foo} = 999", "", "", "yes"])
    # BAD: constraint with selected() referencing impossible value '5'.
    survey.append(["select_multiple mylist", "multi", "Multi", "",
                   "if(selected(., '5'), count-selected(.)=1, count-selected(.)>=1)",
                   "", ""])
    # GOOD: calculation referencing a valid value as quoted string.
    survey.append(["calculate", "good_calc", "", "", "",
                   "if(${foo} = '3', 'yes', 'no')", ""])

    # --- choices sheet ---
    choices = wb.create_sheet("choices")
    choices.append(["list_name", "name", "label"])
    for v in (1, 2, 3, -997):
        choices.append(["mylist", v, f"Option {v}"])

    out = tmp_path / "test_form.xlsx"
    wb.save(out)
    return out


def test_impossible_literal_values_flags_bad_and_skips_others(tmp_path, capsys):
    form_path = _build_form(tmp_path)
    checker = SurveyCTOChecker(form_path)
    assert checker.load_form(), "form should load"

    ok = checker.check_impossible_literal_values()
    assert not ok, "check should fail when impossible literals are present"

    err_text = "\n".join(checker.errors)

    # The 'bad_eq' row (literal 997 vs list {-997,1,2,3}) must be flagged.
    assert "bad_eq" in err_text, f"expected bad_eq in errors, got:\n{err_text}"
    assert "'997'" in err_text

    # The 'multi' row (selected(., '5') with list {1,2,3,-997}) must be flagged.
    assert "multi" in err_text, f"expected multi in errors, got:\n{err_text}"
    assert "'5'" in err_text

    # None of the "good" or "skip" rows should appear.
    for name in ("good_eq", "good_sel", "skip_var_rhs", "skip_dyn",
                 "disabled_row", "good_calc"):
        assert name not in err_text, \
            f"{name!r} should not appear in errors, got:\n{err_text}"


def test_impossible_literal_values_passes_clean_form(tmp_path):
    wb = openpyxl.Workbook()
    survey = wb.active
    survey.title = "survey"
    survey.append(["type", "name", "label", "relevance"])
    survey.append(["select_one mylist", "foo", "Foo?", ""])
    survey.append(["text", "follow", "Follow", "${foo} = -997"])

    choices = wb.create_sheet("choices")
    choices.append(["list_name", "name", "label"])
    for v in (1, 2, 3, -997):
        choices.append(["mylist", v, f"Option {v}"])

    path = tmp_path / "clean.xlsx"
    wb.save(path)

    checker = SurveyCTOChecker(path)
    assert checker.load_form()
    assert checker.check_impossible_literal_values()
    assert not checker.errors

"""Write-side tests: edits go directly to the gsheet via Sheets API.

Each test uses an *ephemeral* fixture (a fresh Drive copy of the persistent
fixture, trashed at end-of-test) so tests don't interfere with each other.

The write API is exercised end-to-end: write -> re-export to xlsx -> verify
with openpyxl AND run surveycto_checker.py to make sure we didn't break the
form's structural validity.
"""
from __future__ import annotations

import subprocess
import sys
import time
from pathlib import Path

import openpyxl
import pytest

from gsheet_edit import (
    DEFAULT_TRANSLATION_ASSIGNEE_EMAIL,
    DEFAULT_TRANSLATION_COMMENT_TEMPLATE,
    StaleDataError,
    UnsupportedCellCommentsError,
    add_choice_list,
    add_translation_comment,
    append_row,
    batch_update_cells,
    bulk_set_column,
    delete_row,
    find_row_by_value,
    get_cell,
    get_row,
    get_text_color,
    insert_row_at,
    open_tab,
    rename_variable,
    set_text_color,
    update_cell,
    update_cell_checked,
)
from gsheet_io import (
    drive_service,
    exported_xlsx,
    get_drive_modified_time,
    get_drive_version,
)

SCRIPTS = Path(__file__).parent.parent / "scripts"
CHECKER = SCRIPTS / "surveycto_checker.py"


def _run_checker(xlsx_path: Path) -> subprocess.CompletedProcess:
    return subprocess.run(
        [sys.executable, str(CHECKER), str(xlsx_path)],
        capture_output=True, text=True, timeout=60,
    )


# ---------- Cell-level reads/writes ----------

@pytest.mark.live
@pytest.mark.destructive
def test_round_trip_update_cell(ephemeral_fixture):
    doc_id = ephemeral_fixture("school_survey")
    tab = open_tab(doc_id, "survey")

    row = find_row_by_value(tab, "name", "s_l1")
    assert row is not None, "fixture missing expected variable s_l1"

    original = get_cell(tab, row, "label")
    new_label = "TEST_LABEL_FROM_PYTEST"
    update_cell(tab, row, "label", new_label)
    assert get_cell(tab, row, "label") == new_label

    update_cell(tab, row, "label", original)
    assert get_cell(tab, row, "label") == original


@pytest.mark.live
@pytest.mark.destructive
def test_find_row_returns_none_for_missing(ephemeral_fixture):
    doc_id = ephemeral_fixture("bulletin_notes")
    tab = open_tab(doc_id, "survey")
    assert find_row_by_value(tab, "name", "_definitely_not_a_real_field") is None


# ---------- Append + visible in export ----------

@pytest.mark.live
@pytest.mark.destructive
def test_append_row_lands_at_end_and_is_visible_in_export(ephemeral_fixture):
    doc_id = ephemeral_fixture("bulletin_notes")
    tab = open_tab(doc_id, "survey")

    new_row = {
        "type": "note",
        "name": "test_note_appended_by_pytest",
        "label": "Test note from pytest run",
        "required": "no",
    }
    landed_row = append_row(tab, new_row)
    assert landed_row > 1

    with exported_xlsx(doc_id) as path:
        wb = openpyxl.load_workbook(path, data_only=True)
        survey = wb["survey"]
        headers = [c.value for c in survey[1]]
        name_col_idx = headers.index("name") + 1
        names = [survey.cell(r, name_col_idx).value
                 for r in range(2, survey.max_row + 1)]
        assert "test_note_appended_by_pytest" in names


# ---------- Rename a variable + all references ----------

@pytest.mark.live
@pytest.mark.destructive
def test_rename_variable_updates_name_and_all_references(ephemeral_fixture):
    doc_id = ephemeral_fixture("school_survey")
    tab = open_tab(doc_id, "survey")

    # niveau_dispo is referenced from rows 111, 259, 266, 272 in the fixture
    # (per the checker's broken-ref report) — perfect test of cross-row rewrite.
    old_name = "niveau_dispo"
    new_name = "niveau_dispo_RENAMED"

    row = find_row_by_value(tab, "name", old_name)
    assert row is not None, (
        f"fixture school_survey missing expected variable {old_name!r}"
    )

    result = rename_variable(tab, old_name, new_name)
    assert result["name_renamed"] == 1
    assert result["references_updated"] >= 1

    new_row = find_row_by_value(tab, "name", new_name)
    assert new_row is not None
    assert find_row_by_value(tab, "name", old_name) is None

    with exported_xlsx(doc_id) as path:
        wb = openpyxl.load_workbook(path, data_only=True)
        survey = wb["survey"]
        headers = [c.value for c in survey[1]]
        old_ref = "${" + old_name + "}"
        new_ref = "${" + new_name + "}"
        ref_cols_idx = [
            i + 1 for i, h in enumerate(headers)
            if h in {"relevance", "constraint", "calculation", "label",
                     "label:Malagasy", "hint", "hint:Malagasy",
                     "choice_filter", "repeat_count", "required", "default"}
        ]
        leaks = []
        new_hits = 0
        for r in range(2, survey.max_row + 1):
            for c in ref_cols_idx:
                v = survey.cell(r, c).value
                if v and old_ref in str(v):
                    leaks.append((r, headers[c-1], v))
                if v and new_ref in str(v):
                    new_hits += 1
        assert not leaks, f"stale references to ${{{old_name}}} remain: {leaks[:5]}"
        assert new_hits >= 1


# ---------- Foreground color ----------

@pytest.mark.live
@pytest.mark.destructive
def test_set_and_read_back_red_text_color(ephemeral_fixture):
    """The 'red text = unverified translation' semantic requires a working
    foreground color primitive — which the upstream sheets_manager.py lacks."""
    doc_id = ephemeral_fixture("school_survey")
    tab = open_tab(doc_id, "survey")

    row = find_row_by_value(tab, "name", "s_l1")
    assert row is not None

    set_text_color(tab, row, "label:Malagasy", (1.0, 0.0, 0.0))
    color = get_text_color(tab, row, "label:Malagasy")
    assert color is not None
    r, g, b = color
    assert r > 0.9 and g < 0.1 and b < 0.1, f"expected ~red, got {color}"

    set_text_color(tab, row, "label:Malagasy", (0.0, 0.0, 0.0))
    color2 = get_text_color(tab, row, "label:Malagasy")
    if color2 is not None:
        assert all(c < 0.05 for c in color2), f"expected ~black, got {color2}"


# ---------- Cell comments ----------

@pytest.mark.live
@pytest.mark.destructive
def test_translation_comment_fails_loudly_without_creating_substitute(
        ephemeral_fixture):
    """Public Google APIs cannot create UI-backed Sheets cell comments.

    The helper must fail before creating an unanchored Drive comment or a
    Sheets note, since neither is a substitute for a cell comment with an
    @mention notification.
    """
    doc_id = ephemeral_fixture("bulletin_notes")
    tab = open_tab(doc_id, "survey")

    expected_default = (
        f"@{DEFAULT_TRANSLATION_ASSIGNEE_EMAIL} traduction \u00e0 "
        "v\u00e9rifier stp"
    )
    assert (DEFAULT_TRANSLATION_COMMENT_TEMPLATE
            .format(email=DEFAULT_TRANSLATION_ASSIGNEE_EMAIL)
            == expected_default)

    drive = drive_service()
    before = drive.comments().list(
        fileId=doc_id,
        fields="comments(id)",
        pageSize=100,
    ).execute().get("comments", [])

    with pytest.raises(UnsupportedCellCommentsError) as exc:
        add_translation_comment(tab, 2, "label")

    assert "cell comments with @mentions cannot be created" in str(exc.value)

    after = drive.comments().list(
        fileId=doc_id,
        fields="comments(id)",
        pageSize=100,
    ).execute().get("comments", [])
    assert len(after) == len(before)


# ---------- Round-trip preserves checker validity ----------

@pytest.mark.live
@pytest.mark.destructive
def test_round_trip_edit_does_not_introduce_new_checker_errors(ephemeral_fixture):
    doc_id = ephemeral_fixture("bulletin_notes")
    tab = open_tab(doc_id, "survey")

    with exported_xlsx(doc_id) as before:
        baseline = _run_checker(before)
    baseline_errors = baseline.stdout.count("\n❌")

    row = 2
    original = get_cell(tab, row, "label") or ""
    update_cell(tab, row, "label", original + " ")
    update_cell(tab, row, "label", original)

    with exported_xlsx(doc_id) as after:
        after_run = _run_checker(after)
    after_errors = after_run.stdout.count("\n❌")
    assert after_errors == baseline_errors, (
        f"round-trip edit introduced new checker errors. "
        f"baseline={baseline_errors}, after={after_errors}.\n"
        f"AFTER STDOUT TAIL:\n{after_run.stdout[-2000:]}"
    )


# ---------- NEW: Multi-tab edit (choice list + select_one) ----------

@pytest.mark.live
@pytest.mark.destructive
def test_add_choice_list_and_select_one_passes_checker(ephemeral_fixture):
    """Realistic workflow: add a new choice list to 'choices' AND a select_one
    row in 'survey' that references it. Round-trip via xlsx and confirm the
    checker doesn't flag it as a broken-reference."""
    doc_id = ephemeral_fixture("bulletin_notes")

    list_name = "test_pet_kind"
    add_choice_list(doc_id, list_name, [
        ("dog", "Dog"),
        ("cat", "Cat"),
        ("other", "Other"),
    ])

    survey = open_tab(doc_id, "survey")
    append_row(survey, {
        "type": f"select_one {list_name}",
        "name": "test_pet_question",
        "label": "What kind of pet?",
        "required": "no",
    })

    # Verify in export
    with exported_xlsx(doc_id) as path:
        wb = openpyxl.load_workbook(path, data_only=True)

        # Choice list rows are present (XLSForm allows either 'name' or 'value'
        # for the choice id column — bulletin_notes uses 'value').
        choices_sheet = wb["choices"]
        choices_headers = [c.value for c in choices_sheet[1]]
        list_col = choices_headers.index("list_name") + 1
        id_header = "name" if "name" in choices_headers else "value"
        id_col = choices_headers.index(id_header) + 1
        choice_pairs = {(choices_sheet.cell(r, list_col).value,
                         choices_sheet.cell(r, id_col).value)
                        for r in range(2, choices_sheet.max_row + 1)}
        assert (list_name, "dog") in choice_pairs
        assert (list_name, "cat") in choice_pairs
        assert (list_name, "other") in choice_pairs

        # Select_one row is present in survey
        survey_sheet = wb["survey"]
        s_headers = [c.value for c in survey_sheet[1]]
        type_col = s_headers.index("type") + 1
        name_col_s = s_headers.index("name") + 1
        survey_pairs = {(survey_sheet.cell(r, type_col).value,
                         survey_sheet.cell(r, name_col_s).value)
                        for r in range(2, survey_sheet.max_row + 1)}
        assert (f"select_one {list_name}", "test_pet_question") in survey_pairs

        # Checker should NOT flag this select_one as referencing an undefined
        # choice list (which it would if our writes didn't land properly).
        result = _run_checker(path)
        # Look for our specific list in the checker output. If it complained
        # about ours, we'd see the list_name in the "undefined choice list" report.
        undefined_section = "Undefined choice lists" in result.stdout
        if undefined_section:
            # Make sure OUR list isn't the one being flagged.
            assert list_name not in _section_after(
                result.stdout, "Undefined choice lists",
            ), (
                f"checker flagged our new list {list_name!r} as undefined; "
                f"the multi-tab edit didn't land correctly.\n"
                f"STDOUT TAIL:\n{result.stdout[-1500:]}"
            )


def _section_after(text: str, header: str, max_lines: int = 30) -> str:
    """Return the next ``max_lines`` lines after a section header in the
    checker output, for substring matching."""
    if header not in text:
        return ""
    after = text.split(header, 1)[1]
    return "\n".join(after.splitlines()[:max_lines])


# ---------- NEW: Insert row at specific position ----------

@pytest.mark.live
@pytest.mark.destructive
def test_insert_row_at_position_preserves_neighbors(ephemeral_fixture):
    """Inserting a row at position N must shift everything below down by 1
    without dropping or corrupting the rows around it. Critical for placing
    a new question inside a `begin group`/`end group` block."""
    doc_id = ephemeral_fixture("bulletin_notes")
    tab = open_tab(doc_id, "survey")

    # Snapshot rows around the insertion point.
    insert_at = 5
    above = get_row(tab, insert_at - 1)
    at = get_row(tab, insert_at)
    below = get_row(tab, insert_at + 1)

    new_row = {
        "type": "note",
        "name": "test_insert_at_5",
        "label": "Inserted at 5 by pytest",
    }
    insert_row_at(tab, insert_at, new_row)

    # Above row should be unchanged.
    above_after = get_row(tab, insert_at - 1)
    assert above_after.get("name") == above.get("name"), (
        f"row above insertion changed: was {above.get('name')!r}, "
        f"now {above_after.get('name')!r}"
    )

    # Inserted row should be at insert_at.
    at_after = get_row(tab, insert_at)
    assert at_after.get("name") == "test_insert_at_5"

    # Original "at" content is now at insert_at + 1; original "below" at +2.
    bumped = get_row(tab, insert_at + 1)
    assert bumped.get("name") == at.get("name"), (
        f"row at insertion point not shifted: expected {at.get('name')!r}, "
        f"got {bumped.get('name')!r}"
    )
    bumped2 = get_row(tab, insert_at + 2)
    assert bumped2.get("name") == below.get("name"), (
        f"row below insertion not shifted: expected {below.get('name')!r}, "
        f"got {bumped2.get('name')!r}"
    )


@pytest.mark.live
@pytest.mark.destructive
def test_insert_row_within_group_preserves_group_balance(ephemeral_fixture):
    """A more focused version of the above: insert *between* a `begin group`
    and its `end group` and verify the begin/end-group counts are still equal
    and that insertion fell within a real group block."""
    doc_id = ephemeral_fixture("school_survey")
    tab = open_tab(doc_id, "survey")

    # Find the first 'begin group' and pick an insertion row 2 rows past it.
    svc_grid = sheets_service_full_grid(tab)
    begin_idx = next(
        (i for i, r in enumerate(svc_grid[1:], start=2)
         if (r[tab.col_idx_0("type")] if len(r) > tab.col_idx_0("type") else "")
         == "begin group"),
        None,
    )
    assert begin_idx is not None, "fixture has no 'begin group' rows"

    # Target a row inside the first group (begin_idx + 2 should still be inside
    # the group as long as the group has > 2 children).
    insert_at = begin_idx + 2

    insert_row_at(tab, insert_at, {
        "type": "note",
        "name": "test_inserted_inside_group",
        "label": "Inserted inside group by pytest",
    })

    # Re-fetch the survey, count begin/end group rows.
    svc_grid_after = sheets_service_full_grid(tab)
    type_col = tab.col_idx_0("type")
    types_after = [
        (r[type_col] if len(r) > type_col else "")
        for r in svc_grid_after[1:]
    ]
    begin_count = sum(1 for t in types_after if t == "begin group")
    end_count = sum(1 for t in types_after if t == "end group")
    assert begin_count == end_count, (
        f"group balance broken: begin={begin_count} end={end_count}"
    )

    # And the inserted row is present at the expected position.
    name_col = tab.col_idx_0("name")
    assert (svc_grid_after[insert_at - 1][name_col]
            == "test_inserted_inside_group"), (
        f"inserted row not at row {insert_at}; "
        f"got {svc_grid_after[insert_at - 1][name_col]!r}"
    )


def sheets_service_full_grid(tab):
    """Helper: read the entire tab as a list-of-lists grid (header + data)."""
    from gsheet_io import sheets_service
    svc = sheets_service()
    res = svc.spreadsheets().values().get(
        spreadsheetId=tab.doc_id,
        range=f"'{tab.tab_title}'",
    ).execute()
    return res.get("values", [])


# ---------- NEW: Concurrent-edit / stale-data detection ----------

@pytest.mark.live
@pytest.mark.destructive
def test_update_cell_checked_succeeds_when_value_unchanged(ephemeral_fixture):
    doc_id = ephemeral_fixture("bulletin_notes")
    tab = open_tab(doc_id, "survey")

    row = 2
    original = get_cell(tab, row, "label") or ""
    update_cell_checked(tab, row, "label", expected_old=original,
                        new_value="MODIFIED_BY_PYTEST")
    assert get_cell(tab, row, "label") == "MODIFIED_BY_PYTEST"


@pytest.mark.live
@pytest.mark.destructive
def test_update_cell_checked_detects_out_of_band_modification(ephemeral_fixture):
    """Simulate a concurrent edit: another writer changes the cell between our
    read and our write. update_cell_checked must raise StaleDataError."""
    doc_id = ephemeral_fixture("bulletin_notes")
    tab = open_tab(doc_id, "survey")

    row = 2
    # Snapshot what we *think* the cell holds.
    snapshot = get_cell(tab, row, "label") or ""

    # Out-of-band edit (simulating a different writer).
    update_cell(tab, row, "label", "OUT_OF_BAND_EDIT")

    # Now try to write based on the stale snapshot.
    with pytest.raises(StaleDataError) as exc:
        update_cell_checked(tab, row, "label",
                            expected_old=snapshot,
                            new_value="WOULD_BE_OUR_WRITE")
    assert exc.value.actual == "OUT_OF_BAND_EDIT"
    # Cell should still hold the out-of-band edit, not our write.
    assert get_cell(tab, row, "label") == "OUT_OF_BAND_EDIT"


@pytest.mark.live
def test_drive_modified_time_endpoint_is_callable(persistent_fixture):
    """`get_drive_modified_time` returns an RFC3339 timestamp string.

    NOTE: Drive's modifiedTime does NOT propagate immediately after a Sheets
    API write — it can lag 30s+ for newly-copied files. Don't rely on it as
    a tight stale-data sentinel. Use ``update_cell_checked`` (compare-and-
    swap) instead — see that test below.
    """
    doc_id = persistent_fixture("bulletin_notes")
    t = get_drive_modified_time(doc_id)
    assert isinstance(t, str)
    # RFC3339 like "2026-05-01T12:00:00.000Z"
    assert "T" in t and t.endswith("Z"), f"unexpected timestamp shape: {t!r}"


@pytest.mark.live
@pytest.mark.destructive
def test_drive_version_endpoint_is_callable(ephemeral_fixture):
    """`get_drive_version` returns an int and survives a write — but its
    propagation lag makes it unsuitable for tight test loops."""
    doc_id = ephemeral_fixture("bulletin_notes")
    v = get_drive_version(doc_id)
    assert isinstance(v, int)
    assert v >= 1


# ---------- NEW: Delete row ----------

@pytest.mark.live
@pytest.mark.destructive
def test_delete_row_removes_target(ephemeral_fixture):
    doc_id = ephemeral_fixture("bulletin_notes")
    tab = open_tab(doc_id, "survey")

    # Append a marker row, find it, delete it, confirm gone.
    append_row(tab, {
        "type": "note",
        "name": "test_delete_marker",
        "label": "to be deleted",
    })
    row = find_row_by_value(tab, "name", "test_delete_marker")
    assert row is not None
    delete_row(tab, row)
    assert find_row_by_value(tab, "name", "test_delete_marker") is None


@pytest.mark.live
@pytest.mark.destructive
def test_delete_row_rejects_header_row(ephemeral_fixture):
    doc_id = ephemeral_fixture("bulletin_notes")
    tab = open_tab(doc_id, "survey")
    with pytest.raises(ValueError):
        delete_row(tab, 1)


# ---------- NEW: batch writes ----------

@pytest.mark.live
@pytest.mark.destructive
def test_batch_update_cells_multi_column(ephemeral_fixture):
    """batch_update_cells writes to several cells across multiple columns
    in one API call. Verify each landed value via per-cell reads."""
    doc_id = ephemeral_fixture("bulletin_notes")
    tab = open_tab(doc_id, "survey")
    edits = [
        (2, "label", "BATCH_LABEL_2"),
        (3, "label", "BATCH_LABEL_3"),
        (2, "name",  "batch_name_2"),
        (3, "name",  "batch_name_3"),
    ]
    res = batch_update_cells(tab, edits)
    assert res.get("totalUpdatedCells") == 4
    assert get_cell(tab, 2, "label") == "BATCH_LABEL_2"
    assert get_cell(tab, 3, "label") == "BATCH_LABEL_3"
    assert get_cell(tab, 2, "name") == "batch_name_2"
    assert get_cell(tab, 3, "name") == "batch_name_3"


@pytest.mark.live
@pytest.mark.destructive
def test_bulk_set_column_sets_same_value_across_rows(ephemeral_fixture):
    """bulk_set_column flips one column to the same value across many rows
    in one API call — the canonical "disable a module" or "mark required"
    pattern."""
    doc_id = ephemeral_fixture("bulletin_notes")
    tab = open_tab(doc_id, "survey")
    rows = [2, 3, 4, 5]
    res = bulk_set_column(tab, rows, "label", "BULK_LABEL")
    assert res.get("totalUpdatedCells") == 4
    for r in rows:
        assert get_cell(tab, r, "label") == "BULK_LABEL"


def test_batch_update_cells_empty_is_noop():
    """Empty edits list returns 0 counts without making an API call.

    No fixture needed — the empty-input early-return must not require auth
    or a live Sheet. Pass a sentinel object that would crash if touched."""
    class _Sentinel:
        def __getattr__(self, name):
            raise AssertionError(
                f"empty-input path must not touch tab (called .{name})"
            )
    res = batch_update_cells(_Sentinel(), [])
    assert res == {"totalUpdatedCells": 0, "totalUpdatedRows": 0}

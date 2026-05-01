"""Read-side tests: gsheet -> temp xlsx -> existing surveycto tooling.

The user's mandate: "for any kind of read commands or anything that requires
examining the survey CTO, it should download to a local temporary Excel form,
and then examine it using the skill in the same way as before".
"""
from __future__ import annotations

import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

from gsheet_io import (
    exported_xlsx,
    export_gsheet_to_xlsx,
    get_metadata,
    sheet_id_for_tab,
)

SCRIPTS = Path(__file__).parent.parent / "scripts"
CHECKER = SCRIPTS / "surveycto_checker.py"
TO_TXT = SCRIPTS / "surveycto_to_txt.py"


# ---------- Plumbing ----------

@pytest.mark.live
def test_export_smallest_fixture_yields_real_xlsx(persistent_fixture, tmp_path):
    doc_id = persistent_fixture("bulletin_notes")
    dest = tmp_path / "out.xlsx"
    export_gsheet_to_xlsx(doc_id, dest)
    assert dest.is_file()
    assert dest.stat().st_size > 10_000
    wb = openpyxl.load_workbook(dest, data_only=True)
    assert "survey" in wb.sheetnames
    assert "choices" in wb.sheetnames
    assert "settings" in wb.sheetnames


@pytest.mark.live
def test_exported_xlsx_context_manager_cleans_up(persistent_fixture):
    doc_id = persistent_fixture("bulletin_notes")
    captured: Path
    with exported_xlsx(doc_id) as path:
        assert path.is_file()
        assert path.stat().st_size > 10_000
        captured = path
    assert not captured.exists()
    assert not captured.parent.exists()


# ---------- Content fidelity ----------

@pytest.mark.live
def test_school_survey_export_preserves_structure(persistent_fixture):
    doc_id = persistent_fixture("school_survey")
    with exported_xlsx(doc_id) as path:
        wb = openpyxl.load_workbook(path, data_only=True)
        expected = {"survey", "choices", "settings",
                    "help-survey", "help-choices", "help-settings"}
        assert expected.issubset(set(wb.sheetnames)), (
            f"missing tabs. got={wb.sheetnames}, expected superset of={expected}"
        )

        survey = wb["survey"]
        assert 1200 < survey.max_row < 1500, (
            f"survey rows={survey.max_row} outside expected 1200-1500"
        )

        headers = [c.value for c in survey[1]]
        for col in ("type", "name", "label", "label:Malagasy",
                    "a_traduire", "module", "relevance", "calculation"):
            assert col in headers, f"missing header column: {col!r}"


@pytest.mark.live
def test_settings_version_formula_is_evaluated(persistent_fixture):
    """The settings.version cell is a NOW()-based formula. Drive's xlsx export
    must materialise the cached value (otherwise SurveyCTO would reject)."""
    doc_id = persistent_fixture("school_survey")
    with exported_xlsx(doc_id) as path:
        wb = openpyxl.load_workbook(path, data_only=True)
        st = wb["settings"]
        headers = [c.value for c in st[1]]
        version_col = headers.index("version")
        version_value = st.cell(row=2, column=version_col + 1).value
        assert version_value is not None, "version cell empty after export"
        assert not str(version_value).startswith("="), (
            f"formula not evaluated; got literal: {version_value!r}"
        )
        v = str(version_value)
        assert v.isdigit() and len(v) == 10, (
            f"unexpected version format: {v!r}"
        )


# ---------- Existing tooling compatibility ----------

@pytest.mark.live
def test_checker_runs_on_exported_xlsx(persistent_fixture, tmp_path):
    doc_id = persistent_fixture("school_survey")
    with exported_xlsx(doc_id) as path:
        result = subprocess.run(
            [sys.executable, str(CHECKER), str(path)],
            capture_output=True, text=True, timeout=60,
        )
        assert result.returncode in (0, 1), (
            f"checker crashed (rc={result.returncode}):\n"
            f"STDOUT:\n{result.stdout[-2000:]}\n"
            f"STDERR:\n{result.stderr[-2000:]}"
        )
        assert "SurveyCTO Form Checker" in result.stdout
        assert "Checking Required Columns" in result.stdout


@pytest.mark.live
def test_to_txt_runs_on_exported_xlsx(persistent_fixture, tmp_path):
    doc_id = persistent_fixture("bulletin_notes")
    with exported_xlsx(doc_id) as path:
        out_txt = tmp_path / "out.txt"
        result = subprocess.run(
            [sys.executable, str(TO_TXT), str(path), str(out_txt)],
            capture_output=True, text=True, timeout=60,
        )
        assert result.returncode == 0, (
            f"to_txt failed (rc={result.returncode}): {result.stderr}"
        )
        assert out_txt.is_file() and out_txt.stat().st_size > 100
        body = out_txt.read_text()
        assert "##" in body or "•" in body, (
            f"output doesn't look like a survey dump:\n{body[:500]}"
        )


# ---------- Sheets API metadata sanity ----------

@pytest.mark.live
def test_sheet_id_for_tab_lookup(persistent_fixture):
    doc_id = persistent_fixture("bulletin_notes")
    md = get_metadata(doc_id)
    titles = [s["properties"]["title"] for s in md["sheets"]]
    assert "survey" in titles
    survey_id = sheet_id_for_tab(doc_id, "survey")
    assert isinstance(survey_id, int)

    with pytest.raises(ValueError):
        sheet_id_for_tab(doc_id, "_nonexistent_tab_")
